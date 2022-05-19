from cmath import nan
from telnetlib import ENCRYPT
import pyodbc
import os
from datetime import datetime, date
from datetime import timedelta
from dateutil.relativedelta import relativedelta
import pandas as pd
import chardet
from xlsxwriter import Workbook
import glob

from openpyxl import formatting, styles, Workbook as openpyxl_workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill, colors
from openpyxl.styles.colors import Color, ColorDescriptor
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.styles.fills import Fill
from openpyxl.formatting.rule import CellIsRule, Rule
from openpyxl.styles.borders import Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

import email, smtplib, ssl

from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import time
from cryptography.fernet import Fernet

cwd = os.path.dirname(os.path.realpath(__file__))
os.chdir(cwd)

DB_ENCRYPTION = 'NO' #YES/NO

fernet_key = b'zJD8OVkFNpd5N4fJw6pqaWiDrvybkselSQ0fF9SwXfw='
fernet = Fernet(fernet_key)

#'Server=localhost\SQLEXPRESS;'
conn = pyodbc.connect('Driver={ODBC Driver 17 for SQL Server};'
                      'Server=DESKTOP-GLMUIDH\SQLEXPRESS;'
                      'Database=KILP_FINAL;'
                      'Trusted_Connection=yes;')

cursor = conn.cursor()



def change_format(date):
    #print('date', date)
    
    if date == "0100-01-01": 
        date = ''
    
    
    date = str(date).strip()
    
    if date is nan or date =='nan':
        date = ''

    if date:
        
        try:
            #return datetime.strptime(date, "%d-%b-%y").strftime('%Y-%m-%d')
            return datetime.strptime(date, "%m/%d/%Y").strftime('%Y-%m-%d')
        except:
            try: 
                return datetime.strptime(date, "%d-%m-%Y").strftime('%Y-%m-%d')
            except: 
                try:
                    return datetime.strptime(date, "%m/%d/%Y").strftime('%Y-%m-%d')
                except: 
                    return ''
    else:
        return date

def change_display_format(date):
    date = str(date).strip()
    if date:
        try:
            #'%d-%b-%y'
            return datetime.strptime(date, "%Y-%m-%d").strftime('%m/%d/%Y')
        except:
            return date
    

def start():
    current_time = datetime.now() 
    month = str(current_time.month).rjust(2, '0')
    day = str(current_time.day).rjust(2, '0')
    todate = month+''+day+''+str(current_time.year)
    from_name = ''
    for name in glob.glob('Source Data/beneficiary*'):
        ##print(os.path.basename(name))
        filename_e = os.path.basename(name)
        filename = os.path.splitext(filename_e)[0]
        extension = os.path.splitext(filename_e)[1]
        ##print(extension)
        if extension == '.csv':
            beneficiary_file_name = 'beneficiary.csv'
            case_file_name = 'process.csv'
            ##print(beneficiary_file_name)

            if os.path.exists('Source Data/'+beneficiary_file_name):
                print('Processing - '+beneficiary_file_name)
                process_beneficiary_file('Source Data/'+beneficiary_file_name, from_name)
            
            if os.path.exists('Source Data/'+case_file_name):
                print('Processing - '+case_file_name)
                process_case_file('Source Data/'+case_file_name, from_name)

    result_filepath = 'Processed Reports/Charter – Active Employee Report_'+str(todate)+'.xlsx'
    ActiveEmployeeReport(result_filepath)
    print('Processed - Charter – Active Employee Report')

    result_filepath2 = 'Processed Reports/Charter – PERM Report_'+str(todate)+'.xlsx'
    PermReport(result_filepath2)
    print('Processed - Charter – PERM Report')

    result_filepath3 = 'Processed Reports/Charter – Weekly NetOps - P&T Transfer Report_'+str(todate)+'.xlsx'
    WeeklyNetOps(result_filepath3)
    print('Processed - Charter – Weekly NetOps - P&T Transfer Report')
       
        
    
        
def process_beneficiary_file(file_path, from_name):
    with open(file_path,'rb') as f:
        rawdata = b''.join([f.readline() for _ in range(20)])
    enc= chardet.detect(rawdata)['encoding'] #UTF-16
    #print(rawdata)
    df = pd.read_csv(file_path, encoding='utf-8',delimiter=',')
    
    list_h = df.columns.tolist()
    #print(df.iterrows())
    total_rows = len(df)
    for index, row in df.iterrows():
        #print(index)
        #print(row['organization_group_id'])
        #if(index==3):
            #break
            #return False

        organization_xref = ''
        if 'organization_group_id' in list_h:
            organization_xref = str(row['organization_group_id']).strip()
            # organization_xref = fernet.encrypt(organization_xref.encode())

        organization_name = ''
        if "organization_group_name" in list_h:
            organization_name = str(str(row['organization_group_name']).replace("'", "")).strip()
            if DB_ENCRYPTION == "YES":
                organization_name = (fernet.encrypt(organization_name.encode())).decode("utf-8")

        organization_id = ''
        if organization_xref and organization_name:
            # print("INSERT INTO dbo.Organization(OrganizationXref, OrganizationName) VALUES ('{}', '{}')".format(organization_xref, organization_name))
            results = cursor.execute(
                "SELECT * FROM dbo.Organization where OrganizationXref='{}'".format(organization_xref)).fetchall()
            length = len(results)
            if length <= 0:
                cursor.execute(
                    "INSERT INTO dbo.Organization(OrganizationXref, OrganizationName) VALUES ('{}', '{}')".format(
                        organization_xref, organization_name))
                cursor.execute("SELECT @@IDENTITY AS ID;")
                organization_id = cursor.fetchone()[0]
                cursor.commit()
                ##print('inserted')
            else:
                organization_id = results[0].OrganizationId

        petitioner_xref = ''
        if "petitioner_company_id" in list_h:
            petitioner_xref = str(row['petitioner_company_id']).strip()

        petitioner_name = ''
        if "petitioner_company_name" in list_h:
            petitioner_name = str(str(row['petitioner_company_name']).replace("'", "")).strip()

        petitioner_company_of_primary_beneficiary = ''
        if "petitioner_company_of_primary_beneficiary" in list_h:
            petitioner_company_of_primary_beneficiary = str(
                str(row['petitioner_company_of_primary_beneficiary']).replace("'", "")).strip()

        beneficiary_xref = ''
        if "beneficiary_id" in list_h and not pd.isna(row["beneficiary_id"]):
            beneficiary_xref = str(row["beneficiary_id"]).strip()

        beneficiary_type = ''
        if "beneficiary_type" in list_h and not pd.isna(row["beneficiary_type"]):
            beneficiary_type = str(row["beneficiary_type"]).strip()
            
        petitioner_id = ''
        is_primary_beneficiary = 1
        if petitioner_xref and petitioner_name:
            if beneficiary_type == 'Dependent':
                if str(row['primary_beneficiary_id']).strip():
                    ##print("SELECT PetitionerId FROM dbo.Beneficiary where BeneficiaryXref='{}'".format(row['Primary Beneficiary Xref'].strip()))
                    results = cursor.execute(
                        "SELECT PetitionerId FROM dbo.Beneficiary where BeneficiaryXref='{}'".format(
                            str(row['primary_beneficiary_id']).strip())).fetchall()
                    length = len(results)
                    if length > 0:
                        petitioner_id = results[0][0]
                        is_primary_beneficiary = 0


            else:
                ##print("SELECT * FROM dbo.Petitioner where PetitionerXref='{}' and PetitionerName = '{}' and OrganizationId={}".format(petitioner_xref, petitioner_name, organization_id))
                results = cursor.execute(
                    "SELECT * FROM dbo.Petitioner where PetitionerXref='{}' and OrganizationId={}".format(
                        petitioner_xref, organization_id)).fetchall()
                length = len(results)
                if length <= 0:
                    ##print("INSERT INTO dbo.Petitioner(PetitionerXref, PetitionerName, OrganizationId) VALUES ('{}', '{}', '{}')".format(petitioner_xref, petitioner_name, organization_id))
                    cursor.execute(
                        "INSERT INTO dbo.Petitioner(PetitionerXref, PetitionerName, OrganizationId) VALUES ('{}', '{}', '{}')".format(
                            petitioner_xref, petitioner_name, organization_id))
                    cursor.execute("SELECT @@IDENTITY AS ID;")
                    petitioner_id = cursor.fetchone()[0]
                    cursor.commit()
                else:
                    petitioner_id = results[0].PetitionerId

        if petitioner_id:
            

            beneficiary_record_creation_date = ''
            if "beneficiary_record_opened_date" in list_h and str(
                    row["beneficiary_record_opened_date"]).strip() and not pd.isna(
                    row["beneficiary_record_opened_date"]):
                beneficiary_record_creation_date = change_format(row["beneficiary_record_opened_date"])

            beneficiary_record_inactivation_date = ''
            if "beneficiary_record_retired_date" in list_h and str(
                    row["beneficiary_record_retired_date"]).strip() and not pd.isna(
                    row["beneficiary_record_retired_date"]):
                beneficiary_record_inactivation_date = change_format(row["beneficiary_record_retired_date"])

            beneficiary_record_status = 0
            if "beneficiary_status" in list_h and not pd.isna(row["beneficiary_status"]):
                beneficiary_record_status_chk = str(row["beneficiary_status"]).strip()
                if beneficiary_record_status_chk == 'Active':
                    beneficiary_record_status = 1

            beneficiary_employee_id = ''
            if "employee_id" in list_h and not pd.isna(row["employee_id"]):
                beneficiary_employee_id = str(row["employee_id"]).strip()

            beneficiary_last_name = ''
            if "beneficiary_last_name" in list_h and not pd.isna(row["beneficiary_last_name"]):
                beneficiary_last_name = str(str(row["beneficiary_last_name"]).strip()).replace("'", "")

            beneficiary_first_name = ''
            if "beneficiary_first_name" in list_h and not pd.isna(row["beneficiary_first_name"]):
                beneficiary_first_name = str(str(row["beneficiary_first_name"]).strip()).replace("'", "")
                if DB_ENCRYPTION == "YES":
                    beneficiary_first_name = (fernet.encrypt(beneficiary_first_name.encode())).decode("utf-8")

            beneficiary_middle_name = ''
            if "Beneficiary Middle Name" in list_h and not pd.isna(row["Beneficiary Middle Name"]):
                beneficiary_middle_name = str(str(row["Beneficiary Middle Name"]).strip()).replace("'", "")
                if DB_ENCRYPTION == "YES":
                    beneficiary_middle_name = (fernet.encrypt(beneficiary_middle_name.encode())).decode("utf-8")

            primary_beneficiary_id = ''
            if "primary_beneficiary_id" in list_h and not pd.isna(row["primary_beneficiary_id"]):
                primary_beneficiary_id = str(row["primary_beneficiary_id"]).strip()

            # print(primary_beneficiary_id)
            if primary_beneficiary_id == beneficiary_xref:
                is_primary_beneficiary = 1
            else:
                is_primary_beneficiary = 0

            primary_beneficiary_last_name = ''
            if "primary_beneficiary_last_name" in list_h and not pd.isna(row["primary_beneficiary_last_name"]):
                primary_beneficiary_last_name = str(str(row["primary_beneficiary_last_name"]).strip()).replace("'", "")
                if DB_ENCRYPTION == "YES":
                    primary_beneficiary_last_name = (fernet.encrypt(primary_beneficiary_last_name.encode())).decode(
                        "utf-8")

            primary_beneficiary_first_name = ''
            if "primary_beneficiary_first_name" in list_h and not pd.isna(row["primary_beneficiary_first_name"]):
                primary_beneficiary_first_name = str(str(row["primary_beneficiary_first_name"]).strip()).replace("'",
                                                                                                                 "")
                if DB_ENCRYPTION == "YES":
                    primary_beneficiary_first_name = (fernet.encrypt(primary_beneficiary_first_name.encode())).decode(
                        "utf-8")

            relation = ''
            if "relation" in list_h and not pd.isna(row["relation"]):
                relation = str(str(row["relation"]).strip()).replace("'", "")
                if DB_ENCRYPTION == "YES":
                    relation = (fernet.encrypt(relation.encode())).decode("utf-8")

            gender = ''
            if "gender" in list_h and not pd.isna(row["gender"]):
                gender = str(str(row["gender"]).strip()).replace("'", "")
                if DB_ENCRYPTION == "YES":
                    gender = (fernet.encrypt(gender.encode())).decode("utf-8")

            date_of_birth = ''
            if "date_of_birth" in list_h and row["date_of_birth"] and not pd.isna(row["date_of_birth"]):
                date_of_birth = change_format(row["date_of_birth"])

            country_of_birth = ''
            if "country_of_birth" in list_h and not pd.isna(row["country_of_birth"]):
                country_of_birth = str(str(row["country_of_birth"]).strip()).replace("'", "")
                if DB_ENCRYPTION == "YES":
                    country_of_birth = (fernet.encrypt(country_of_birth.encode())).decode("utf-8")

            country_of_citizenship = ''
            if "country_of_citizenship" in list_h and not pd.isna(row["country_of_citizenship"]):
                country_of_citizenship = row["country_of_citizenship"].replace("'", "")
                if DB_ENCRYPTION == "YES":
                    country_of_citizenship = (fernet.encrypt(country_of_citizenship.encode())).decode("utf-8")

            alien_number = ''
            if "alien_number" in list_h and not pd.isna(row["alien_number"]):
                alien_number = str(row["alien_number"]).strip()
                if DB_ENCRYPTION == "YES":
                    alien_number = (fernet.encrypt(str(alien_number).encode())).decode("utf-8")

            date_of_last_entry_into_the_us = ''
            if "date_of_last_entry_into_the_us" in list_h and str(
                    row["date_of_last_entry_into_the_us"]).strip() and not pd.isna(
                    row["date_of_last_entry_into_the_us"]):
                date_of_last_entry_into_the_us = change_format(row["date_of_last_entry_into_the_us"])

            i94_number = ''
            if "i94_number" in list_h and not pd.isna(row["i94_number"]):
                i94_number = str(row["i94_number"]).strip()
                if DB_ENCRYPTION == "YES":
                    i94_number = (fernet.encrypt(i94_number.encode())).decode("utf-8")

            immigration_status = ''
            if "i94_status" in list_h and not pd.isna(row["i94_status"]):
                immigration_status = row["i94_status"].replace("'", "")

            immigration_status_valid_from = ''
            if "i94_valid_from" in list_h and str(row["i94_valid_from"]).strip() and not pd.isna(row["i94_valid_from"]):
                immigration_status_valid_from = change_format(row["i94_valid_from"])

            immigration_status_expiration_status = ''
            if "i94_exp_date" in list_h and str(row["i94_exp_date"]).strip() and not pd.isna(row["i94_exp_date"]):
                if str(row["i94_exp_date"]).strip() == 'D/S':
                    immigration_status_expiration_status = 'D/S'
                else:
                    if 'D/S' in str(row["i94_exp_date"]).strip():
                        split1 = (str(row["i94_exp_date"]).strip()).split('(D/S)')
                        ##print(split1)
                        immigration_status_expiration_status = change_format(split1[0])
                        immigration_status_expiration_status = str(immigration_status_expiration_status) + ' (D/S)'
                    else:
                        immigration_status_expiration_status = change_format(str(row["i94_exp_date"]).strip())

            #print(immigration_status_expiration_status)
            i797_approved_date = ''
            if "I-797 Approved Date" in list_h and str(row["I-797 Approved Date"]).strip() and not pd.isna(
                    row["I-797 Approved Date"]):
                i797_approved_date = change_format(row["I-797 Approved Date"])

            i797_status = ''
            if "I-797 Status" in list_h and not pd.isna(row["I-797 Status"]):
                i797_status = str(row["I-797 Status"]).strip()
                if DB_ENCRYPTION == "YES":
                    i797_status = (fernet.encrypt(i797_status.encode())).decode("utf-8")

            i797_valid_from = ''
            if "i797_valid_from" in list_h and str(row["i797_valid_from"]).strip() and not pd.isna(
                    row["i797_valid_from"]):
                i797_valid_from = change_format(str(row["i797_valid_from"]))

            i797_expiration_date = ''
            if "i797_exp_date" in list_h and str(row["i797_exp_date"]).strip() and not pd.isna(row["i797_exp_date"]):
                i797_expiration_date = change_format(row["i797_exp_date"])
            
            #print('i797_expiration_date ', i797_expiration_date)
            final_niv_status_valid_from = ''
            if "#" in list_h and str(row["final_niv_hl_status_valid_from"]).strip() and not pd.isna(
                    row["final_niv_hl_status_valid_from"]):
                final_niv_status_valid_from = change_format(row["final_niv_hl_status_valid_from"])

            final_niv_maxout_date = ''
            if "final_niv_maxout_date" in list_h and str(row["final_niv_maxout_date"]).strip() and not pd.isna(
                    row["final_niv_maxout_date"]):
                final_niv_maxout_date = change_format(row["final_niv_maxout_date"])

            #print('final_niv_maxout_date ', final_niv_maxout_date)
            maxout_note = ''
            if "Maxout Date Applicability and Note" in list_h and not pd.isna(
                    row["Maxout Date Applicability and Note"]):
                maxout_note = str(str(row["Maxout Date Applicability and Note"]).strip()).replace("'", "")
                if DB_ENCRYPTION == "YES":
                    maxout_note = (fernet.encrypt(maxout_note.encode())).decode("utf-8")

            ped = ''
            if "ped_petition_end_date" in list_h and str(row["ped_petition_end_date"]).strip() and not pd.isna(
                    row["ped_petition_end_date"]):
                ped = change_format(row["ped_petition_end_date"])

            ead_type = ''
            if "ead_type" in list_h and not pd.isna(row["ead_type"]):
                ead_type = str(str(row["ead_type"]).strip()).replace("'", "")
                if DB_ENCRYPTION == "YES":
                    ead_type = (fernet.encrypt(ead_type.encode())).decode("utf-8")

            ead_valid_from = ''
            if "EAD Valid From" in list_h and str(row["EAD Valid From"]).strip() and not pd.isna(row["EAD Valid From"]):
                ead_valid_from = change_format(row["EAD Valid From"])

            ead_expiration_date = ''
            if "ead_exp_date" in list_h and str(row["ead_exp_date"]).strip() and not pd.isna(row["ead_exp_date"]):
                ead_expiration_date = change_format(row["ead_exp_date"])

            ap_valid_from = ''
            if "AP Valid From" in list_h and row["AP Valid From"].strip() and not pd.isna(row["AP Valid From"]):
                ap_valid_from = change_format(row["AP Valid From"])

            ap_expiration_date = ''
            if "ap_exp_date" in list_h and str(row["ap_exp_date"]).strip() and not pd.isna(row["ap_exp_date"]):
                ap_expiration_date = change_format(row["ap_exp_date"])

            ead_ap_type = ''
            if "EAD/AP Type" in list_h and not pd.isna(row["EAD/AP Type"]):
                ead_ap_type = str(row["EAD/AP Type"]).strip()
                if DB_ENCRYPTION == "YES":
                    ead_ap_type = (fernet.encrypt(ead_ap_type.encode())).decode("utf-8")

            ead_ap_valid_from = ''
            if "EAD/AP Valid From" in list_h and str(row["EAD/AP Valid From"]).strip() and not pd.isna(
                    row["EAD/AP Valid From"]):
                ead_ap_valid_from = change_format(row["EAD/AP Valid From"])

            ead_ap_expiration_date = ''
            if "EAD/AP Expiration Date" in list_h and str(row["EAD/AP Expiration Date"]).strip() and not pd.isna(
                    row["EAD/AP Expiration Date"]):
                ead_ap_expiration_date = change_format(row["EAD/AP Expiration Date"])

            ds_2019_valid_from = ''
            if "ds2019_valid_from" in list_h and str(row["ds2019_valid_from"].strip()) and not pd.isna(
                    row["ds2019_valid_from"]):
                ds_2019_valid_from = change_format(row["ds2019_valid_from"])

            ds_2019_expiration_date = ''
            if "ds2019_exp_date" in list_h and str(row["ds2019_exp_date"]).strip() and not pd.isna(
                    row["ds2019_exp_date"]):
                ds_2019_expiration_date = change_format(row["ds2019_exp_date"])

            reentry_permit_expiration_date = ''
            if "re_entry_permit_exp_date" in list_h and row["re_entry_permit_exp_date"] and not pd.isna(
                    row["re_entry_permit_exp_date"]):
                reentry_permit_expiration_date = change_format(row["re_entry_permit_exp_date"])

            green_card_valid_from = ''
            if "Green Card Valid From" in list_h and row["Green Card Valid From"] and not pd.isna(
                    row["Green Card Valid From"]):
                green_card_valid_from = change_format(row["Green Card Valid From"])

            green_card_expiration_date = ''
            if "green_card_exp_date" in list_h and row["green_card_exp_date"] and not pd.isna(
                    row["green_card_exp_date"]):
                green_card_expiration_date = change_format(row["green_card_exp_date"])

            passport_last_name = ''
            if "Passport Last Name" in list_h and not pd.isna(row["Passport Last Name"]):
                passport_last_name = str(str(row["Passport Last Name"]).strip()).replace("'", "")
                if DB_ENCRYPTION == "YES":
                    passport_last_name = (fernet.encrypt(passport_last_name.encode())).decode("utf-8")

            passport_first_name = ''
            if "Passport First Name" in list_h and not pd.isna(row["Passport First Name"]):
                passport_first_name = str(str(row["Passport First Name"]).strip()).replace("'", "")
                if DB_ENCRYPTION == "YES":
                    passport_first_name = (fernet.encrypt(passport_first_name.encode())).decode("utf-8")

            passport_middle_name = ''
            if "Passport Middle Name" in list_h and not pd.isna(row["Passport Middle Name"]):
                passport_middle_name = str(str(row["Passport Middle Name"]).strip()).replace("'", "")
                if DB_ENCRYPTION == "YES":
                    passport_middle_name = (fernet.encrypt(passport_middle_name.encode())).decode("utf-8")

            passport_number = ''
            if "passport_number" in list_h and not pd.isna(row["passport_number"]):
                passport_number = str(row["passport_number"]).strip()
                if DB_ENCRYPTION == "YES":
                    passport_number = (fernet.encrypt(passport_number.encode())).decode("utf-8")

            passport_issuing_country = ''
            if "passport_issuing_country" in list_h and not pd.isna(row["passport_issuing_country"]):
                passport_issuing_country = str(str(row["passport_issuing_country"]).strip()).replace("'", "")
                if DB_ENCRYPTION == "YES":
                    passport_issuing_country = (fernet.encrypt(passport_issuing_country.encode())).decode("utf-8")

            passport_valid_from = ''
            if "passport_valid_from" in list_h and row["passport_valid_from"] and not pd.isna(
                    row["passport_valid_from"]):
                passport_valid_from = change_format(row["passport_valid_from"])
                if DB_ENCRYPTION == "YES":
                    ead_ap_type = (fernet.encrypt(ead_ap_type.encode())).decode("utf-8")

            passport_expiration_date = ''
            if "passport_exp_date" in list_h and row["passport_exp_date"] and not pd.isna(row["passport_exp_date"]):
                passport_expiration_date = change_format(row["passport_exp_date"])

            visa_type = ''
            if "visa_type" in list_h and not pd.isna(row["visa_type"]):
                visa_type = str(row["visa_type"]).strip()
                if DB_ENCRYPTION == "YES":
                    visa_type = (fernet.encrypt(visa_type.encode())).decode("utf-8")

            visa_valid_from = ''
            if "visa_issue_date" in list_h and row["visa_issue_date"] and not pd.isna(row["visa_issue_date"]):
                visa_valid_from = change_format(row["visa_issue_date"])

            visa_expiration_date = ''
            if "visa_exp_date" in list_h and row["visa_exp_date"] and not pd.isna(row["visa_exp_date"]):
                visa_expiration_date = change_format(row["visa_exp_date"])

            employee_hire_date = ''
            if "hire_date" in list_h and row["hire_date"] and not pd.isna(row["hire_date"]):
                employee_hire_date = change_format(row["hire_date"])

            current_job_title = ''
            if "job_title" in list_h and not pd.isna(row["job_title"]):
                current_job_title = str(str(row["job_title"]).strip()).replace("'", "")
                if DB_ENCRYPTION == "YES":
                    current_job_title = (fernet.encrypt(current_job_title.encode())).decode("utf-8")

            work_address_street = ''
            if "job_location_street" in list_h and not pd.isna(row["job_location_street"]):
                work_address_street = str(str(row["job_location_street"]).strip()).replace("'", "")
                if DB_ENCRYPTION == "YES":
                    work_address_street = (fernet.encrypt(work_address_street.encode())).decode("utf-8")

            work_address_city = ''
            if "job_location_city" in list_h and not pd.isna(row["job_location_city"]):
                work_address_city = str(str(row["job_location_city"]).strip()).replace("'", "")
                if DB_ENCRYPTION == "YES":
                    work_address_city = (fernet.encrypt(work_address_city.encode())).decode("utf-8")

            work_address_state = ''
            if "job_location_state" in list_h and not pd.isna(row["job_location_state"]):
                work_address_state = str(str(row["job_location_state"]).strip()).replace("'", "")
                if DB_ENCRYPTION == "YES":
                    work_address_state = (fernet.encrypt(work_address_state.encode())).decode("utf-8")

            work_address_zip = ''
            if "Work Address-Zip" in list_h and not pd.isna(row["Work Address-Zip"]):
                work_address_zip = str(str(row["Work Address-Zip"]).strip()).replace("'", "")
                if DB_ENCRYPTION == "YES":
                    work_address_zip = (fernet.encrypt(work_address_zip.encode())).decode("utf-8")

            work_address_country = ''
            if "Work Address-Country" in list_h and not pd.isna(row["Work Address-Country"]):
                work_address_country = str(row["Work Address-Country"].strip()).replace("'", "")
                if DB_ENCRYPTION == "YES":
                    work_address_country = (fernet.encrypt(work_address_country.encode())).decode("utf-8")

            priority_date_1_date = ''
            if "primary_visa_priority_date" in list_h and row["primary_visa_priority_date"] and not pd.isna(
                    row["primary_visa_priority_date"]):
                priority_date_1_date = change_format(row["primary_visa_priority_date"])

            priority_date_1_category = ''
            if "primary_visa_priority_category" in list_h and not pd.isna(row["primary_visa_priority_category"]):
                priority_date_1_category = str(str(row["primary_visa_priority_category"]).strip()).replace("'", "")

            priority_date_1_country_of_charge = ''
            if "primary_visa_country_of_chargeability" in list_h and not pd.isna(
                    row["primary_visa_country_of_chargeability"]):
                priority_date_1_country_of_charge = str(row["primary_visa_country_of_chargeability"]).strip()

            priority_date_2_date = ''
            if "Priority Date 2-Date" in list_h and row["Priority Date 2-Date"].strip() and not pd.isna(
                    row["Priority Date 2-Date"]):
                priority_date_2_date = change_format(row["Priority Date 2-Date"]).replace("'", "")

            priority_date_2_category = ''
            if "Priority Date 2-Category" in list_h and not pd.isna(row["Priority Date 2-Category"]):
                priority_date_2_category = str(str(row["Priority Date 2-Category"]).strip()).replace("'", "")

            priority_date_2_country_of_charge = ''
            if "Priority Date 2-Country of Chargeability" in list_h and not pd.isna(
                    row["Priority Date 2-Country of Chargeability"]):
                priority_date_2_country_of_charge = str(
                    str(row["Priority Date 2-Country of Chargeability"]).strip()).replace("'", "")

            priority_date_3_date = ''
            if "Priority Date 3-Date" in list_h and row["Priority Date 3-Date"].strip() and not pd.isna(
                    row["Priority Date 3-Date"]):
                priority_date_3_date = change_format(row["Priority Date 3-Date"])

            priority_date_3_category = ''
            if "Priority Date 3-Category" in list_h and not pd.isna(row["Priority Date 3-Category"]):
                priority_date_3_category = str(str(row["Priority Date 3-Category"]).strip()).replace("'", "")

            priority_date_3_country_of_charge = ''
            if "Priority Date 3-Country of Chargeability" in list_h and not pd.isna(
                    row["Priority Date 3-Country of Chargeability"]):
                priority_date_3_country_of_charge = str(
                    str(row["Priority Date 3-Country of Chargeability"]).strip()).replace("'", "")

            priority_date_4_date = ''
            if "Priority Date 4-Date" in list_h and row["Priority Date 4-Date"] and not pd.isna(
                    row["Priority Date 4-Date"]):
                priority_date_4_date = change_format(row["Priority Date 4-Date"])

            priority_date_4_category = ''
            if "Priority Date 4-Category" in list_h and not pd.isna(row["Priority Date 4-Category"]):
                priority_date_4_category = str(str(row["Priority Date 4-Category"]).strip()).replace("'", "")

            priority_date_4_country_of_charge = ''
            if "Priority Date 4-Country of Chargeability" in list_h and not pd.isna(
                    row["Priority Date 4-Country of Chargeability"]):
                priority_date_4_country_of_charge = str(
                    str(row["Priority Date 4-Country of Chargeability"]).strip()).replace("'", "")

            priority_date_5_date = ''
            if "Priority Date 5-Date" in list_h and row["Priority Date 5-Date"] and not pd.isna(
                    row["Priority Date 5-Date"]):
                priority_date_5_date = change_format(row["Priority Date 5-Date"])

            priority_date_5_category = ''
            if "Priority Date 5-Category" in list_h and not pd.isna(row["Priority Date 5-Category"]):
                priority_date_5_category = str(str(row["Priority Date 5-Category"]).strip()).replace("'", "")

            priority_date_5_country_of_charge = ''
            if "Priority Date 5-Country of Chargeability" in list_h and not pd.isna(
                    row["Priority Date 5-Country of Chargeability"]):
                priority_date_5_country_of_charge = str(
                    str(row["Priority Date 5-Country of Chargeability"]).strip()).replace("'", "")

            Beneficiary_Xref2 = ''
            if "beneficiary_case_no" in list_h:
                Beneficiary_Xref2 = str(row['beneficiary_case_no']).strip()

            Current_Immigration_Status = ''
            if "current_status" in list_h:
                Current_Immigration_Status = str(row['current_status']).strip()

            FullName = ''
            if "beneficiary_full_name" in list_h:
                FullName = str(str(row['beneficiary_full_name']).replace("'", "")).strip()

            ImmigrationStatusExpirationDate2 = ''
            if "current_status_exp_date" in list_h and str(row["current_status_exp_date"]).strip() and not pd.isna(
                    row["current_status_exp_date"]):
                ImmigrationStatusExpirationDate2 = change_format(row["current_status_exp_date"])

            I129SEndDate = ''
            if "i129s_exp_date" in list_h and str(row["i129s_exp_date"]).strip() and not pd.isna(row["i129s_exp_date"]):
                I129SEndDate = change_format(row["i129s_exp_date"])

            GreenCardMethod = ''
            if "green_card_method" in list_h:
                GreenCardMethod = str(row['green_card_method']).strip()

            WorkEmail = ''
            if "work_email_id" in list_h:
                WorkEmail = str(row['work_email_id']).strip()

            current_employer = ''
            if "current_employer" in list_h:
                current_employer = str(row['current_employer']).strip()

            EmployeeId = ''
            if "management_info_employee_id" in list_h:
                EmployeeId = str(row['management_info_employee_id']).strip()

            Department = ''
            if "management_info_department" in list_h:
                Department = str(row['management_info_department']).strip()

            Department_Group = ''
            if "management_info_dept_group" in list_h:
                Department_Group = str(row['management_info_dept_group']).strip()

            Department_Number = ''
            if "management_info_dept_number" in list_h:
                Department_Number = str(row['management_info_dept_number']).strip()

            Business_Unit_Code = ''
            if "business_unit_code" in list_h:
                Business_Unit_Code = str(row['business_unit_code']).strip()

            Client_Billing_Code = ''
            if "hr_info_client_billing_code" in list_h:
                Client_Billing_Code = str(row['hr_info_client_billing_code']).strip()
                Client_Billing_Code = Client_Billing_Code.replace("'", "`") 

            ManagerName = ''
            if "management_info_manager" in list_h:
                ManagerName = str(str(row['management_info_manager']).replace("'", "")).strip()

            ManagerEmail = ''
            if "management_info_manager_email" in list_h:
                ManagerEmail = str(row['management_info_manager_email']).strip()

            SecondLevelManager = ''
            if "management_info_second_level_manager" in list_h:
                SecondLevelManager = str(str(row['management_info_second_level_manager']).replace("'", "")).strip()

            SecondLevelManagerEmail = ''
            if "management_info_second_level_manager_email" in list_h:
                SecondLevelManagerEmail = str(row['management_info_second_level_manager_email']).strip()

            BusinessPartnerName = ''
            if "management_info_partner_name" in list_h:
                BusinessPartnerName = str(str(row['management_info_partner_name']).replace("'", "")).strip()

            BusinessPartnerEmail = ''
            if "management_info_partner_email" in list_h:
                BusinessPartnerEmail = str(row['management_info_partner_email']).strip()

            CostCenter = ''
            if "management_info_cost_center" in list_h:
                CostCenter = str(row['management_info_cost_center']).strip()

            CostCenterNumber = ''
            if "management_info_cost_center_number" in list_h:
                CostCenterNumber = str(row['management_info_cost_center_number']).strip()

            ClientBillingCode = ''
            if "management_info_client_billing_code" in list_h:
                ClientBillingCode = str(row['management_info_client_billing_code']).strip()
                ClientBillingCode = ClientBillingCode.replace("'", "`") 

            BusinessUnitCode = ''
            if "management_info_business_unit_code" in list_h:
                BusinessUnitCode = str(row['management_info_business_unit_code']).strip()

            JobTitle = ''
            if "management_info_job_title" in list_h:
                JobTitle = str(str(row['management_info_job_title']).replace("'", "")).strip()

            JobCode = ''
            if "management_info_job_code" in list_h:
                JobCode = str(row['management_info_job_code']).strip()

            EmploymentStartDate = ''
            if "management_info_job_start_date" in list_h and str(
                    row["management_info_job_start_date"]).strip() and not pd.isna(
                    row["management_info_job_start_date"]):
                EmploymentStartDate = change_format(row["management_info_job_start_date"])

            EmploymentEndDate = ''
            if "management_info_job_end_date" in list_h and str(
                    row["management_info_job_end_date"]).strip() and not pd.isna(
                    row["management_info_job_end_date"]):
                EmploymentEndDate = change_format(row["management_info_job_end_date"])

            WorkAddressFull = ''
            if "management_info_work_address" in list_h:
                WorkAddressFull = str(str(row['management_info_work_address']).replace("'", "")).strip()

            WorkLocationCity = ''
            if "management_info_job_location_city" in list_h:
                WorkLocationCity = str(row['management_info_job_location_city']).strip()

            WorkLocationState = ''
            if "management_info_job_location_state" in list_h:
                WorkLocationState = str(row['management_info_job_location_state']).strip()

            Visa_GreenCardMethod = ''
            if "visa_priority_green_card_method" in list_h:
                Visa_GreenCardMethod = str(row['visa_priority_green_card_method']).strip()

            PriorityDate1Note = ''
            if "visa_priority_note" in list_h:
                PriorityDate1Note = str(str(row['visa_priority_note']).replace("'", "")).strip()

            beneficiary_id = ''
            if beneficiary_xref:
                results = cursor.execute("SELECT * FROM dbo.Beneficiary where BeneficiaryXref='{}' and from_name='{}'".format(beneficiary_xref, from_name)).fetchall()
                length = len(results)
                if length <= 0:

                    cursor.execute("INSERT INTO dbo.Beneficiary(PetitionerId, BeneficiaryXref, BeneficiaryType, SourceCreatedDate, IsActive, InactiveDate, LastName, FirstName, MiddleName, PrimaryBeneficiaryXref, PrimaryBeneficiaryLastName, PrimaryBeneficiaryFirstName, RelationType, Gender, BirthDate, BirthCountry, CitizenshipCountry, AlienNumber, MostRecentUSEntryDate, I94Number, ImmigrationStatus, ImmigrationStatusValidFromDate, ImmigrationStatusExpirationDate, MostRecentI797IssueApprovalDate, MostRecentI797Status, MostRecentI797ValidFromDate, I797ExpirationDate, InitialHlEntryDate, FinalNivDate, MaxOutDateNote, EadType, VisaPedDate, EadValidFromDate, EadExpirationDate, AdvanceParoleValidFromDate, AdvanceParoleExpirationDate, EADAPType, EadApValidFromDate, EadApExpirationDate, Ds2019ValidFromDate, Ds2019ExpirationDate, ReEntryPermitExpirationDate, GreenCardValidFromDate, GreenCardExpirationDate, MostRecentPassportLastName, MostRecentPassportFirstName, MostRecentPassportMiddleName, MostRecentPassportNumber, MostRecentPassportIssuingCountry, MostRecentPassportValidFromDate, MostRecentPassportExpirationDate, VisaType, VisaValidFromDate, VisaExpirationDate, from_name, is_primary_beneficiary,Beneficiary_Xref2,FullName,Current_Immigration_Status,CurrentImmigrationStatusExpirationDate2,I129SEndDate,GreenCardMethod,WorkEmail,current_employer,Visa_GreenCardMethod,PriorityDate1Note) VALUES ('{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}','{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}','{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}','{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}','{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}','{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}','{}')".format(petitioner_id, beneficiary_xref, beneficiary_type, beneficiary_record_creation_date, beneficiary_record_status, beneficiary_record_inactivation_date, beneficiary_last_name, beneficiary_first_name, beneficiary_middle_name, primary_beneficiary_id, primary_beneficiary_last_name, primary_beneficiary_first_name, relation, gender, date_of_birth, country_of_birth, country_of_citizenship,  alien_number, date_of_last_entry_into_the_us, i94_number, immigration_status, immigration_status_valid_from, immigration_status_expiration_status, i797_approved_date, i797_status, i797_valid_from, i797_expiration_date,  final_niv_status_valid_from, final_niv_maxout_date, maxout_note, ead_type, ped, ead_valid_from, ead_expiration_date, ap_valid_from, ap_expiration_date, ead_ap_type, ead_ap_valid_from, ead_ap_expiration_date, ds_2019_valid_from, ds_2019_expiration_date, reentry_permit_expiration_date, green_card_valid_from, green_card_expiration_date, passport_last_name, passport_first_name, passport_middle_name, passport_number, passport_issuing_country, passport_valid_from, passport_expiration_date, visa_type, visa_valid_from, visa_expiration_date, from_name, is_primary_beneficiary,Beneficiary_Xref2,FullName,Current_Immigration_Status,ImmigrationStatusExpirationDate2,I129SEndDate,GreenCardMethod,WorkEmail,current_employer,Visa_GreenCardMethod,PriorityDate1Note))
                    cursor.execute("SELECT @@IDENTITY AS ID;")
                    beneficiary_id = cursor.fetchone()[0]
                    cursor.commit()
                else:

                    beneficiary_id = results[0].BeneficiaryId
                    
                    cursor.execute("UPDATE dbo.Beneficiary SET PetitionerId='{}', BeneficiaryXref='{}', BeneficiaryType='{}', SourceCreatedDate='{}', IsActive='{}', InactiveDate='{}', LastName='{}', FirstName='{}', MiddleName='{}', PrimaryBeneficiaryXref='{}', PrimaryBeneficiaryLastName='{}', PrimaryBeneficiaryFirstName='{}', RelationType='{}', Gender='{}', BirthDate='{}', BirthCountry='{}', CitizenshipCountry='{}', AlienNumber='{}', MostRecentUSEntryDate='{}', I94Number='{}', ImmigrationStatus='{}', ImmigrationStatusValidFromDate='{}', ImmigrationStatusExpirationDate='{}', MostRecentI797IssueApprovalDate='{}', MostRecentI797Status='{}', MostRecentI797ValidFromDate='{}', I797ExpirationDate='{}', InitialHlEntryDate='{}', FinalNivDate='{}', MaxOutDateNote='{}', EadType='{}', VisaPedDate='{}', EadValidFromDate='{}', EadExpirationDate='{}', AdvanceParoleValidFromDate='{}', AdvanceParoleExpirationDate='{}', EADAPType='{}', EadApValidFromDate='{}', EadApExpirationDate='{}', Ds2019ValidFromDate='{}', Ds2019ExpirationDate='{}', ReEntryPermitExpirationDate='{}', GreenCardValidFromDate='{}', GreenCardExpirationDate='{}', MostRecentPassportLastName='{}', MostRecentPassportFirstName='{}', MostRecentPassportMiddleName='{}', MostRecentPassportNumber='{}', MostRecentPassportIssuingCountry='{}', MostRecentPassportValidFromDate='{}', MostRecentPassportExpirationDate='{}', VisaType='{}', VisaValidFromDate='{}', VisaExpirationDate='{}', from_name='{}', is_primary_beneficiary='{}',Beneficiary_Xref2='{}',FullName='{}',Current_Immigration_Status='{}',CurrentImmigrationStatusExpirationDate2='{}',I129SEndDate='{}',GreenCardMethod='{}',WorkEmail='{}',current_employer='{}',Visa_GreenCardMethod='{}',PriorityDate1Note='{}' WHERE BeneficiaryId='{}' ".format(petitioner_id, beneficiary_xref, beneficiary_type, beneficiary_record_creation_date, beneficiary_record_status, beneficiary_record_inactivation_date, beneficiary_last_name, beneficiary_first_name, beneficiary_middle_name, primary_beneficiary_id, primary_beneficiary_last_name, primary_beneficiary_first_name, relation, gender, date_of_birth, country_of_birth, country_of_citizenship,  alien_number, date_of_last_entry_into_the_us, i94_number, immigration_status, immigration_status_valid_from, immigration_status_expiration_status, i797_approved_date, i797_status, i797_valid_from, i797_expiration_date,  final_niv_status_valid_from, final_niv_maxout_date, maxout_note, ead_type, ped, ead_valid_from, ead_expiration_date, ap_valid_from, ap_expiration_date, ead_ap_type, ead_ap_valid_from, ead_ap_expiration_date, ds_2019_valid_from, ds_2019_expiration_date, reentry_permit_expiration_date, green_card_valid_from, green_card_expiration_date, passport_last_name, passport_first_name, passport_middle_name, passport_number, passport_issuing_country, passport_valid_from, passport_expiration_date, visa_type, visa_valid_from, visa_expiration_date, from_name, is_primary_beneficiary,Beneficiary_Xref2,FullName,Current_Immigration_Status,ImmigrationStatusExpirationDate2,I129SEndDate,GreenCardMethod,WorkEmail,current_employer,Visa_GreenCardMethod,PriorityDate1Note,beneficiary_id))
                    cursor.commit()
            
            if beneficiary_id:
                results = cursor.execute("SELECT * FROM dbo.BeneficiaryPriorityDate where BeneficiaryId='{}'".format(beneficiary_id)).fetchall()
                length = len(results)
                if length <= 0:
                    cursor.execute("INSERT INTO dbo.BeneficiaryPriorityDate(BeneficiaryId, Priority1Date, Priority1Category, Priority1Country, Priority2Date, Priority2Category, Priority2Country, Priority3Date, Priority3Category, Priority3Country, Priority4Date, Priority4Category, Priority4Country, Priority5Date, Priority5Category, Priority5Country) VALUES ('{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}')".format(beneficiary_id, priority_date_1_date, priority_date_1_category, priority_date_1_country_of_charge, priority_date_2_date, priority_date_2_category, priority_date_2_country_of_charge, priority_date_3_date, priority_date_3_category, priority_date_3_country_of_charge, priority_date_4_date, priority_date_4_category, priority_date_4_country_of_charge, priority_date_5_date, priority_date_5_category, priority_date_5_country_of_charge))
                    cursor.commit()
                else:
                    cursor.execute("UPDATE dbo.BeneficiaryPriorityDate SET BeneficiaryId='{}', Priority1Date='{}', Priority1Category='{}', Priority1Country='{}', Priority2Date='{}', Priority2Category='{}', Priority2Country='{}', Priority3Date='{}', Priority3Category='{}', Priority3Country='{}', Priority4Date='{}', Priority4Category='{}', Priority4Country='{}', Priority5Date='{}', Priority5Category='{}', Priority5Country='{}' WHERE BeneficiaryId='{}'".format(beneficiary_id, priority_date_1_date, priority_date_1_category, priority_date_1_country_of_charge, priority_date_2_date, priority_date_2_category, priority_date_2_country_of_charge, priority_date_3_date, priority_date_3_category, priority_date_3_country_of_charge, priority_date_4_date, priority_date_4_category, priority_date_4_country_of_charge, priority_date_5_date, priority_date_5_category, priority_date_5_country_of_charge, beneficiary_id))
                    cursor.commit()

            if beneficiary_id:
                results = cursor.execute("SELECT * FROM dbo.BeneficiaryEmployment where BeneficiaryId='{}'".format(beneficiary_id)).fetchall()
                length = len(results)
                if length <= 0:
                    cursor.execute("INSERT INTO dbo.BeneficiaryEmployment(BeneficiaryId, EmployeeId, HireDate, JobTitle, Address1, City, StateProvince, ZipCode, Country,Department,Department_Group,Department_Number,Business_Unit_Code,Client_Billing_Code,ManagerName,ManagerEmail,SecondLevelManager,SecondLevelManagerEmail,BusinessPartnerName,BusinessPartnerEmail,CostCenter,CostCenterNumber,ClientBillingCode,BusinessUnitCode,JobCode,EmploymentStartDate,EmploymentEndDate,WorkAddressFull,WorkLocationCity,WorkLocationState) VALUES ('{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}','{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}')".format(beneficiary_id, beneficiary_employee_id, employee_hire_date, current_job_title, work_address_street, work_address_city, work_address_state, work_address_zip, work_address_country,Department,Department_Group,Department_Number,Business_Unit_Code,Client_Billing_Code,ManagerName,ManagerEmail,SecondLevelManager,SecondLevelManagerEmail,BusinessPartnerName,BusinessPartnerEmail,CostCenter,CostCenterNumber,ClientBillingCode,BusinessUnitCode,JobCode,EmploymentStartDate,EmploymentEndDate,WorkAddressFull,WorkLocationCity,WorkLocationState))
                    cursor.commit()
                else:
                    cursor.execute("UPDATE dbo.BeneficiaryEmployment SET BeneficiaryId='{}', EmployeeId='{}', HireDate='{}', JobTitle='{}', Address1='{}', City='{}', StateProvince='{}', ZipCode='{}', Country='{}',Department='{}',Department_Group='{}',Department_Number='{}',Business_Unit_Code='{}',Client_Billing_Code='{}',ManagerName='{}',ManagerEmail='{}',SecondLevelManager='{}',SecondLevelManagerEmail='{}',BusinessPartnerName='{}',BusinessPartnerEmail='{}',CostCenter='{}',CostCenterNumber='{}',ClientBillingCode='{}',BusinessUnitCode='{}',JobCode='{}',EmploymentStartDate='{}',EmploymentEndDate='{}',WorkAddressFull='{}',WorkLocationCity='{}',WorkLocationState='{}' WHERE BeneficiaryId='{}'".format(beneficiary_id, beneficiary_employee_id, employee_hire_date, current_job_title, work_address_street, work_address_city, work_address_state, work_address_zip, work_address_country,Department,Department_Group,Department_Number,Business_Unit_Code,Client_Billing_Code,ManagerName,ManagerEmail,SecondLevelManager,SecondLevelManagerEmail,BusinessPartnerName,BusinessPartnerEmail,CostCenter,CostCenterNumber,ClientBillingCode,BusinessUnitCode,JobCode,EmploymentStartDate,EmploymentEndDate,WorkAddressFull,WorkLocationCity,WorkLocationState,beneficiary_id))
                    cursor.commit()
            

def process_case_file(file_path, from_name):
    with open(file_path,'rb') as f:
        rawdata = b''.join([f.readline() for _ in range(20)])
    enc= chardet.detect(rawdata)['encoding'] #UTF-16

    df = pd.read_csv(file_path, encoding='utf-8',delimiter=',')
    list_h = df.columns.tolist()
    total_rows = len(df)
    for index, row in df.iterrows():
        organization_xref = ''
        if 'organization_group_id' in list_h:
            organization_xref = str(row['organization_group_id']).strip()

        organization_name = ''
        if "organization_group_name" in list_h:
            organization_name = str(str(row['organization_group_name']).replace("'", "")).strip()
            if DB_ENCRYPTION == "YES":
                organization_name = (fernet.encrypt(organization_name.encode())).decode("utf-8")

        organization_id = ''
        if organization_xref and organization_name:
            ##print("SELECT * FROM dbo.Organization where OrganizationXref='{}' and OrganizationName = '{}'".format(organization_xref, organization_name))
            results = cursor.execute(
                "SELECT * FROM dbo.Organization where OrganizationXref='{}'".format(organization_xref)).fetchall()
            length = len(results)
            if length <= 0:
                ##print("INSERT INTO dbo.Organization(OrganizationXref, OrganizationName) VALUES ('{}', '{}')".format(organization_xref, organization_name))
                cursor.execute(
                    "INSERT INTO dbo.Organization(OrganizationXref, OrganizationName) VALUES ('{}', '{}')".format(
                        organization_xref, organization_name))
                cursor.execute("SELECT @@IDENTITY AS ID;")
                organization_id = cursor.fetchone()[0]
                cursor.commit()
                ##print('inserted')
            else:
                organization_id = results[0].OrganizationId

        ##print('oid ', organization_id)
        petitioner_xref = ''
        if "petitioner_id" in list_h:
            petitioner_xref = str(row['petitioner_id']).strip()

        petitioner_name = ''
        if "petitioner_name" in list_h:
            petitioner_name = str(str(row['petitioner_name']).replace("'", "")).strip()

        petitioner_of_primary_beneficiary = ''
        if 'petitioner_of_primary_beneficiary' in list_h:
            petitioner_of_primary_beneficiary = str(row['petitioner_of_primary_beneficiary']).strip()

        beneficiary_xref = ''
        if "beneficiary_id" in list_h and not pd.isna(row["beneficiary_id"]):
            beneficiary_xref = str(row["beneficiary_id"]).strip()
        
        beneficiary_xref2 = ''
        if "beneficiary_case_no" in list_h and not pd.isna(row["beneficiary_case_no"]):
            beneficiary_xref2 = str(row["beneficiary_case_no"]).strip()

        beneficiary_type = ''
        if "beneficiary_type" in list_h and not pd.isna(row["beneficiary_type"]):
            beneficiary_type = str(row["beneficiary_type"]).strip()
        
            
        petitioner_id = ''
        is_primary_beneficiary = 1
        if petitioner_xref and petitioner_name:
            if beneficiary_type == 'Dependent':
                if str(row['primary_fnl_id']).strip():
                    ##print("SELECT PetitionerId FROM dbo.Beneficiary where BeneficiaryXref='{}'".format(row['Primary Beneficiary Xref'].strip()))
                    results = cursor.execute(
                        "SELECT PetitionerId FROM dbo.Beneficiary where BeneficiaryXref='{}'".format(
                            str(row['primary_fnl_id']).strip())).fetchall()
                    length = len(results)
                    if length > 0:
                        petitioner_id = results[0][0]
                        is_primary_beneficiary = 0


            else:
                ##print("SELECT * FROM dbo.Petitioner where PetitionerXref='{}' and PetitionerName = '{}' and OrganizationId={}".format(petitioner_xref, petitioner_name, organization_id))
                results = cursor.execute(
                    "SELECT * FROM dbo.Petitioner where PetitionerXref='{}' and OrganizationId={}".format(
                        petitioner_xref, organization_id)).fetchall()
                length = len(results)
                if length <= 0:
                    ##print("INSERT INTO dbo.Petitioner(PetitionerXref, PetitionerName, OrganizationId) VALUES ('{}', '{}', '{}')".format(petitioner_xref, petitioner_name, organization_id))
                    cursor.execute(
                        "INSERT INTO dbo.Petitioner(PetitionerXref, PetitionerName, OrganizationId) VALUES ('{}', '{}', '{}')".format(
                            petitioner_xref, petitioner_name, organization_id))
                    cursor.execute("SELECT @@IDENTITY AS ID;")
                    petitioner_id = cursor.fetchone()[0]
                    cursor.commit()
                else:
                    petitioner_id = results[0].PetitionerId

        ##print('pid ', petitioner_id)
        # if petitioner_id :
        if True:
            beneficiary_record_creation_date = ''
            if "beneficiary_record_opened_date" in list_h and row["beneficiary_record_opened_date"] and not pd.isna(
                    row["beneficiary_record_opened_date"]):
                beneficiary_record_creation_date = change_format(row["beneficiary_record_opened_date"])

            beneficiary_record_inactivation_date = ''
            if "beneficiary_retired_date" in list_h and row["beneficiary_retired_date"] and not pd.isna(
                    row["beneficiary_retired_date"]):
                beneficiary_record_inactivation_date = change_format(row["beneficiary_retired_date"])

            beneficiary_record_status = 0
            if "beneficiary_status" in list_h and not pd.isna(row["beneficiary_status"]):
                beneficiary_record_status_chk = str(row["beneficiary_status"]).strip()
                if beneficiary_record_status_chk == 'Active':
                    beneficiary_record_status = 1
                else:
                    beneficiary_record_status = 0

            beneficiary_last_name = ''
            if "beneficiary_last_name" in list_h and not pd.isna(row["beneficiary_last_name"]):
                beneficiary_last_name = str(str(row["beneficiary_last_name"]).strip()).replace("'", "")

            beneficiary_first_name = ''
            if "beneficiary_first_name" in list_h and not pd.isna(row["beneficiary_first_name"]):
                beneficiary_first_name = str(str(row["beneficiary_first_name"]).strip()).replace("'", "")
                if DB_ENCRYPTION == "YES":
                    beneficiary_first_name = (fernet.encrypt(beneficiary_first_name.encode())).decode("utf-8")

            beneficiary_full_name = ''
            if "full_name" in list_h and not pd.isna(row["full_name"]):
                beneficiary_full_name = str(str(row["full_name"]).strip()).replace("'", "")

            primary_beneficiary_id = ''
            if "primary_fnl_id" in list_h and not pd.isna(row["primary_fnl_id"]):
                primary_beneficiary_id = str(row["primary_fnl_id"]).strip()

            if beneficiary_type == 'Dependent':
                is_primary_beneficiary = 0
            else:
                is_primary_beneficiary = 1

            primary_beneficiary_last_name = ''
            if "primary_last_name" in list_h and not pd.isna(row["primary_last_name"]):
                primary_beneficiary_last_name = (row["primary_last_name"].strip()).replace("'", "")
                if DB_ENCRYPTION == "YES":
                    primary_beneficiary_last_name = (fernet.encrypt(primary_beneficiary_last_name.encode())).decode(
                        "utf-8")

            primary_beneficiary_first_name = ''
            if "primary_first_name" in list_h and not pd.isna(row["primary_first_name"]):
                primary_beneficiary_first_name = (row["primary_first_name"].strip()).replace("'", "")
                if DB_ENCRYPTION == "YES":
                    primary_beneficiary_first_name = (fernet.encrypt(primary_beneficiary_first_name.encode())).decode(
                        "utf-8")
            
            primary_beneficiary_full_name = ''
            if "primary_full_name" in list_h and not pd.isna(row["primary_full_name"]):
                primary_beneficiary_full_name = (row["primary_full_name"].strip()).replace("'", "")
                if DB_ENCRYPTION == "YES":
                    primary_beneficiary_full_name = (fernet.encrypt(primary_beneficiary_full_name.encode())).decode(
                        "utf-8")

            relation = ''
            if "relationship" in list_h and not pd.isna(row["relationship"]):
                relation = row["relationship"].strip()
                if DB_ENCRYPTION == "YES":
                    relation = (fernet.encrypt(relation.encode())).decode("utf-8")

            immigration_status = ''
            if "i94_status" in list_h and not pd.isna(row["i94_status"]):
                immigration_status = str(row["i94_status"]).strip()

            immigration_status_expiration_status = ''
            if "i94_expiration" in list_h and row["i94_expiration"] and not pd.isna(row["i94_expiration"]):
                if row["i94_expiration"].strip() == 'D/S':
                    immigration_status_expiration_status = 'D/S'
                else:
                    if 'D/S' in row["i94_expiration"]:
                        split1 = str(str(row["i94_expiration"]).strip()).split('(D/S)')
                        immigration_status_expiration_status = change_format(split1[0])
                        immigration_status_expiration_status = str(immigration_status_expiration_status) + ' (D/S)'
                    else:
                        immigration_status_expiration_status = change_format(row["i94_expiration"])

            Current_Immigration_Status = ''
            if "current_status" in list_h:
                Current_Immigration_Status =  str(row["current_status"]).strip()
            
            current_status_expiration = ''
            if "current_status_expiration" in list_h:
                current_status_expiration = change_format(row['current_status_expiration'])

            beneficiary_id = ''
            if beneficiary_xref:
                results = cursor.execute("SELECT * FROM dbo.Beneficiary where BeneficiaryXref='{}'".format(beneficiary_xref)).fetchall()
                length = len(results)
                if length <= 0:
                    cursor.execute("INSERT INTO dbo.Beneficiary(PetitionerId, BeneficiaryXref, BeneficiaryType, SourceCreatedDate, IsActive, InactiveDate, LastName, FirstName, PrimaryBeneficiaryXref, PrimaryBeneficiaryLastName, PrimaryBeneficiaryFirstName, RelationType, ImmigrationStatus, ImmigrationStatusExpirationDate, from_name, is_primary_beneficiary,Beneficiary_Xref2, FullName, PrimaryBeneficiaryFullName, Current_Immigration_Status,CurrentImmigrationStatusExpirationDate2  ) VALUES ('{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}')".format(petitioner_id, beneficiary_xref, beneficiary_type, beneficiary_record_creation_date, beneficiary_record_status, beneficiary_record_inactivation_date, beneficiary_last_name, beneficiary_first_name, primary_beneficiary_id, primary_beneficiary_last_name, primary_beneficiary_first_name, relation, immigration_status, immigration_status_expiration_status, from_name, is_primary_beneficiary,beneficiary_xref2, beneficiary_full_name, primary_beneficiary_full_name, Current_Immigration_Status, current_status_expiration))
                    cursor.execute("SELECT @@IDENTITY AS ID;")
                    beneficiary_id = cursor.fetchone()[0]
                    cursor.commit()
                else:
                    beneficiary_id = results[0].BeneficiaryId
                    #cursor.execute("UPDATE  dbo.Beneficiary SET PetitionerId='{}', BeneficiaryXref='{}', BeneficiaryType='{}', SourceCreatedDate='{}', IsActive='{}', InactiveDate='{}', LastName='{}', FirstName='{}', PrimaryBeneficiaryXref='{}', PrimaryBeneficiaryLastName='{}', PrimaryBeneficiaryFirstName='{}', RelationType='{}', ImmigrationStatus='{}', ImmigrationStatusExpirationDate='{}', from_name='{}', is_primary_beneficiary='{}',Beneficiary_Xref2='{}',FullName='{}',PrimaryBeneficiaryFullName='{}', Current_Immigration_Status='{}', CurrentImmigrationStatusExpirationDate2='{}'   WHERE BeneficiaryId='{}'  ".format(petitioner_id, beneficiary_xref, beneficiary_type, beneficiary_record_creation_date, beneficiary_record_status, beneficiary_record_inactivation_date, beneficiary_last_name, beneficiary_first_name, primary_beneficiary_id, primary_beneficiary_last_name, primary_beneficiary_first_name, relation, immigration_status, immigration_status_expiration_status, from_name, is_primary_beneficiary, beneficiary_xref2, beneficiary_full_name, primary_beneficiary_full_name, Current_Immigration_Status, current_status_expiration, beneficiary_id))
                    #cursor.commit()

            ##print('bid ',beneficiary_id)
            if beneficiary_id:
                case_xref = ''
                if "process_id" in list_h and not pd.isna(row["process_id"]):
                    case_xref = (str(row["process_id"]).strip()).replace(".0", "")
                
                current_process = ''
                if "current_process" in list_h and not pd.isna(row["current_process"]):
                    current_process = str(row["current_process"]).strip()

                
                current_process_name = ''
                if "current_process_name" in list_h and not pd.isna(row["current_process_name"]):
                    current_process_name = str(row["current_process_name"]).strip()
                
                case_creation_date = ''
                if "process_date_opened" in list_h and row["process_date_opened"] and not pd.isna(row["process_date_opened"]):
                    case_creation_date = change_format(row["process_date_opened"])

                case_petition_name = ''
                if "process_type" in list_h and not pd.isna(row["process_type"]):
                    case_petition_name = str(str(row["process_type"]).strip()).replace("'", "")
                    if DB_ENCRYPTION == "YES":
                        case_petition_name = (fernet.encrypt(case_petition_name.encode())).decode("utf-8")

                case_type = ''
                if "process_reference" in list_h and not pd.isna(row["process_reference"]):
                    case_type = str(str(row["process_reference"]).strip()).replace("'", "")
                    if DB_ENCRYPTION == "YES":
                        case_type = (fernet.encrypt(case_type.encode())).decode("utf-8")

                case_description = ''
                if "Case Description" in list_h and not pd.isna(row["Case Description"]):
                    case_description = str(str(row["Case Description"]).strip()).replace("'", "")
                    if DB_ENCRYPTION == "YES":
                        case_description = (fernet.encrypt(case_description.encode())).decode("utf-8")
                
                case_filed_date = ''
                if "filed_date" in list_h and row["filed_date"] and not pd.isna(row["filed_date"]):
                    case_filed_date = change_format(row["filed_date"])
                
                
                case_receipt_number = ''
                if "receipt_number" in list_h and not pd.isna(row["receipt_number"]):
                    case_receipt_number = str(row["receipt_number"]).strip()
                    if DB_ENCRYPTION == "YES":
                        case_receipt_number = (fernet.encrypt(case_receipt_number.encode())).decode("utf-8")

                case_receipt_status = ''
                if "Case Receipt Status" in list_h and not pd.isna(row["Case Receipt Status"]):
                    case_receipt_status = str(row["Case Receipt Status"]).strip()
                    if DB_ENCRYPTION == "YES":
                        case_receipt_status = (fernet.encrypt(case_receipt_status.encode())).decode("utf-8")

                rfe_audit_received_date = ''
                if "rfe_received" in list_h and row["rfe_received"] and not pd.isna(row["rfe_received"]):
                    rfe_audit_received_date = change_format(row["rfe_received"])
                
                rfe_audit_due_date = ''
                if "rfe_due_date" in list_h and row["rfe_due_date"] and not pd.isna(row["rfe_due_date"]):
                    rfe_audit_due_date = change_format(row["rfe_due_date"])
                
                rfe_audit_submitted_date = ''
                if "rfe_response_submitted" in list_h and row["rfe_response_submitted"] and not pd.isna(row["rfe_response_submitted"]):
                    rfe_audit_submitted_date = change_format(row["rfe_response_submitted"])

                primary_case_status = ''
                if "process_status" in list_h and not pd.isna(row["process_status"]):
                    primary_case_status = str(row["process_status"]).strip()

                secondary_case_status = ''
                if "final_action" in list_h and not pd.isna(row["final_action"]):
                    secondary_case_status = str(str(row["final_action"]).strip()).replace("'", "")
                
                case_comments = ''
                if "summary_case_disposition" in list_h and not pd.isna(row["summary_case_disposition"]):
                    case_comments = str(str(row["summary_case_disposition"]).strip()).replace("'", "")
                    if DB_ENCRYPTION == "YES":
                        case_comments = (fernet.encrypt(case_comments.encode())).decode("utf-8")

                case_last_step_completed = ''
                if "last_process_activity" in list_h and not pd.isna(row["last_process_activity"]):
                    case_last_step_completed = str(str(row["last_process_activity"]).strip()).replace("'", "")
                    case_last_step_completed = case_last_step_completed.replace("'", "`")
                    if DB_ENCRYPTION == "YES":
                        case_last_step_completed = (fernet.encrypt(case_last_step_completed.encode())).decode("utf-8")

                case_last_step_completed_date = ''
                if "last_process_activity_date" in list_h and row["last_process_activity_date"] and not pd.isna(row["last_process_activity_date"]):
                    case_last_step_completed_date = change_format(row["last_process_activity_date"])

                case_next_step_to_be_completed = ''
                if "next_unfinished_reminder_subject" in list_h and not pd.isna(row["next_unfinished_reminder_subject"]):
                    case_next_step_to_be_completed = str(str(row["next_unfinished_reminder_subject"]).strip()).replace("'", "")
                    if DB_ENCRYPTION == "YES":
                        case_next_step_to_be_completed = (fernet.encrypt(case_next_step_to_be_completed.encode())).decode("utf-8")
                
                case_next_step_to_be_completed_date = ''
                if "next_unfinished_reminder_expiry" in list_h and row["next_unfinished_reminder_expiry"] and not pd.isna(row["next_unfinished_reminder_expiry"]):
                    case_next_step_to_be_completed_date = change_format(row["next_unfinished_reminder_expiry"])
                
                case_priority_date = ''
                if "Case Priority Date" in list_h and row["Case Priority Date"] and not pd.isna(row["Case Priority Date"]):
                    case_priority_date = change_format(row["Case Priority Date"])

                case_priority_category = ''
                if "Case Priority Category" in list_h and not pd.isna(row["Case Priority Category"]):
                    case_priority_category = str(row["Case Priority Category"]).strip()
                    if DB_ENCRYPTION == "YES":
                        case_priority_category = (fernet.encrypt(case_priority_category.encode())).decode("utf-8")

                case_priority_country = ''
                if "Case Priority Country" in list_h and not pd.isna(row["Case Priority Country"]):
                    case_priority_country = str(row["Case Priority Country"]).strip()
                    if DB_ENCRYPTION == "YES":
                        case_priority_country = (fernet.encrypt(case_priority_country.encode())).decode("utf-8")

                case_approved_date = '' 
                if "approval_date" in list_h and row["approval_date"] and not pd.isna(row["approval_date"]):
                    case_approved_date = change_format(row["approval_date"])
                
                case_valid_from = ''
                if "status_valid_from" in list_h and row["status_valid_from"] and not pd.isna(row["status_valid_from"]):
                    case_valid_from = change_format(row["status_valid_from"])
                
                case_valid_to = ''
                if "status_valid_to" in list_h and row["status_valid_to"] and not pd.isna(row["status_valid_to"]):
                    case_valid_to = change_format(row["status_valid_to"])
                
                case_closed_date = ''
                if "date_closed" in list_h and row["date_closed"] and not pd.isna(row["date_closed"]):
                    case_closed_date = change_format(row["date_closed"])
                
                case_denied_date = ''
                if "date_denied" in list_h and row["date_denied"] and not pd.isna(row["date_denied"]):
                    case_denied_date = change_format(row["date_denied"])
                
                case_withdrawn_date = ''
                if "Case Withdrawn Date" in list_h and row["Case Withdrawn Date"] and not pd.isna(row["Case Withdrawn Date"]):
                    case_withdrawn_date = change_format(row["Case Withdrawn Date"])
                
                case_primary_attorney = ''
                if "Case Primary Attorney" in list_h and not pd.isna(row["Case Primary Attorney"]):
                    case_primary_attorney = str(str(row["Case Primary Attorney"]).strip()).replace("'", "")
                    if DB_ENCRYPTION == "YES":
                        case_primary_attorney = (fernet.encrypt(case_primary_attorney.encode())).decode("utf-8")
                
                case_reviewing_attorney = ''
                if "Case Reviewing Attorney" in list_h and not pd.isna(row["Case Reviewing Attorney"]):
                    case_reviewing_attorney = str(str(row["Case Reviewing Attorney"]).strip()).replace("'", "")
                    if DB_ENCRYPTION == "YES":
                        case_reviewing_attorney = (fernet.encrypt(case_reviewing_attorney.encode())).decode("utf-8")
                
                case_primary_case_manager = ''
                if "Case Primary Case Manager" in list_h and not pd.isna(row["Case Primary Case Manager"]):
                    case_primary_case_manager = str(str(row["Case Primary Case Manager"]).strip()).replace("'", "")
                    if DB_ENCRYPTION == "YES":
                        case_primary_case_manager = (fernet.encrypt(case_primary_case_manager.encode())).decode("utf-8")
                
                petition_xref = ''
                if "process_petition_id" in list_h and not pd.isna(row["process_petition_id"]):
                    petition_xref = (str(row["process_petition_id"]).strip()).replace(".0", "")


                ####
                CaseReceivedDate= ''
                if "receipt_date" in list_h and str(row["receipt_date"]).strip() and not pd.isna(row["receipt_date"]):
                    CaseReceivedDate= change_format(row["receipt_date"])

                RFEDocsReqestedDate = ''
                if "rfe_docs_requested" in list_h and str(row["rfe_docs_requested"]).strip() and not pd.isna(row["rfe_docs_requested"]):
                    RFEDocsReqestedDate = change_format(row["rfe_docs_requested"])

                RFEDocsReceivedDate = ''
                if "rfe_docs_received" in list_h and str(row["rfe_docs_received"]).strip() and not pd.isna(
                        row["rfe_docs_received"]):
                    RFEDocsReceivedDate = change_format(row["rfe_docs_received"])

                PERMAuditReceivedDate = ''
                if "audit_notice_received" in list_h and str(row["audit_notice_received"]).strip() and not pd.isna(
                        row["audit_notice_received"]):
                    PERMAuditReceivedDate = change_format(row["audit_notice_received"])

                PERMAuditSubmittedDate = ''
                if "audit_response_sent_to_dol" in list_h and str(row["audit_response_sent_to_dol"]).strip() and not pd.isna(
                        row["audit_response_sent_to_dol"]):
                    PERMAuditSubmittedDate = change_format(row["audit_response_sent_to_dol"])

                SecondaryCaseStatusDate= ''
                if "final_action_date" in list_h and str(row["final_action_date"]).strip() and not pd.isna(row["final_action_date"]):
                    SecondaryCaseStatusDate= change_format(row["final_action_date"])

                DaysSinceLastStepCompleted= ''
                if "days_since_last_activity" in list_h:
                    DaysSinceLastStepCompleted= (str(row['days_since_last_activity']).strip()).replace('.0', '')

                visa_preference_category= ''
                if "visa_preference_category" in list_h:
                    visa_preference_category= str(row['visa_preference_category']).strip()

                visa_priority_country= ''
                if "visa_priority_country" in list_h:
                    visa_priority_country= str(row['visa_priority_country']).strip()

                PartnerXref= ''
                if "partner_id" in list_h:
                    PartnerXref= str(row['partner_id']).strip()

                PartnerLastName= ''
                if "partner_last_name" in list_h:
                    PartnerLastName= str(row['partner_last_name']).strip()

                PartnerFirstName= ''
                if "partner_first_name" in list_h:
                    PartnerFirstName= str(row['partner_first_name']).strip()

                AssociateXref= ''
                if "associate_id" in list_h:
                    AssociateXref= str(row['associate_id']).strip()

                AssociateLastName= ''
                if "associate_last_name" in list_h:
                    AssociateLastName= str(row['associate_last_name']).strip()

                AssociateFirstName= ''
                if "associate_first_name" in list_h:
                    AssociateFirstName= str(row['associate_first_name']).strip()

                SupervisoryParalegalXref= ''
                if "supervisory_paralegal_id" in list_h:
                    SupervisoryParalegalXref= (str(row['supervisory_paralegal_id']).strip()).replace('.0', '')

                SupervisoryParalegalLastName= ''
                if "supervisory_paralegal_last_name" in list_h:
                    SupervisoryParalegalLastName= str(row['supervisory_paralegal_last_name']).strip()

                SupervisoryParalegalFirstName= ''
                if "supervisory_paralegal_first_name" in list_h:
                    SupervisoryParalegalFirstName= str(row['supervisory_paralegal_first_name']).strip()

                ParalegalXref= ''
                if "paralegal_id" in list_h:
                    ParalegalXref= (str(row['paralegal_id']).strip()).replace('.0', '')

                ParalegalLastName= ''
                if "paralegal_last_name" in list_h:
                    ParalegalLastName= str(row['paralegal_last_name']).strip()

                ParalegalFirstName= ''
                if "paralegal_first_name" in list_h:
                    ParalegalFirstName= str(row['paralegal_first_name']).strip()

                AccountManagerXref= ''
                if "account_manager_id" in list_h:
                    AccountManagerXref= str(row['account_manager_id']).strip()

                AccountManagerLastName= ''
                if "account_manager_last_name" in list_h:
                    AccountManagerLastName= str(row['account_manager_last_name']).strip()

                AccountManagerFirstName= ''
                if "account_manager_first_name" in list_h:
                    AccountManagerFirstName= str(row['account_manager_first_name']).strip()

                SpecialInstructionFlag= ''
                if "special_instruction_flag" in list_h:
                    SpecialInstructionFlag= str(row['special_instruction_flag']).strip()

                SpecialInstructionInfo= ''
                if "special_instruction_info" in list_h:
                    SpecialInstructionInfo= str(row['special_instruction_info']).strip()

                ClientBillingCode= ''
                if "client_billing_code" in list_h:
                    ClientBillingCode= str(row['client_billing_code']).strip()

                OnlineIntakeDate= ''
                if "online_intake_date" in list_h and str(row["online_intake_date"]).strip() and not pd.isna(row["online_intake_date"]):
                    OnlineIntakeDate= change_format(row["online_intake_date"])


                questionnairesenttomanager= ''
                if "questionnaire_sent_to_manager" in list_h:
                    questionnairesenttomanager= str(row['questionnaire_sent_to_manager']).strip()

                questionnairessenttofn= ''
                if "questionnaires_sent_to_fn" in list_h:
                    questionnairessenttofn= str(row['questionnaires_sent_to_fn']).strip()

                followupwithfnforrequestedinformation= ''
                if "follow_up_with_fn_for_requested_information" in list_h:
                    followupwithfnforrequestedinformation= change_format(str(row['follow_up_with_fn_for_requested_information']).strip())

                questionnairecompletedandreturnedbymanager= ''
                if "questionnaire_completed_and_returned_by_manager" in list_h:
                    questionnairecompletedandreturnedbymanager= change_format(str(row['questionnaire_completed_and_returned_by_manager']).strip())

                questionnairecompletedandreturnedbyfn= ''
                if "questionnaire_completed_and_returned_by_fn" in list_h:
                    questionnairecompletedandreturnedbyfn= change_format(str(row['questionnaire_completed_and_returned_by_fn']).strip())

                employersubmissionquestionnairecompleted= ''
                if "employer_submission_questionnaire_completed" in list_h:
                    employersubmissionquestionnairecompleted= change_format(str(row['employer_submission_questionnaire_completed']).strip())

                allpetitioningcompanyinforeceived= ''
                if "all_petitioning_company_info_received" in list_h:
                    allpetitioningcompanyinforeceived= change_format(str(row['all_petitioning_company_info_received']).strip())

                allfndocsreceived= ''
                if "all_fn_docs_received" in list_h:
                    allfndocsreceived= change_format(str(row['all_fn_docs_received']).strip())

                fncompletedquestionnairesandacknowledgement= ''
                if "fn_completed_questionnaires_and_acknowledgement" in list_h:
                    fncompletedquestionnairesandacknowledgement= change_format(str(row['fn_completed_questionnaires_and_acknowledgement']).strip())

                fnquestionnairescompleted= ''
                if "fn_questionnaires_completed" in list_h:
                    fnquestionnairescompleted= change_format(str(row['fn_questionnaires_completed']).strip())

                lcafiled= ''
                if "lca_filed" in list_h:
                    lcafiled= change_format(str(row['lca_filed']).strip())

                lcacasenumber= ''
                if "lca_case_number" in list_h:
                    lcacasenumber= str(row['lca_case_number']).strip()

                lcacertified= ''
                if "lca_certified" in list_h:
                    lcacertified= change_format(str(row['lca_certified']).strip())

                formsanddocumentationprepped= ''
                if "forms_and_documentation_prepped" in list_h:
                    formsanddocumentationprepped= change_format(str(row['forms_and_documentation_prepped']).strip())

                formsanddocumentationsubmittedforsignature= ''
                if "forms_and_documentation_submitted_for_signature" in list_h:
                    formsanddocumentationsubmittedforsignature= change_format(str(row['forms_and_documentation_submitted_for_signature']).strip())

                signedformsandletterreceived= ''
                if "signed_forms_and_letter_received" in list_h:
                    signedformsandletterreceived= change_format(str(row['signed_forms_and_letter_received']).strip())

                dateaosformssentforsignature= ''
                if "date_aos_forms_sent_for_signature" in list_h and str(row["date_aos_forms_sent_for_signature"]).strip() and not pd.isna(row["date_aos_forms_sent_for_signature"]):
                    dateaosformssentforsignature= change_format(row["date_aos_forms_sent_for_signature"])

                datesignedaosformsreceived= ''
                if "date_signed_aos_forms_received" in list_h and str(row["date_signed_aos_forms_received"]).strip() and not pd.isna(row["date_signed_aos_forms_received"]):
                    datesignedaosformsreceived= change_format(row["date_signed_aos_forms_received"])

                targetfiledate= ''
                if "target_file_date" in list_h and str(row["target_file_date"]).strip() and not pd.isna(row["target_file_date"]):
                    targetfiledate= change_format(row["target_file_date"])

                applicationfiled= ''
                if "application_filed" in list_h:
                    applicationfiled= change_format(str(row['application_filed']).strip())

                applicationfiledwithcis= ''
                if "application_filed_with_cis" in list_h:
                    applicationfiledwithcis= change_format(str(row['application_filed_with_cis']).strip())

                petitionfiledwithcis= ''
                if "petition_filed_with_cis" in list_h:
                    petitionfiledwithcis= change_format(str(row['petition_filed_with_cis']).strip())

                formi129filedwithcis= ''
                if "form_i129_filed_with_cis" in list_h:
                    formi129filedwithcis= change_format(str(row['form_i129_filed_with_cis']).strip())

                aosapplicationfiled = ''
                if "aos_application_filed" in list_h:
                    aosapplicationfiled = change_format(str(row['aos_application_filed']).strip())

                tnpacketsenttofnforpoeprocessing= ''
                if "tn_packet_sent_to_fn_for_poe_processing" in list_h:
                    tnpacketsenttofnforpoeprocessing = change_format(str(row['tn_packet_sent_to_fn_for_poe_processing']).strip())

                appealmotionduedate= ''
                if "appeal_motion_due_date" in list_h and str(row["appeal_motion_due_date"]).strip() and not pd.isna(row["appeal_motion_due_date"]):
                    appealmotionduedate= change_format(row["appeal_motion_due_date"])

                appealmotionfiled= ''
                if "appeal_motion_filed" in list_h:
                    appealmotionfiled= change_format(str(row['appeal_motion_filed']).strip())

                consularinterviewdate = ''
                if "consular_interview_date" in list_h and str(row["consular_interview_date"]).strip() and not pd.isna(
                        row["consular_interview_date"]):
                    consularinterviewdate = change_format(row["consular_interview_date"])

                supplementalbriefdocsfiled= ''
                if "supplemental_brief_docs_filed" in list_h:
                    supplementalbriefdocsfiled= change_format(str(row['supplemental_brief_docs_filed']).strip())

                docketdatebalca = ''
                if "docket_date_balca" in list_h and str(row["docket_date_balca"]).strip() and not pd.isna(row["docket_date_balca"]):
                    docketdatebalca = change_format(row["docket_date_balca"])

                datewithdrawrequestsenttouscis = ''
                if "date_withdraw_request_sent_to_uscis" in list_h and str(row["date_withdraw_request_sent_to_uscis"]).strip() and not pd.isna(
                        row["date_withdraw_request_sent_to_uscis"]):
                    datewithdrawrequestsenttouscis = change_format(row["date_withdraw_request_sent_to_uscis"])

                withdrawalrequestconfirmedbydoluscis= ''
                if "withdrawal_request_confirmed_by_dol_uscis" in list_h:
                    withdrawalrequestconfirmedbydoluscis= change_format(str(row['withdrawal_request_confirmed_by_dol_uscis']).strip())

                approvalpackagesent= ''
                if "approval_package_sent" in list_h:
                    approvalpackagesent= change_format(str(row['approval_package_sent']).strip())

                h1bregistrationsubmitted= ''
                if "h1b_registration_submitted" in list_h:
                    h1bregistrationsubmitted= change_format(str(row['h1b_registration_submitted']).strip())

                h1bregistrationresult= ''
                if "h1b_registration_result" in list_h:
                    h1bregistrationresult= change_format(str(row['h1b_registration_result']).strip())

                h1bcapregistrationselected= ''
                if "h1b_cap_registration_selected" in list_h:
                    h1bcapregistrationselected= change_format(str(row['h1b_cap_registration_selected']).strip())

                I907filedupgradedtopremprocessing= ''
                if "i907_filed_upgraded_to_prem_processing" in list_h:
                    I907filedupgradedtopremprocessing= change_format(str(row['i907_filed_upgraded_to_prem_processing']).strip())

                premiumprocessingfeereceivedfromfn= ''
                if "premium_processing_fee_received_from_fn" in list_h:
                    premiumprocessingfeereceivedfromfn= change_format(str(row['premium_processing_fee_received_from_fn']).strip())

                receipts= ''
                if "receipts" in list_h:
                    receipts= str(row['receipts']).strip()

                I485receiptdate= ''
                if "i485_receipt_date" in list_h and str(row["i485_receipt_date"]).strip() and not pd.isna(row["i485_receipt_date"]):
                    I485receiptdate= change_format(row["i485_receipt_date"])

                I485jportabilityreceiptdate= ''
                if "i485j_portability_receipt_date" in list_h and str(row["i485j_portability_receipt_date"]).strip() and not pd.isna(row["i485j_portability_receipt_date"]):
                    I485jportabilityreceiptdate= change_format(row["i485j_portability_receipt_date"])

                I131receiptdate= ''
                if "i131_receipt_date" in list_h and str(row["i131_receipt_date"]).strip() and not pd.isna(row["i131_receipt_date"]):
                    I131receiptdate= change_format(row["i131_receipt_date"])


                apreceiptnoticereceived= ''
                if "ap_receipt_notice_received" in list_h:
                    apreceiptnoticereceived= change_format(str(row['ap_receipt_notice_received']).strip())

                eadreceiptnoticereceived= ''
                if "ead_receipt_notice_received" in list_h:
                    eadreceiptnoticereceived= change_format(str(row['ead_receipt_notice_received']).strip())

                petitioningjobtitle= ''
                if "petitioning_job_title" in list_h:
                    petitioningjobtitle= str(row['petitioning_job_title']).strip()

                petitioningjoblocation= ''
                if "petitioning_job_location" in list_h:
                    petitioningjoblocation= str(row['petitioning_job_location']).strip()

                permmemosenttoemployer= ''
                if "perm_memo_sent_to_employer" in list_h:
                    permmemosenttoemployer= str(row['perm_memo_sent_to_employer']).strip()

                approvalofpermmemoreceived= ''
                if "approvalofpermmemoreceived" in list_h:
                    approvalofpermmemoreceived= change_format(str(row['approvalofpermmemoreceived']).strip())

                employeeworkexperiencechartsent= ''
                if "employee_work_experience_chart_sent" in list_h:
                    employeeworkexperiencechartsent= change_format(str(row['employee_work_experience_chart_sent']).strip())

                employeeworkexperiencechartreceived= ''
                if "employee_work_experience_chart_received" in list_h:
                    employeeworkexperiencechartreceived = change_format(str(row['employee_work_experience_chart_received']).strip())

                employmentverificationletterssenttoemployee= ''
                if "employment_verification_letters_sent_to_employee" in list_h:
                    employmentverificationletterssenttoemployee= change_format(str(row['employment_verification_letters_sent_to_employee']).strip())

                signedemploymentverificationlettersreceived= ''
                if "signed_employment_verification_letters_received" in list_h:
                    signedemploymentverificationlettersreceived= change_format(str(row['signed_employment_verification_letters_received']).strip())

                prevailingwagedeterminationrequestsubmittedtodol= ''
                if "prevailing_wage_determination_request_submitted_to_dol" in list_h:
                    prevailingwagedeterminationrequestsubmittedtodol = change_format(str(row['prevailing_wage_determination_request_submitted_to_dol']).strip())

                prevailingwagedeterminationissuedbydol= ''
                if "prevailing_wage_determination_issued_by_dol" in list_h:
                    prevailingwagedeterminationissuedbydol= change_format(str(row['prevailing_wage_determination_issued_by_dol']).strip())

                recruitmentinstructionssenttocompany= ''
                if "recruitment_instructions_sent_to_company" in list_h:
                    recruitmentinstructionssenttocompany= change_format(str(row['recruitment_instructions_sent_to_company']).strip())

                joborderplacedwithswa= ''
                if "job_order_placed_with_swa" in list_h:
                    joborderplacedwithswa= change_format(str(row['job_order_placed_with_swa']).strip())

                noticeoffilingposted= ''
                if "notice_of_filing_posted" in list_h:
                    noticeoffilingposted= change_format(str(row['notice_of_filing_posted']).strip())

                intranetnoticeoffilingposted= ''
                if "intranet_notice_of_filing_posted" in list_h:
                    intranetnoticeoffilingposted= change_format(str(row['intranet_notice_of_filing_posted']).strip())

                noticeoffilingremovedsigned= ''
                if "notice_of_filing_removed_signed" in list_h:
                    noticeoffilingremovedsigned= change_format(str(row['notice_of_filing_removed_signed']).strip())

                intranetnoticeoffilingremoved= ''
                if "intranet_notice_of_filing_removed" in list_h:
                    intranetnoticeoffilingremoved= change_format(str(row['intranet_notice_of_filing_removed']).strip())

                _1stsundayadplaced= ''
                if "_1st_sunday_ad_placed" in list_h:
                    _1stsundayadplaced= change_format(str(row['_1st_sunday_ad_placed']).strip())

                _2ndsundayadplaced= ''
                if "_2nd_sunday_ad_placed" in list_h:
                    _2ndsundayadplaced= change_format(str(row['_2nd_sunday_ad_placed']).strip())

                _1stadditionalrecruitmentstepplaced= ''
                if "_1st_additional_recruitment_step_placed" in list_h:
                    _1stadditionalrecruitmentstepplaced = change_format(str(row['_1st_additional_recruitment_step_placed']).strip())

                _2ndadditionalrecruitmentstepplaced= ''
                if "_2nd_additional_recruitment_step_placed" in list_h:
                    _2ndadditionalrecruitmentstepplaced= change_format(str(row['_2nd_additional_recruitment_step_placed']).strip())

                _3rdadditionalrecruitmentstepplaced= ''
                if "_3rd_additional_recruitment_step_placed" in list_h:
                    _3rdadditionalrecruitmentstepplaced= change_format(str(row['_3rd_additional_recruitment_step_placed']).strip())

                datedcopiesofallrecruitmentreceived= ''
                if "dated_copies_of_all_recruitment_received" in list_h:
                    datedcopiesofallrecruitmentreceived = change_format(str(row['dated_copies_of_all_recruitment_received']).strip())

                completedevaluationquestionnairesandresumesreceived= ''
                if "completed_evaluation_questionnaires_and_resumes_received" in list_h:
                    completedevaluationquestionnairesandresumesreceived = change_format(str(row['completed_evaluation_questionnaires_and_resumes_received']).strip())

                recruitmentreportsenttocompany= ''
                if "recruitment_report_sent_to_company" in list_h:
                    recruitmentreportsenttocompany= change_format(str(row['recruitment_report_sent_to_company']).strip())

                recruitmentreportreceived= ''
                if "recruitment_report_received" in list_h:
                    recruitmentreportreceived= change_format(str(row['recruitment_report_received']).strip())

                form9089senttofnandemployer= ''
                if "form_9089_sent_to_fn_and_employer" in list_h:
                    form9089senttofnandemployer= change_format(str(row['form_9089_sent_to_fn_and_employer']).strip())

                editstoform9089receivedfromfnandemployer= ''
                if "edits_to_form_9089_received_from_fn_and_employer" in list_h:
                    editstoform9089receivedfromfnandemployer= change_format(str(row['edits_to_form_9089_received_from_fn_and_employer']).strip())

                form9089submittedtodol= ''
                if "form_9089_submitted_to_dol" in list_h:
                    form9089submittedtodol= change_format(str(row['form_9089_submitted_to_dol']).strip())

                inputcallconducted= ''
                if "input_call_conducted" in list_h:
                    inputcallconducted= change_format(str(row['input_call_conducted']).strip())

                inputstatementreceived= ''
                if "input_statement_received" in list_h:
                    inputstatementreceived= change_format(str(row['input_statement_received']).strip())

                casestrategyandlettersplansent= ''
                if "case_strategy_and_letters_plan_sent" in list_h:
                    casestrategyandlettersplansent= change_format(str(row['case_strategy_and_letters_plan_sent']).strip())

                longlettersenttofn= ''
                if "long_letter_sent_to_fn" in list_h:
                    longlettersenttofn= change_format(str(row['long_letter_sent_to_fn']).strip())

                shortletterssenttofn= ''
                if "short_letters_sent_to_fn" in list_h:
                    shortletterssenttofn= change_format(str(row['short_letters_sent_to_fn']).strip())

                numberoftotalapplicants= ''
                if "number_of_total_applicants" in list_h:
                    numberoftotalapplicants= change_format(str(row['number_of_total_applicants']).strip())

                numberofnonusworkers= ''
                if "number_of_non_us_workers" in list_h:
                    numberofnonusworkers= change_format(str(row['number_of_non_us_workers']).strip())

                numberofphonescreensconducted= ''
                if "number_of_phone_screens_conducted" in list_h:
                    numberofphonescreensconducted= change_format(str(row['number_of_phone_screens_conducted']).strip())

                numberofmanagerinterviewsconducted= ''
                if "number_of_manager_interviews_conducted" in list_h:
                    numberofmanagerinterviewsconducted= change_format(str(row['number_of_manager_interviews_conducted']).strip())



                case_id = ''
                ##print('cx ', case_xref)
                if case_xref:
                    
                    ##print("SELECT * FROM [dbo].[Case] where BeneficiaryId='{}' and CaseXref='{}' and from_name='{}'".format(beneficiary_id, case_xref, from_name))
                    results = cursor.execute("SELECT * FROM [dbo].[Case] where CaseXref='{}' ".format(case_xref)).fetchall()
                    length = len(results)
                    if length <= 0:
                        cursor.execute("INSERT INTO [dbo].[Case](CaseXref, BeneficiaryId, SourceCreatedDate, CasePetitionName, CaseType, CaseDescription, CaseFiledDate, ReceiptNumber, ReceiptStatus, RFEAuditReceivedDate,RFEAuditDueDate, RFEAuditSubmittedDate, PrimaryCaseStatus, SecondaryCaseStatus, CaseComments, LastStepCompleted, LastStepCompletedDate, NextStepAction, NextStepActionDueDate, PriorityDate, PriorityCategory, PriorityCountry, CaseApprovedDate, CaseValidFromDate, CaseExpirationDate, CaseClosedDate, CaseDeniedDate, CaseWithdrawnDate, CasePrimaryAttorney, CaseReviewingAttorney, CasePrimaryCaseManager, PetitionXref, IsCurrentProcess, CurrentProcessName, RFEDocsReqestedDate, RFEDocsReceivedDate, SecondaryCaseStatusDate, DaysSinceLastStepCompleted, visa_preference_category, visa_priority_country, PartnerXref, PartnerLastName, PartnerFirstName, AssociateXref, AssociateLastName, AssociateFirstName, SupervisoryParalegalXref, SupervisoryParalegalLastName, SupervisoryParalegalFirstName, ParalegalXref, ParalegalLastName, ParalegalFirstName, AccountManagerXref, AccountManagerLastName, AccountManagerFirstName, SpecialInstructionFlag, SpecialInstructionInfo, ClientBillingCode, OnlineIntakeDate, questionnairesenttomanager, questionnairessenttofn, followupwithfnforrequestedinformation, questionnairecompletedandreturnedbymanager, questionnairecompletedandreturnedbyfn, employersubmissionquestionnairecompleted, allpetitioningcompanyinforeceived, allfndocsreceived, fncompletedquestionnairesandacknowledgement, fnquestionnairescompleted, lcafiled, lcacasenumber, lcacertified, formsanddocumentationprepped, formsanddocumentationsubmittedforsignature, signedformsandletterreceived, dateaosformssentforsignature, datesignedaosformsreceived, targetfiledate, applicationfiled, applicationfiledwithcis, petitionfiledwithcis, formi129filedwithcis, aosapplicationfiled, tnpacketsenttofnforpoeprocessing, appealmotionduedate, appealmotionfiled, consularinterviewdate, supplementalbriefdocsfiled, docketdatebalca, datewithdrawrequestsenttouscis, withdrawalrequestconfirmedbydoluscis, approvalpackagesent, h1bregistrationsubmitted, h1bregistrationresult, h1bcapregistrationselected, I907filedupgradedtopremprocessing, premiumprocessingfeereceivedfromfn, receipts, I485receiptdate, I485jportabilityreceiptdate, I131receiptdate, apreceiptnoticereceived, eadreceiptnoticereceived, petitioningjobtitle, petitioningjoblocation, permmemosenttoemployer, approvalofpermmemoreceived, employeeworkexperiencechartsent, employeeworkexperiencechartreceived, employmentverificationletterssenttoemployee, signedemploymentverificationlettersreceived, prevailingwagedeterminationrequestsubmittedtodol, prevailingwagedeterminationissuedbydol, recruitmentinstructionssenttocompany, joborderplacedwithswa, noticeoffilingposted, intranetnoticeoffilingposted, noticeoffilingremovedsigned, intranetnoticeoffilingremoved, _1stsundayadplaced, _2ndsundayadplaced, _1stadditionalrecruitmentstepplaced, _2ndadditionalrecruitmentstepplaced, _3rdadditionalrecruitmentstepplaced, datedcopiesofallrecruitmentreceived, completedevaluationquestionnairesandresumesreceived, recruitmentreportsenttocompany, recruitmentreportreceived, form9089senttofnandemployer, editstoform9089receivedfromfnandemployer, form9089submittedtodol, inputcallconducted, inputstatementreceived, casestrategyandlettersplansent, longlettersenttofn, shortletterssenttofn, numberoftotalapplicants, numberofnonusworkers, numberofphonescreensconducted, numberofmanagerinterviewsconducted ) VALUES ('{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}','{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}','{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}','{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}','{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}','{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}' )".format(case_xref, beneficiary_id, case_creation_date, case_petition_name, case_type, case_description, case_filed_date, case_receipt_number, case_receipt_status, rfe_audit_received_date, rfe_audit_due_date, rfe_audit_submitted_date, primary_case_status, secondary_case_status, case_comments, case_last_step_completed, case_last_step_completed_date, case_next_step_to_be_completed, case_next_step_to_be_completed_date, case_priority_date, case_priority_category, case_priority_country, case_approved_date, case_valid_from, case_valid_to, case_closed_date, case_denied_date, case_withdrawn_date, case_primary_attorney, case_reviewing_attorney, case_primary_case_manager, petition_xref,  current_process, current_process_name, RFEDocsReqestedDate, RFEDocsReceivedDate, SecondaryCaseStatusDate, DaysSinceLastStepCompleted, visa_preference_category, visa_priority_country, PartnerXref, PartnerLastName, PartnerFirstName, AssociateXref, AssociateLastName, AssociateFirstName, SupervisoryParalegalXref, SupervisoryParalegalLastName, SupervisoryParalegalFirstName, ParalegalXref, ParalegalLastName, ParalegalFirstName, AccountManagerXref, AccountManagerLastName, AccountManagerFirstName, SpecialInstructionFlag, SpecialInstructionInfo, ClientBillingCode, OnlineIntakeDate, questionnairesenttomanager, questionnairessenttofn, followupwithfnforrequestedinformation, questionnairecompletedandreturnedbymanager, questionnairecompletedandreturnedbyfn, employersubmissionquestionnairecompleted, allpetitioningcompanyinforeceived, allfndocsreceived, fncompletedquestionnairesandacknowledgement, fnquestionnairescompleted, lcafiled, lcacasenumber, lcacertified, formsanddocumentationprepped, formsanddocumentationsubmittedforsignature, signedformsandletterreceived, dateaosformssentforsignature, datesignedaosformsreceived, targetfiledate, applicationfiled, applicationfiledwithcis, petitionfiledwithcis, formi129filedwithcis, aosapplicationfiled, tnpacketsenttofnforpoeprocessing,appealmotionduedate, appealmotionfiled, consularinterviewdate, supplementalbriefdocsfiled, docketdatebalca, datewithdrawrequestsenttouscis, withdrawalrequestconfirmedbydoluscis, approvalpackagesent, h1bregistrationsubmitted, h1bregistrationresult, h1bcapregistrationselected, I907filedupgradedtopremprocessing, premiumprocessingfeereceivedfromfn, receipts, I485receiptdate, I485jportabilityreceiptdate, I131receiptdate, apreceiptnoticereceived, eadreceiptnoticereceived, petitioningjobtitle, petitioningjoblocation, permmemosenttoemployer, approvalofpermmemoreceived, employeeworkexperiencechartsent, employeeworkexperiencechartreceived, employmentverificationletterssenttoemployee, signedemploymentverificationlettersreceived, prevailingwagedeterminationrequestsubmittedtodol, prevailingwagedeterminationissuedbydol, recruitmentinstructionssenttocompany, joborderplacedwithswa, noticeoffilingposted, intranetnoticeoffilingposted, noticeoffilingremovedsigned, intranetnoticeoffilingremoved, _1stsundayadplaced, _2ndsundayadplaced, _1stadditionalrecruitmentstepplaced, _2ndadditionalrecruitmentstepplaced, _3rdadditionalrecruitmentstepplaced, datedcopiesofallrecruitmentreceived, completedevaluationquestionnairesandresumesreceived, recruitmentreportsenttocompany, recruitmentreportreceived, form9089senttofnandemployer, editstoform9089receivedfromfnandemployer, form9089submittedtodol, inputcallconducted, inputstatementreceived, casestrategyandlettersplansent, longlettersenttofn, shortletterssenttofn, numberoftotalapplicants, numberofnonusworkers, numberofphonescreensconducted, numberofmanagerinterviewsconducted))
                        cursor.execute("SELECT @@IDENTITY AS ID;")
                        case_id = cursor.fetchone()[0]
                        cursor.commit()
                    else:
                        case_id = results[0].CaseId
                        cursor.execute("UPDATE [dbo].[Case] SET CaseXref='{}', BeneficiaryId='{}', SourceCreatedDate='{}', CasePetitionName='{}', CaseType='{}', CaseDescription='{}', CaseFiledDate='{}', ReceiptNumber='{}', ReceiptStatus='{}', RFEAuditReceivedDate='{}', RFEAuditDueDate='{}', RFEAuditSubmittedDate='{}', PrimaryCaseStatus='{}', SecondaryCaseStatus='{}', CaseComments='{}', LastStepCompleted='{}', LastStepCompletedDate='{}', NextStepAction='{}', NextStepActionDueDate='{}', PriorityDate='{}', PriorityCategory='{}', PriorityCountry='{}', CaseApprovedDate='{}', CaseValidFromDate='{}', CaseExpirationDate='{}', CaseClosedDate='{}', CaseDeniedDate='{}', CaseWithdrawnDate='{}', CasePrimaryAttorney='{}', CaseReviewingAttorney='{}', CasePrimaryCaseManager='{}', PetitionXref='{}',IsCurrentProcess='{}', CurrentProcessName='{}' ,RFEDocsReqestedDate='{}',  RFEDocsReceivedDate='{}',  SecondaryCaseStatusDate='{}',  DaysSinceLastStepCompleted='{}',  visa_preference_category='{}',  visa_priority_country='{}',  PartnerXref='{}',  PartnerLastName='{}',  PartnerFirstName='{}',  AssociateXref='{}',  AssociateLastName='{}',  AssociateFirstName='{}',  SupervisoryParalegalXref='{}',  SupervisoryParalegalLastName='{}',  SupervisoryParalegalFirstName='{}',  ParalegalXref='{}',  ParalegalLastName='{}',  ParalegalFirstName='{}',  AccountManagerXref='{}',  AccountManagerLastName='{}',  AccountManagerFirstName='{}',  SpecialInstructionFlag='{}',  SpecialInstructionInfo='{}',  ClientBillingCode='{}',  OnlineIntakeDate='{}',  questionnairesenttomanager='{}',  questionnairessenttofn='{}',  followupwithfnforrequestedinformation='{}',  questionnairecompletedandreturnedbymanager='{}',  questionnairecompletedandreturnedbyfn='{}',  employersubmissionquestionnairecompleted='{}',  allpetitioningcompanyinforeceived='{}',  allfndocsreceived='{}',  fncompletedquestionnairesandacknowledgement='{}',  fnquestionnairescompleted='{}',  lcafiled='{}',  lcacasenumber='{}',  lcacertified='{}',  formsanddocumentationprepped='{}',  formsanddocumentationsubmittedforsignature='{}',  signedformsandletterreceived='{}',  dateaosformssentforsignature='{}',  datesignedaosformsreceived='{}',  targetfiledate='{}',  applicationfiled='{}',  applicationfiledwithcis='{}',  petitionfiledwithcis='{}',  formi129filedwithcis='{}',  aosapplicationfiled='{}',  tnpacketsenttofnforpoeprocessing='{}',  appealmotionduedate='{}',  appealmotionfiled='{}',  consularinterviewdate='{}',  supplementalbriefdocsfiled='{}',  docketdatebalca='{}',  datewithdrawrequestsenttouscis='{}',  withdrawalrequestconfirmedbydoluscis='{}',  approvalpackagesent='{}',  h1bregistrationsubmitted='{}',  h1bregistrationresult='{}',  h1bcapregistrationselected='{}',  I907filedupgradedtopremprocessing='{}',  premiumprocessingfeereceivedfromfn='{}',  receipts='{}',  I485receiptdate='{}',  I485jportabilityreceiptdate='{}',  I131receiptdate='{}',  apreceiptnoticereceived='{}',  eadreceiptnoticereceived='{}',  petitioningjobtitle='{}',  petitioningjoblocation='{}',  permmemosenttoemployer='{}',  approvalofpermmemoreceived='{}',  employeeworkexperiencechartsent='{}',  employeeworkexperiencechartreceived='{}',  employmentverificationletterssenttoemployee='{}',  signedemploymentverificationlettersreceived='{}',  prevailingwagedeterminationrequestsubmittedtodol='{}',  prevailingwagedeterminationissuedbydol='{}',  recruitmentinstructionssenttocompany='{}',  joborderplacedwithswa='{}',  noticeoffilingposted='{}',  intranetnoticeoffilingposted='{}',  noticeoffilingremovedsigned='{}',  intranetnoticeoffilingremoved='{}',  _1stsundayadplaced='{}',  _2ndsundayadplaced='{}',  _1stadditionalrecruitmentstepplaced='{}',  _2ndadditionalrecruitmentstepplaced='{}',  _3rdadditionalrecruitmentstepplaced='{}',  datedcopiesofallrecruitmentreceived='{}',  completedevaluationquestionnairesandresumesreceived='{}',  recruitmentreportsenttocompany='{}',  recruitmentreportreceived='{}',  form9089senttofnandemployer='{}',  editstoform9089receivedfromfnandemployer='{}',  form9089submittedtodol='{}',  inputcallconducted='{}',  inputstatementreceived='{}',  casestrategyandlettersplansent='{}',  longlettersenttofn='{}',  shortletterssenttofn='{}',  numberoftotalapplicants='{}',  numberofnonusworkers='{}',  numberofphonescreensconducted='{}',  numberofmanagerinterviewsconducted='{}' WHERE CaseId='{}'".format(case_xref, beneficiary_id, case_creation_date, case_petition_name, case_type, case_description, case_filed_date, case_receipt_number, case_receipt_status, rfe_audit_received_date, rfe_audit_due_date, rfe_audit_submitted_date, primary_case_status, secondary_case_status, case_comments, case_last_step_completed, case_last_step_completed_date, case_next_step_to_be_completed, case_next_step_to_be_completed_date, case_priority_date, case_priority_category, case_priority_country, case_approved_date, case_valid_from, case_valid_to, case_closed_date, case_denied_date, case_withdrawn_date, case_primary_attorney, case_reviewing_attorney, case_primary_case_manager, petition_xref,  current_process, current_process_name, RFEDocsReqestedDate, RFEDocsReceivedDate, SecondaryCaseStatusDate, DaysSinceLastStepCompleted, visa_preference_category, visa_priority_country, PartnerXref, PartnerLastName, PartnerFirstName, AssociateXref, AssociateLastName, AssociateFirstName, SupervisoryParalegalXref, SupervisoryParalegalLastName, SupervisoryParalegalFirstName, ParalegalXref, ParalegalLastName, ParalegalFirstName, AccountManagerXref, AccountManagerLastName, AccountManagerFirstName, SpecialInstructionFlag, SpecialInstructionInfo, ClientBillingCode, OnlineIntakeDate, questionnairesenttomanager, questionnairessenttofn, followupwithfnforrequestedinformation, questionnairecompletedandreturnedbymanager, questionnairecompletedandreturnedbyfn, employersubmissionquestionnairecompleted, allpetitioningcompanyinforeceived, allfndocsreceived, fncompletedquestionnairesandacknowledgement, fnquestionnairescompleted, lcafiled, lcacasenumber, lcacertified, formsanddocumentationprepped, formsanddocumentationsubmittedforsignature, signedformsandletterreceived, dateaosformssentforsignature, datesignedaosformsreceived, targetfiledate, applicationfiled, applicationfiledwithcis, petitionfiledwithcis, formi129filedwithcis, aosapplicationfiled, tnpacketsenttofnforpoeprocessing,appealmotionduedate, appealmotionfiled, consularinterviewdate, supplementalbriefdocsfiled, docketdatebalca, datewithdrawrequestsenttouscis, withdrawalrequestconfirmedbydoluscis, approvalpackagesent, h1bregistrationsubmitted, h1bregistrationresult, h1bcapregistrationselected, I907filedupgradedtopremprocessing, premiumprocessingfeereceivedfromfn, receipts, I485receiptdate, I485jportabilityreceiptdate, I131receiptdate, apreceiptnoticereceived, eadreceiptnoticereceived, petitioningjobtitle, petitioningjoblocation, permmemosenttoemployer, approvalofpermmemoreceived, employeeworkexperiencechartsent, employeeworkexperiencechartreceived, employmentverificationletterssenttoemployee, signedemploymentverificationlettersreceived, prevailingwagedeterminationrequestsubmittedtodol, prevailingwagedeterminationissuedbydol, recruitmentinstructionssenttocompany, joborderplacedwithswa, noticeoffilingposted, intranetnoticeoffilingposted, noticeoffilingremovedsigned, intranetnoticeoffilingremoved, _1stsundayadplaced, _2ndsundayadplaced, _1stadditionalrecruitmentstepplaced, _2ndadditionalrecruitmentstepplaced, _3rdadditionalrecruitmentstepplaced, datedcopiesofallrecruitmentreceived, completedevaluationquestionnairesandresumesreceived, recruitmentreportsenttocompany, recruitmentreportreceived, form9089senttofnandemployer, editstoform9089receivedfromfnandemployer, form9089submittedtodol, inputcallconducted, inputstatementreceived, casestrategyandlettersplansent, longlettersenttofn, shortletterssenttofn, numberoftotalapplicants, numberofnonusworkers, numberofphonescreensconducted, numberofmanagerinterviewsconducted, case_id))
                        cursor.commit()




def ActiveEmployeeReport(result_filepath):
    ###################################### Tab 1 Header #############################################
    #Tab 1 - Active Employee Report
    headers = ['Beneficiary Full Name', 'Management Info Employee ID', 'Management Info Department','Management Info Business Unit Code', 'Management Info Dept Group', 'Management Info Job Start Date', 'Birth Country', 'Citizenship', 'Petitioning Job Title', 'Petitioning Job Location', 'Current Status', 'Current Status Expires', 'I-797 Expires', 'I94 Expires','EAD Expiration', 'AP Expiration', 'Management Info Manager', 'Management Info Second Level Manager', 'NIV Max Out Date', 'Visa Priority Date']

     
    headers_table = ['FullName', 'EmployeeId', 'Department','Business_Unit_Code', 'Department_Group', 'EmploymentStartDate', 'BirthCountry', 'CitizenshipCountry', 'petitioningjobtitle','petitioningjoblocation', 'Current_Immigration_Status', 'CurrentImmigrationStatusExpirationDate2', 'I797ExpirationDate', 'ImmigrationStatusExpirationDate', 'EadExpirationDate','AdvanceParoleExpirationDate', 'ManagerName', 'SecondLevelManager', 'FinalNivDate', 'Priority1Date']
    
    date_columns = ['EmploymentStartDate', 'CurrentImmigrationStatusExpirationDate2', 'I797ExpirationDate', 'ImmigrationStatusExpirationDate', 'EadExpirationDate','AdvanceParoleExpirationDate', 'FinalNivDate', 'Priority1Date']
    
    header_names = [{'header': x} for x in headers]
    
    results_active_qry ="""SELECT b.*,
        CASE WHEN b.IsActive=1 THEN 'Active' ELSE 'Retired' END as BeneficiaryRecordStatus, 
        e.EmployeeId, e.Department, e.Business_Unit_Code, e.Department_Group, e.EmploymentStartDate, e.ManagerName, e.SecondLevelManager,
        c.petitioningjobtitle, c.petitioningjoblocation, bp.Priority1Date 
        FROM dbo.Beneficiary as b 
        LEFT JOIN dbo.[Case] as c on c.BeneficiaryId=b.BeneficiaryId
		AND c.SourceCreatedDate = (
		  SELECT
			MAX(SourceCreatedDate)
		  FROM
			dbo.[Case] AS c2
		  WHERE
			c2.BeneficiaryId = c.BeneficiaryId
		)
        LEFT JOIN dbo.BeneficiaryEmployment as e on e.BeneficiaryId=b.BeneficiaryId\
        LEFT JOIN dbo.BeneficiaryPriorityDate as bp on bp.BeneficiaryId=b.BeneficiaryId\
        where b.IsActive = '1' \
        ORDER BY b.FullName ASC"""
    
    results_active = cursor.execute(results_active_qry).fetchall()
    
    df = pd.read_sql(results_active_qry, conn)
    #print(df)
    for dfcol in df.columns:
        if dfcol not in headers_table:
            df.drop(dfcol, axis=1, inplace=True)
    
    # altering the DataFrame - Column order
    df = df[headers_table]
    
    for d_h in date_columns:
        if d_h in df:
            if "1900-01-01" in df[d_h]:
                df[d_h] = ""
            else:
                df[d_h] = pd.to_datetime(df[d_h], format='%Y-%m-%d', errors='coerce').dt.date
    df.columns = headers #changing dataframe all column names
    writer = pd.ExcelWriter(result_filepath, engine='xlsxwriter', date_format='m/d/yyyy')
    df.to_excel(writer, "Active Employee Report", startrow=0, columns=headers, index=False)
    writer.save()
    
    wb_pyxl = load_workbook(result_filepath)  
    wb_pyxl.active = 0 #active first sheet
    sheet = wb_pyxl.active
    for hdr in headers_table:
        col = headers_table.index(hdr)
        sheet.cell(row=1, column=col+1).alignment=Alignment(wrap_text=True, horizontal="justify", vertical="justify")
        sheet.column_dimensions[get_column_letter(col+1)].width = 14
    
    if sheet.max_row > 1:
        table = Table(displayName="Table1", ref="A1:" + get_column_letter(sheet.max_column) + str(sheet.max_row))
    else:
        table = Table(displayName="Table1", ref="A1:" + get_column_letter(sheet.max_column) + str(2))
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    sheet.add_table(table)
    
    if len(results_active) > 0:
        for _key, s in enumerate(results_active):
            num = _key+1
            
            for hdr in headers_table:
                col = headers_table.index(hdr)
                
                if hdr:
                    value_obj = str(getattr(s, hdr))
                else: 
                    value_obj = ''
                
                if hdr in date_columns:
                    ##print(value_obj)
                    value_obj = str(value_obj)
                    if str(value_obj) == '1900-01-01 00:00:00':
                        value_obj = ''
                        sheet.cell(row=int(num)+1, column=col+1).value = ''
                    else:
                        value_obj = change_display_format(str(value_obj).replace('00:00:00', ''))

                    if value_obj is None or value_obj == "None" or value_obj == "nan":
                        value_obj = ''

                    #sheet.cell(row=int(num) + 1, column=col + 1).value = str(value_obj)
                    sheet.cell(row=int(num) + 1, column=col + 1).alignment = Alignment(wrap_text=True, horizontal="justify",
                                                                                       vertical="justify")
                    sheet.cell(row=int(num) + 1, column=col + 1).number_format = 'mm/dd/yyyy'
                    sheet.cell(row=int(num) + 1, column=col + 1).font = Font(name='Calibri (Body)', size=11)
                    sheet.cell(row=int(num) + 1, column=col + 1).border = Border(left=Side(style='thin'),
                                                                                 right=Side(style='thin'),
                                                                                 top=Side(style='thin'),
                                                                                 bottom=Side(style='thin'))
                    

                else:
                    
                    if value_obj is None or value_obj == "None" or value_obj == "nan":
                        value_obj = ''

                    sheet.cell(row=int(num)+1, column = int(col)+1).value = str(value_obj) 
                    sheet.cell(row=int(num)+1, column=col+1).alignment=Alignment(wrap_text=True, horizontal="justify", vertical="justify")
                    sheet.cell(row=int(num)+1, column=col+1).font= Font(name = 'Calibri (Body)', size= 11)
                    sheet.cell(row=int(num)+1, column=col+1).border= Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    
                    pass
                
            #return False
        else:
            sheet.cell(row=2, column = 1).value = "No Records Found"
        
    wb_pyxl.save(result_filepath)
    

def PermReport(result_filepath2):
    ###################################### Tab 1 Header #############################################
    # Tab 1 - PERM Report
    
    headers = ["Beneficiary Full Name","Management Info Department","Management Info Business Unit Code","Management Info Dept Group","Petitioning Job Title","Petitioning Job Location","Paralegal","Birth Country","Citizenship","Current Status","NIV Max Out Date","Visa Priority Date","Date Opened","Online Intake Date","Questionnaire sent to manager","Questionnaires Sent to FN","Questionnaire completed and returned by Manager","Questionnaire completed and returned by FN","PERM Memo sent to employer","Approval of PERM Memo received","Employee Work Experience Chart sent","Employee Work Experience Chart received","Prevailing Wage Determination request submitted to DOL","Employment Verification Letters sent to employee","Prevailing Wage Determination issued by DOL","PWD expiration date","Recruitment instructions sent to company","1st Additional Recruitment Step Placed","Dated copies of all recruitment received","Completed evaluation questionnaires and resumes received","Recruitment report sent to company","Recruitment Report Received","Form 9089 sent to FN and Employer","Edits to Form 9089 received from FN and Employer","Form 9089 submitted to DOL","Date Closed","Days Since Last Activity"]

    headers_table = ["FullName","Department","BusinessUnitCode","Department_Group","petitioningjobtitle","petitioningjoblocation","ParalegalXref","BirthCountry","CitizenshipCountry","Current_Immigration_Status","FinalNivDate","Priority1Date","SourceCreatedDate","OnlineIntakeDate","questionnairesenttomanager","questionnairessenttofn","questionnairecompletedandreturnedbymanager","questionnairecompletedandreturnedbyfn","permmemosenttoemployer","approvalofpermmemoreceived","employeeworkexperiencechartsent","employeeworkexperiencechartreceived","prevailingwagedeterminationrequestsubmittedtodol","employmentverificationletterssenttoemployee","prevailingwagedeterminationissuedbydol","PWDExpirationDate","recruitmentinstructionssenttocompany","_1stadditionalrecruitmentstepplaced","datedcopiesofallrecruitmentreceived","completedevaluationquestionnairesandresumesreceived","recruitmentreportsenttocompany","recruitmentreportreceived","form9089senttofnandemployer","editstoform9089receivedfromfnandemployer","form9089submittedtodol","CaseClosedDate","DaysSinceLastStepCompleted"]
    
    date_columns = ["FinalNivDate","Priority1Date","SourceCreatedDate","OnlineIntakeDate","questionnairesenttomanager","questionnairessenttofn","questionnairecompletedandreturnedbymanager","questionnairecompletedandreturnedbyfn","permmemosenttoemployer","approvalofpermmemoreceived","employeeworkexperiencechartsent","employeeworkexperiencechartreceived","prevailingwagedeterminationrequestsubmittedtodol","employmentverificationletterssenttoemployee","prevailingwagedeterminationissuedbydol","PWDExpirationDate","recruitmentinstructionssenttocompany","_1stadditionalrecruitmentstepplaced","datedcopiesofallrecruitmentreceived","completedevaluationquestionnairesandresumesreceived","recruitmentreportsenttocompany","recruitmentreportreceived","form9089senttofnandemployer","editstoform9089receivedfromfnandemployer","form9089submittedtodol","CaseClosedDate","DaysSinceLastStepCompleted"]

    header_names = [{'header': x} for x in headers]
    
    results_active_qry = """
        SELECT b.FullName,be.Department,be.BusinessUnitCode,be.Department_Group,c.petitioningjobtitle,c.petitioningjoblocation,c.ParalegalXref,b.BirthCountry,b.CitizenshipCountry,b.Current_Immigration_Status,b.FinalNivDate,bp.Priority1Date,c.SourceCreatedDate,c.OnlineIntakeDate,c.questionnairesenttomanager,c.questionnairessenttofn,c.questionnairecompletedandreturnedbymanager,c.questionnairecompletedandreturnedbyfn,c.permmemosenttoemployer,c.approvalofpermmemoreceived,c.employeeworkexperiencechartsent,c.employeeworkexperiencechartreceived,c.prevailingwagedeterminationrequestsubmittedtodol,c.employmentverificationletterssenttoemployee,c.prevailingwagedeterminationissuedbydol, '' as PWDExpirationDate,c.recruitmentinstructionssenttocompany,c._1stadditionalrecruitmentstepplaced,c.datedcopiesofallrecruitmentreceived,c.completedevaluationquestionnairesandresumesreceived,c.recruitmentreportsenttocompany,c.recruitmentreportreceived,c.form9089senttofnandemployer,c.editstoform9089receivedfromfnandemployer,c.form9089submittedtodol,c.CaseClosedDate,c.DaysSinceLastStepCompleted,
        CASE WHEN b.IsActive=1 THEN 'Active' ELSE 'Retired' END as BeneficiaryRecordStatus
        FROM dbo.[Case] as c  
        LEFT JOIN  dbo.Beneficiary as b on c.BeneficiaryId=b.BeneficiaryId
        LEFT JOIN dbo.BeneficiaryEmployment as be on be.BeneficiaryId=b.BeneficiaryId
        LEFT JOIN dbo.BeneficiaryPriorityDate as bp on bp.BeneficiaryId=b.BeneficiaryId
        where b.IsActive = '1' and c.PrimaryCaseStatus='Open' and c.PetitionXref='100003008'
        ORDER BY b.FullName ASC"""
    
    results_active = cursor.execute(results_active_qry).fetchall()
    
    df = pd.read_sql(results_active_qry, conn)
    for dfcol in df.columns:
        if dfcol not in headers_table:
            df.drop(dfcol, axis=1, inplace=True)
    
    # altering the DataFrame - Column order
    df = df[headers_table]
    for d_h in date_columns:
        if d_h in df:
            if "1900-01-01" in df[d_h]:
                df[d_h] = ""
            else:
                df[d_h] = pd.to_datetime(df[d_h], format='%Y-%m-%d', errors='coerce').dt.date
    df.columns = headers #changing dataframe all column names
    writer = pd.ExcelWriter(result_filepath2, engine='xlsxwriter', date_format='m/d/yyyy')
    df.to_excel(writer, "PERM Report", startrow=0, columns=headers, index=False)
    writer.save()
    
    wb_pyxl = load_workbook(result_filepath2)  
    wb_pyxl.active = 0 #active first sheet
    sheet = wb_pyxl.active
    for hdr in headers_table:
        col = headers_table.index(hdr)
        sheet.cell(row=1, column=col+1).alignment=Alignment(wrap_text=True, horizontal="justify", vertical="justify")
        sheet.column_dimensions[get_column_letter(col+1)].width = 14
    
    if sheet.max_row > 1:
        table = Table(displayName="Table2", ref="A1:" + get_column_letter(sheet.max_column) + str(sheet.max_row))
    else:
        table = Table(displayName="Table2", ref="A1:" + get_column_letter(sheet.max_column) + str(2))
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    sheet.add_table(table)
    
    if len(results_active) > 0:
        for _key, s in enumerate(results_active):
            num = _key + 1

            for hdr in headers_table:
                col = headers_table.index(hdr)

                if hdr:
                    value_obj = str(getattr(s, hdr))
                else:
                    value_obj = ''

                if hdr in date_columns:
                    ##print(value_obj)
                    value_obj = str(value_obj)
                    if str(value_obj) == '1900-01-01 00:00:00':
                        value_obj = ''
                        sheet.cell(row=int(num)+1, column=col+1).value = ''
                    else:
                        value_obj = change_display_format(str(value_obj).replace('00:00:00', ''))

                    if value_obj is None or value_obj == "None" or value_obj == "nan":
                        value_obj = ''

                    #sheet.cell(row=int(num) + 1, column=col + 1).value = str(value_obj)
                    sheet.cell(row=int(num) + 1, column=col + 1).alignment = Alignment(wrap_text=True, horizontal="justify",
                                                                                       vertical="justify")
                    sheet.cell(row=int(num) + 1, column=col + 1).number_format = 'mm/dd/yyyy'
                    sheet.cell(row=int(num) + 1, column=col + 1).font = Font(name='Calibri (Body)', size=11)
                    sheet.cell(row=int(num) + 1, column=col + 1).border = Border(left=Side(style='thin'),
                                                                                 right=Side(style='thin'),
                                                                                 top=Side(style='thin'),
                                                                                 bottom=Side(style='thin'))


                else:

                    if value_obj is None or value_obj == "None" or value_obj == "nan":
                        value_obj = ''

                    if str(value_obj) == '1900-01-01 00:00:00':
                        value_obj = ''

                    sheet.cell(row=int(num) + 1, column=int(col) + 1).value = str(value_obj)
                    sheet.cell(row=int(num) + 1, column=col + 1).alignment = Alignment(wrap_text=True, horizontal="justify",
                                                                                       vertical="justify")
                    sheet.cell(row=int(num) + 1, column=col + 1).font = Font(name='Calibri (Body)', size=11)
                    sheet.cell(row=int(num) + 1, column=col + 1).border = Border(left=Side(style='thin'),
                                                                                 right=Side(style='thin'),
                                                                                 top=Side(style='thin'),
                                                                                 bottom=Side(style='thin'))

                    pass

            # return False
    else:
        sheet.cell(row=2, column = 1).value = "No Records Found"
    wb_pyxl.save(result_filepath2)


def  WeeklyNetOps(result_filepath3):
    ###################################### Tab 1 Header #############################################
    # Tab 1 - Weekly NetOps - P&T Transfer Report
    

    headers = ["Beneficiary Full Name","Management Info Employee ID","Management Info Department","Management Info Business Unit Code","Management Info Dept Group","Management Info Job Start Date","Birth Country","Citizenship","Petitioning Job Title","Petitioning Job Location","Current Status","Current Status Expires","I-797 Expires","I94 Expires","EAD Expiration","AP Expiration","Management Info Manager","Management Info Second Level Manager","NIV Max Out Date","Visa Priority Date","Special Instruction Flag","Case Opened","Process Type","Process Reference","Questionnaires Sent to FN","FN Completed Questionnaires and Acknowledgement","All FN Docs Received","LCA Filed","LCA Certified","Forms and Documentation Submitted for Signature","Application Filed","Petition Filed with CIS","Receipt Date","RFE Received","RFE Due Date","RFE Docs Received","RFE Response Submitted","Final Action","Date Status Valid From","Date Status Valid To","Case Closed","Days Since Last Activity"]

    headers_table = ["FullName","EmployeeId","Department","BusinessUnitCode","Department_Group","EmploymentStartDate","BirthCountry","CitizenshipCountry","petitioningjobtitle","petitioningjoblocation","Current_Immigration_Status","CurrentImmigrationStatusExpirationDate2","I797ExpirationDate","ImmigrationStatusExpirationDate","EadExpirationDate","AdvanceParoleExpirationDate","ManagerName","SecondLevelManager","FinalNivDate","Priority1Date","SpecialInstructionFlag","SourceCreatedDate","CasePetitionName","CaseType","questionnairessenttofn","fncompletedquestionnairesandacknowledgement","allfndocsreceived","lcafiled","lcacertified","formsanddocumentationsubmittedforsignature","applicationfiled","petitionfiledwithcis","CaseReceivedDate","RFEAuditReceivedDate","RFEAuditDueDate","RFEDocsReceivedDate","RFEAuditSubmittedDate","SecondaryCaseStatus","CaseValidFromDate","CaseExpirationDate","CaseClosedDate","DaysSinceLastStepCompleted"]

    date_columns = ["EmploymentStartDate","CurrentImmigrationStatusExpirationDate2","I797ExpirationDate","ImmigrationStatusExpirationDate","EadExpirationDate","AdvanceParoleExpirationDate","FinalNivDate","Priority1Date","SourceCreatedDate","questionnairessenttofn","fncompletedquestionnairesandacknowledgement","allfndocsreceived","lcafiled","lcacertified","formsanddocumentationsubmittedforsignature","applicationfiled","petitionfiledwithcis","CaseReceivedDate","RFEAuditReceivedDate","RFEAuditDueDate","RFEDocsReceivedDate","RFEAuditSubmittedDate","CaseValidFromDate","CaseExpirationDate","CaseClosedDate","DaysSinceLastStepCompleted"]


    header_names = [{'header': x} for x in headers]
    
    results_active_qry = """ SELECT 
        CASE WHEN b.IsActive=1 THEN 'Active' ELSE 'Retired' END as BeneficiaryRecordStatus, 
        b.FullName,be.EmployeeId,be.Department,be.BusinessUnitCode,be.Department_Group,be.EmploymentStartDate,b.BirthCountry,b.CitizenshipCountry,c.petitioningjobtitle,c.petitioningjoblocation,b.Current_Immigration_Status,b.CurrentImmigrationStatusExpirationDate2,b.I797ExpirationDate,b.ImmigrationStatusExpirationDate,b.EadExpirationDate,b.AdvanceParoleExpirationDate,be.ManagerName,be.SecondLevelManager,b.FinalNivDate,bp.Priority1Date,c.SpecialInstructionFlag,c.SourceCreatedDate,c.CasePetitionName,c.CaseType,c.questionnairessenttofn,c.fncompletedquestionnairesandacknowledgement,c.allfndocsreceived,c.lcafiled,c.lcacertified,c.formsanddocumentationsubmittedforsignature,c.applicationfiled,c.petitionfiledwithcis,c.CaseReceivedDate,c.RFEAuditReceivedDate,c.RFEAuditDueDate,c.RFEDocsReceivedDate,c.RFEAuditSubmittedDate,c.SecondaryCaseStatus,c.CaseValidFromDate,c.CaseExpirationDate,c.CaseClosedDate,c.DaysSinceLastStepCompleted 
        FROM dbo.[Case] as c  
        LEFT JOIN  dbo.Beneficiary as b on c.BeneficiaryId=b.BeneficiaryId
        LEFT JOIN dbo.BeneficiaryEmployment as be on be.BeneficiaryId=b.BeneficiaryId
        LEFT JOIN dbo.BeneficiaryPriorityDate as bp on bp.BeneficiaryId=b.BeneficiaryId
        where b.IsActive = '1' and c.PrimaryCaseStatus='Open' and (c.CaseType='Change of Employer' or c.CaseType='New Hire Assessment')
        ORDER BY b.FullName ASC """
        
    results_active = cursor.execute(results_active_qry).fetchall()
    
    df = pd.read_sql(results_active_qry, conn)
    for dfcol in df.columns:
        if dfcol not in headers_table:
            df.drop(dfcol, axis=1, inplace=True)
    
    # altering the DataFrame - Column order
    df = df[headers_table]
    for d_h in date_columns:
        if d_h in df:
            if "1900-01-01" in df[d_h]:
                df[d_h] = ""
            else:
                df[d_h] = pd.to_datetime(df[d_h], format='%Y-%m-%d', errors='coerce').dt.date
    df.columns = headers #changing dataframe all column names
    writer = pd.ExcelWriter(result_filepath3, engine='xlsxwriter', date_format='m/d/yyyy')
    df.to_excel(writer, "Weekly NetOps-P&T", startrow=0, columns=headers, index=False)
    writer.save()
    
    wb_pyxl = load_workbook(result_filepath3)  
    wb_pyxl.active = 0 #active first sheet
    sheet = wb_pyxl.active
    for hdr in headers_table:
        col = headers_table.index(hdr)
        sheet.cell(row=1, column=col+1).alignment=Alignment(wrap_text=True, horizontal="justify", vertical="justify")
        sheet.column_dimensions[get_column_letter(col+1)].width = 14
    
    if sheet.max_row > 1:
        table = Table(displayName="Table3", ref="A1:" + get_column_letter(sheet.max_column) + str(sheet.max_row))
    else:
        table = Table(displayName="Table3", ref="A1:" + get_column_letter(sheet.max_column) + str(2))
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    sheet.add_table(table)
    
    if len(results_active) > 0:
        for _key, s in enumerate(results_active):
            num = _key + 1

            for hdr in headers_table:
                col = headers_table.index(hdr)

                if hdr:
                    value_obj = str(getattr(s, hdr))
                else:
                    value_obj = ''

                if hdr in date_columns:
                    ##print(value_obj)
                    value_obj = str(value_obj)
                    if str(value_obj) == '1900-01-01 00:00:00':
                        value_obj = ''
                        sheet.cell(row=int(num)+1, column=col+1).value = ''
                    else:
                        value_obj = change_display_format(str(value_obj).replace('00:00:00', ''))

                    if value_obj is None or value_obj == "None" or value_obj == "nan":
                        value_obj = ''

                    #sheet.cell(row=int(num) + 1, column=col + 1).value = str(value_obj)
                    sheet.cell(row=int(num) + 1, column=col + 1).alignment = Alignment(wrap_text=True, horizontal="justify",
                                                                                       vertical="justify")
                    sheet.cell(row=int(num) + 1, column=col + 1).number_format = 'mm/dd/yyyy'
                    sheet.cell(row=int(num) + 1, column=col + 1).font = Font(name='Calibri (Body)', size=11)
                    sheet.cell(row=int(num) + 1, column=col + 1).border = Border(left=Side(style='thin'),
                                                                                 right=Side(style='thin'),
                                                                                 top=Side(style='thin'),
                                                                                 bottom=Side(style='thin'))


                else:

                    if value_obj is None or value_obj == "None" or value_obj == "nan":
                        value_obj = ''

                    if str(value_obj) == '1900-01-01 00:00:00':
                        value_obj = ''

                    sheet.cell(row=int(num) + 1, column=int(col) + 1).value = str(value_obj)
                    sheet.cell(row=int(num) + 1, column=col + 1).alignment = Alignment(wrap_text=True, horizontal="justify",
                                                                                       vertical="justify")
                    sheet.cell(row=int(num) + 1, column=col + 1).font = Font(name='Calibri (Body)', size=11)
                    sheet.cell(row=int(num) + 1, column=col + 1).border = Border(left=Side(style='thin'),
                                                                                 right=Side(style='thin'),
                                                                                 top=Side(style='thin'),
                                                                                 bottom=Side(style='thin'))

                    pass

            # return False
    else:
        sheet.cell(row=2, column = 1).value = "No Records Found"
    wb_pyxl.save(result_filepath3)


if __name__ == '__main__':
    start()
    print('Finished')
    pass
    
    
    

