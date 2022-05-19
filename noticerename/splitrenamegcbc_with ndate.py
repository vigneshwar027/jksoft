from email.mime import base
import glob, os, sys
from posixpath import basename
from PyPDF2 import PdfFileReader, PdfFileWriter, PdfFileMerger
from PIL import Image
import shutil
from os import listdir
from os.path import isfile, join
import fitz
import re
import json
import requests
from base64 import b64encode
import shutil
import subprocess
import time
from datetime import date, datetime, timedelta
from time import strptime
from dateutil import parser
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import smtplib, ssl
import cv2
from pyzbar.pyzbar import decode
import pandas as pd
import warnings

pd.options.mode.chained_assignment = None
warnings.filterwarnings("ignore")

port = 465  # For SSL
smtp_server = "smtp.gmail.com"
sender_email = "processautomation@immilytics.com"  
#receiver_email = "marie@gip-us.com,rah@gip-us.com,jana@immilytics.com,processautomation@immilytics.com"   #use comma , separator to send multiple emailids
receiver_email = "jana@immilytics.com"
password = "iRPA@2020!"


ENDPOINT_URL = 'https://vision.googleapis.com/v1/images:annotate'
api_key = 'AIzaSyBqww2vOLc3thM38ikV0x_ExZTjiuX-mYg'


#fdname="P:/Scans/Notice Rename Tool - Daily Scans/"
fdname="F:/PROJECTS/JK_SOFT/noticerename/"


ingestion = fdname+"/ingestion/"
#ingestion = fdname+"/I-129 approvals - to be processed/"


os.chdir(fdname+"/ingestion/")
#os.chdir(fdname+"/I-129 approvals - to be processed/")
dest=fdname+"/processing/"

start_time = ''

start_time  = datetime.today().strftime("%H:%M")

end_time = ''
processed_files = []
m=0


def make_image_data_list(image_filenames):
    try:
        img_requests = []
        
        with open(image_filenames, 'rb') as f:
            ctxt = b64encode(f.read()).decode()
            img_requests.append({
                    'image': {'content': ctxt},
                    'features': [{
                        'type': 'DOCUMENT_TEXT_DETECTION',
                        'maxResults': 1
                    }]
            })
        return img_requests
    except Exception as e:
        error_log(e, 'Failed', 'Something Went Wrong. Process Stopped','')
        pass

def make_image_data(image_filenames):
    try:
        imgdict = make_image_data_list(image_filenames)
        return json.dumps({"requests": imgdict }).encode()
    except Exception as e:
        error_log(e, 'Failed', 'Something Went Wrong. Process Stopped','')
        pass


def request_ocr(api_key, image_filenames):
    try:
        response = requests.post(ENDPOINT_URL,
                                data=make_image_data(image_filenames),
                                params={'key': api_key},
                                headers={'Content-Type': 'application/json'})
        return response    
    except Exception as e:
        error_log(e, 'Failed', 'Something Went Wrong. Process Stopped','')
        pass
        
def nd_log(fdname, sfname, dfname, status, comment,zoomdate,docid, wspace):
    try:
        today = date.today()
        year = today.year
        month = today.month
        day = today.day
        end_time  = datetime.now().strftime("%H:%M")
        today = str(month)+'/'+str(day)+'/'+str(year)+" "+end_time
        if not os.path.exists(fdname + 'I129log.xlsx'):
            book = Workbook()
            ws = book.active
            ws.cell(1,1).value = 'Filename'
            ws.cell(1,2).value = 'Renamed Filename'
            ws.cell(1,3).value = 'File Upload Status'
            ws.cell(1,4).value = 'Comment'
            ws.cell(1,5).value = 'Date Processed'
            ws.cell(1,6).value = 'Zoom Date Processed'
            ws.cell(1,7).value = 'Document ID'
            ws.cell(1,8).value = 'Workspace Name'
            
            ws.column_dimensions["A"].width = 50.0
            ws.column_dimensions["B"].width = 60.0
            ws.column_dimensions["C"].width = 40.0
            ws.column_dimensions["D"].width = 40.0
            ws.column_dimensions["E"].width = 10.0
            ws.column_dimensions["F"].width = 40.0
            ws.column_dimensions["G"].width = 40.0
            ws.column_dimensions["H"].width = 40.0
            
            ws['A1'].font = Font(bold=True)
            ws['B1'].font = Font(bold=True)
            ws['C1'].font = Font(bold=True)
            ws['D1'].font = Font(bold=True)
            ws['E1'].font = Font(bold=True)
            ws['F1'].font = Font(bold=True)
            ws['G1'].font = Font(bold=True)
            ws['H1'].font = Font(bold=True)
            
            book.save(fdname +'I129log.xlsx')
            
        book2 = load_workbook(fdname +'I129log.xlsx')
        sheet = book2.active                
        sheet.append((sfname,dfname,status,comment,today,zoomdate,docid,wspace))
        sheet.column_dimensions["A"].width = 50.0
        sheet.column_dimensions["B"].width = 60.0
        sheet.column_dimensions["C"].width = 40.0
        sheet.column_dimensions["D"].width = 40.0
        sheet.column_dimensions["E"].width = 10.0
        sheet.column_dimensions["F"].width = 40.0
        sheet.column_dimensions["G"].width = 40.0
        book2.save(fdname +'I129log.xlsx')
    except Exception as e:
        #print(e)
        pass
        
def temp_log(fdname1,fname,ndate1,ndurl,rno,frmdate,todate,fdname):
    try:
        today = date.today()
        year = today.year
        month = today.month
        day = today.day
        end_time  = datetime.now().strftime("%H:%M")
        today = str(month)+'/'+str(day)+'/'+str(year)+" "+end_time
        if not os.path.exists(fdname + 'templog.xlsx'):
            book = Workbook()
            ws = book.active
            ws.cell(1,1).value = 'Filename'            
            ws.cell(1,2).value = 'Zoom Date Processed'
            ws.cell(1,3).value = 'Notice Date'
            ws.cell(1,4).value = 'Workspace URL'
            ws.cell(1,5).value = 'Receipt Number'
            ws.cell(1,6).value = 'From Date'
            ws.cell(1,7).value = 'To Date'
            ws.cell(1,8).value = 'Folder Name'
            
            
            ws.column_dimensions["A"].width = 50.0
            ws.column_dimensions["B"].width = 20.0
            ws.column_dimensions["C"].width = 20.0
            ws.column_dimensions["D"].width = 60.0
            ws.column_dimensions["E"].width = 40.0
            ws.column_dimensions["F"].width = 20.0
            ws.column_dimensions["G"].width = 20.0
            ws.column_dimensions["H"].width = 30.0
            
            
            ws['A1'].font = Font(bold=True)
            ws['B1'].font = Font(bold=True)
            ws['C1'].font = Font(bold=True)
            ws['D1'].font = Font(bold=True)
            ws['E1'].font = Font(bold=True)
            ws['F1'].font = Font(bold=True)
            ws['G1'].font = Font(bold=True)
            ws['H1'].font = Font(bold=True)
            
            
            
            book.save(fdname +'templog.xlsx')
            
        book2 = load_workbook(fdname +'templog.xlsx')
        sheet = book2.active                
        sheet.append((fname,today,ndate1,ndurl,rno,frmdate,todate,fdname1))
        sheet.column_dimensions["A"].width = 50.0
        sheet.column_dimensions["B"].width = 20.0
        sheet.column_dimensions["C"].width = 20.0
        sheet.column_dimensions["D"].width = 60.0
        sheet.column_dimensions["E"].width = 40.0
        sheet.column_dimensions["F"].width = 40.0
        sheet.column_dimensions["G"].width = 40.0
        sheet.column_dimensions["H"].width = 40.0
        
        book2.save(fdname +'templog.xlsx')
    except Exception as e:
        #print(e)
        pass
    
def update_log(fdname, sfname, file ):
    try:
        shutil.move(fdname+'/processing/'+file, fdname+'/errors/'+sfname+"_"+file)
        today = date.today()
        year = today.year
        month = today.month
        day = today.day
        today = str(month)+'/'+str(day)+'/'+str(year)
        if not os.path.exists(fdname + 'nrtlog.xlsx'):
            book = Workbook()
            ws = book.active
            ws.cell(1,1).value = 'Source File Name'
            ws.cell(1,2).value = 'Renamed File Name'
            ws.cell(1,3).value = 'File Rename Status'
            ws.cell(1,4).value = 'Date File Renamed'
            ws.cell(1,5).value = 'Zoom Update Status'
            ws.cell(1,6).value = 'Date zoom updated'
            ws.cell(1,7).value = 'File Upload Status'
            ws.cell(1,8).value = 'Date File Uploaded to ND'
            ws.cell(1,9).value = 'Document ID'
            ws.cell(1,10).value = 'Workspace Name'
            ws.cell(1,11).value = 'Receipt Number'
            
            ws.column_dimensions["A"].width = 50.0
            ws.column_dimensions["B"].width = 60.0
            ws.column_dimensions["C"].width = 40.0
            ws.column_dimensions["D"].width = 10.0
            ws.column_dimensions["E"].width = 40.0
            ws.column_dimensions["F"].width = 40.0
            ws.column_dimensions["G"].width = 40.0
            ws.column_dimensions["H"].width = 40.0
            ws.column_dimensions["I"].width = 40.0
            ws.column_dimensions["J"].width = 40.0
            ws.column_dimensions["K"].width = 40.0
            
            ws['A1'].font = Font(bold=True)
            ws['B1'].font = Font(bold=True)
            ws['C1'].font = Font(bold=True)
            ws['D1'].font = Font(bold=True)
            ws['E1'].font = Font(bold=True)
            ws['F1'].font = Font(bold=True)
            ws['G1'].font = Font(bold=True)
            ws['H1'].font = Font(bold=True)
            ws['I1'].font = Font(bold=True)
            ws['J1'].font = Font(bold=True)
            ws['K1'].font = Font(bold=True)
            
            book.save(fdname +'nrtlog.xlsx')
            
        book2 = load_workbook(fdname +'nrtlog.xlsx')
        sheet = book2.active                
        sheet.append((sfname+"_"+file,'','Unable to merge split pages',today))
        sheet.column_dimensions["A"].width = 50.0
        sheet.column_dimensions["B"].width = 60.0
        sheet.column_dimensions["C"].width = 40.0
        sheet.column_dimensions["D"].width = 10.0
        sheet.column_dimensions["E"].width = 40.0
        sheet.column_dimensions["F"].width = 40.0
        sheet.column_dimensions["G"].width = 40.0
        sheet.column_dimensions["H"].width = 40.0
        sheet.column_dimensions["I"].width = 40.0
        sheet.column_dimensions["J"].width = 40.0
        sheet.column_dimensions["K"].width = 40.0
        book2.save(fdname +'nrtlog.xlsx')
    except Exception as e:
        error_log(e, 'Failed', 'Something Went Wrong. Process Stopped','')
        pass

def BarcodeReader(image):
    data = None 
    try: 
        # read the image in numpy array using cv2
        img = cv2.imread(image)
        
        # Decode the barcode image
        detectedBarcodes = decode(img)
        
        # If not detected then print the message
        if not detectedBarcodes:
            print("Barcode Not Detected or your barcode is blank/corrupted!")
            
        else:
            # Traverse through all the detected barcodes in image
            for barcode in detectedBarcodes: 
            
                # Locate the barcode position in image
                (x, y, w, h) = barcode.rect
                
                # Put the rectangle in image using
                # cv2 to heighlight the barcode
                cv2.rectangle(img, (x-10, y-10),
                            (x + w+10, y + h+10),
                            (255, 0, 0), 2)
                
                if barcode.data!="":
                
                # Print the barcode data
                    #print(barcode.data)
                    
                    #print(barcode.type)
                    data = (barcode.data).decode("utf-8") 
    except: 
        pass
    return data
                 

def initf():
    #start_time  = datetime.now().strftime("%H:%M")
    try:
        source_dir = fdname+"/processing/"
        target_dir = fdname+"/errors/"
        
        file_names = os.listdir(source_dir)
        
        for file_name in file_names:
            try:
                dest1 = get_next_file(file_name, target_dir)
                shutil.move(os.path.join(source_dir, file_name), dest1)
            except:
                pass

    except Exception as e:
        error_log(e, 'Failed', 'Something Went Wrong. Process Stopped','')
        pass

        
    split()    

def split():
    try:
        os.chdir(fdname+"/ingestion/")
        for file in glob.glob("*.pdf"):
            print(file)
            #source=fdname+"/I-129 approvals - to be processed/"
            source=ingestion
            doc = fitz.open(join(source, file))
            page = doc[0]
                        
            image_matrix = fitz.Matrix(2,2)
            image_matrix.preScale(2, 2)
            #image_matrix.preRotate(90)
            pix = page.getPixmap(alpha = True, matrix=image_matrix)
            pix.writePNG(fdname +'page_1.png')
            txt_len = 0
            text=""
            bbox=""
            doc.close()
                
            file_name_arr= fdname +'page_1.png'
            response = request_ocr(api_key, file_name_arr )
                
            if response.status_code != 200 or response.json().get('error'):
                print(response.text)

            else:
                for idx, resp in enumerate(response.json()['responses']):
                    t = resp['textAnnotations'][0]
                    #print(t['description'].casefold())
                    text = t['description'].replace('\n',' ')
                        
            
            #print(text)
            if '1129' in text or 'i129' in text or '1-129' in text or 'I-129' in text or 'I129' in text or \
            '1539' in text or 'i539' in text or 'I-539' in text  or '1-539' in text or 'I539' in text or \
            '1765' in text or 'i765' in text or '1-765' in text or 'I-765'  in text or 'I765'  in text or \
            '1131' in text or 'i131' in text or '1-131' in text or 'I-131'  in text or 'I131'  in text or \
            '1140' in text or 'i140' in text or '1-140' in text or 'I-140'  in text or 'I140'  in text or \
            '1485' in text or 'i485' in text or '1-485' in text or 'I-485' in text or 'I485' in text :
                #print('*****************IN**********************')
                #os.chdir(fdname+"/I-129 approvals - to be processed/")
                os.chdir(ingestion)
                pdf_file = open(file,'rb')
                pdf_reader = PdfFileReader(pdf_file)
                cover_pages = ['1-129 Approval Notice','1-129 Premium Processing Receipt Notice',\
                    '1-129 Receipt Notice','1-539 Approval Notice','1-539 Biometrics Appointment Notice',\
                    '1-539 Receipt Notice']

                pageNumbers = pdf_reader.getNumPages()

                for i in range (pageNumbers):
                    #print(i)
                    if (text.strip() not in cover_pages and i == 0) or i > 0: #neglate first page if it has predefined cover page text
                        #print('splitted')
                        pdf_writer = PdfFileWriter()
                        pdf_writer.addPage(pdf_reader.getPage(i))
                        split_motive = open(dest+str(i+1)+'.pdf','wb')
                        pdf_writer.write(split_motive)
                        split_motive.close()

                pdf_file.close()
                
                merge(file)
                #os.remove(fdname+"/I-129 approvals - to be processed/"+file)
                os.remove(ingestion+file)
            else:
                dest1 = get_next_file(file, fdname+'/processing/')
                #shutil.move(fdname+'/I-129 approvals - to be processed/'+file, dest1)
                shutil.move(ingestion+file, dest1)
                
            
            processed_files.append(file)
            rename()
        
        #if os.path.exists(fdname + 'templog.xlsx'):
            #os.remove(fdname + 'templog.xlsx')
        error_log('', 'Successful', '','mail')
    except Exception as e:
        error_log(e, 'Failed', 'Something Went Wrong. Process Stopped','mail')
        pass
        
        
    #onlyfiles = [f for f in listdir(dest) if isfile(join(dest, f))]

def merge(sfname):
    try: 
        #print('sfname', sfname)
        source=fdname+"/processing/"
        i = 1;
        os.chdir(source)
        pdfs = [];
        for file in sorted(glob.glob("*.pdf"),key=lambda x: int(x.split(".")[0])):
            #print(file)
            
            new_file_name ='';
            doc = fitz.open(join(source, file))
            page = doc[0]
                        
            image_matrix = fitz.Matrix(2,2)
            image_matrix.preScale(2, 2)
            #image_matrix.preRotate(90)
            pix = page.getPixmap(alpha = True, matrix=image_matrix)
            pix.writePNG(fdname +'page_1.png')
            txt_len = 0
            text=""
            bbox=""
            doc.close()
                
            file_name_arr= fdname +'page_1.png'
            response = request_ocr(api_key, file_name_arr )
                
            if response.status_code != 200 or response.json().get('error'):
                print(response.text)
            else:
                for idx, resp in enumerate(response.json()['responses']):
                    t = resp['textAnnotations'][0]
                    #print(t['description'].casefold())
                    text = t['description'].replace('\n',' ')
                        
            #print(text)
            #print('$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$')
            try:
            #if True:
                if '1129' in text or 'i129' in text or '1-129' in text or 'I-129'  in text :
                    #print('form i129')
                    try:
                        text1 = text.split('Beneficiary')
                        try:
                            text2= text1[1].split('of')
                            text3 = str(text2[0].strip())
                            text4 = str(text2[1].strip())
                            current_page = text3[-1]
                            total_page = text4[0]
                            if current_page.isdigit() and total_page.isdigit():
                                pass
                            else:
                                text1 = text.lower().split('page')
                                text2= text1[1].split('of')
                                text3 = str(text2[0].strip())
                                text4 = str(text2[1].strip())
                                current_page = text3[-1]
                                total_page = text4[0]
                            pdfs.append(file)
                        except:
                            text1 = text.lower().split('page')
                            text2= text1[1].split('of')
                            text3 = str(text2[0].strip())
                            text4 = str(text2[1].strip())
                            current_page = text3[-1]
                            total_page = text4[0]
                            pdfs.append(file)
                            
                    except:
                        try:
                            text1 = text.lower().split('page')
                            text2= text1[1].split('of')
                            text3 = str(text2[0].strip())
                            text4 = str(text2[1].strip())
                            current_page = text3[-1]
                            total_page = text4[0]

                            
                            pdfs.append(file)
                    
                        except:
                            update_log(fdname, sfname, file )
                            pass
                        pass    
                elif '1539' in text or 'i539' in text or 'I-539' in text  or '1-539' in text or \
                    '1765' in text or 'i765' in text or '1-765' in text or 'I-765'  in text or 'I765'  in text or \
                    '1140' in text or 'i140' in text or '1-140' in text or 'I-140' in text or 'I140' in text or \
                    '1485' in text or 'i485' in text or '1-485' in text or 'I-485' in text or 'I485' in text :
                    #print('i485')
                    try:
                    #if True:
                        #print(text)
                        text1 = text.split('Beneficiary')
                        try:
                            text2= text1[1].split('of')
                            #print(text2)
                            text3 = str(text2[0].strip())
                            text4 = str(text2[1].strip())
                            current_page = text3[-1]
                            total_page = text4[0]
                            #print(current_page, total_page)
                            if current_page.isdigit() and total_page.isdigit():
                                #print('Yes')
                                pass
                            else:
                                #print('No')
                                text1 = text.lower().split('page')
                                text2= text1[1].split('of')
                                text3 = str(text2[0].strip())
                                text4 = str(text2[1].strip())
                                current_page = text3[-1]
                                total_page = text4[0]
                            pdfs.append(file)
                        except:
                            text1 = text.lower().split('page')
                            text2= text1[1].split('of')
                            text3 = str(text2[0].strip())
                            text4 = str(text2[1].strip())
                            current_page = text3[-1]
                            total_page = text4[0]
                            pdfs.append(file)
                            
                    except:
                        try:
                            text1 = text.split('PAGE')
                            text2= text1[1].split('of')
                            text3 = str(text2[0].strip())
                            text4 = str(text2[1].strip())
                            current_page = text3[-1]
                            total_page = text4[0]

                            
                            pdfs.append(file)
                    
                        except:
                            update_log(fdname, sfname, file )
                            pass
                        pass
                        
                elif '1131' in text or 'i131' in text or '1-131' in text or 'I-131'  in text or 'I131'  in text:
                    #print('form i131')
                    try:
                    #if True:
                        #print(text)
                        
                        text1 = text.lower().split('page')
                        text2= text1[1].split('of')
                        text3 = str(text2[0].strip())
                        text4 = str(text2[1].strip())
                        current_page = text3[-1]
                        total_page = text4[0]
                        pdfs.append(file)
                            
                    except:
                        try:
                            text1 = text.split('PAGE')
                            text2= text1[1].split('of')
                            text3 = str(text2[0].strip())
                            text4 = str(text2[1].strip())
                            current_page = text3[-1]
                            total_page = text4[0]

                            
                            pdfs.append(file)
                    
                        except:
                            update_log(fdname, sfname, file )
                            pass
                        pass
                        
                else:
                    #print('else')
                    if len(pdfs) == 0:
                        #if first page having issues
                        update_log(fdname, sfname, file )
                        continue
                        
                    else:
                        try:
                            #if not in the above condition the last page might not have page number
                            current_page = int(current_page) + 1 
                            pdfs.append(file)
                        except:
                            update_log(fdname, sfname, file )
                            pass
                
                if total_page == 'i':
                    total_page = 1

                #print(current_page ,total_page, len(pdfs))            
                if int(current_page) == int(total_page) and len(pdfs)== int(total_page):
                    merger = PdfFileMerger()

                    for pdf in pdfs:
                        #print(pdf)
                        merger.append(pdf)
                    
                    sfname_1 = sfname.split('.pdf')
                    sfname = sfname_1[0]
                    
                    merger.write(sfname+"_"+str(i)+".pdf")
                    merger.close()
                    [os.remove(f) for f in pdfs]
                    pdfs = []
                    i +=1
                
                
                os.remove(fdname +'page_1.png')
                #print('**************************************') 

            except:
                pass  
    except Exception as e:
        error_log(e, 'Failed', 'Something Went Wrong. Process Stopped')


def merge_old(sfname):
    try: 
        #print('sfname', sfname)
        source=fdname+"/processing/"
        i = 1
        os.chdir(source)
        pdfs = []
        barcode_page = []
        
        for file in sorted(glob.glob("*.pdf"),key=lambda x: int(x.split(".")[0])):
            #print(file)
            
            new_file_name ='';
            doc = fitz.open(join(source, file))
            page = doc[0]
                        
            image_matrix = fitz.Matrix(2,2)
            image_matrix.preScale(2, 2)
            #image_matrix.preRotate(90)
            pix = page.getPixmap(alpha = True, matrix=image_matrix)
            pix.writePNG(fdname +'page_1.png')
            txt_len = 0
            text=""
            bbox=""
            doc.close()
                
            file_name_arr= fdname +'page_1.png'
            response = request_ocr(api_key, file_name_arr )
                
            if response.status_code != 200 or response.json().get('error'):
                print(response.text)
            else:
                for idx, resp in enumerate(response.json()['responses']):
                    t = resp['textAnnotations'][0]
                    #print(t['description'].casefold())
                    text = t['description'].replace('\n',' ')
                        
            #print(text)
            #print('$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$')
            try:
                barcode_text = BarcodeReader(file_name_arr)
                #print(barcode_text)
                if barcode_text is not None:
                    if len(pdfs)==0 and barcode_text.strip() not in barcode_page:
                        pdfs.append(file)
                        barcode_page.append(barcode_text.strip())
                    elif len(pdfs)>0 and barcode_text.strip() in barcode_page:
                        pdfs.append(file)
                        barcode_page.append(barcode_text.strip())
                    else:
                        merger = PdfFileMerger()

                        for pdf in pdfs:
                            #print(pdf)
                            merger.append(pdf)
                        
                        sfname_1 = sfname.split('.pdf')
                        sfname = sfname_1[0]
                        
                        merger.write(sfname+"_"+str(i)+".pdf")
                        merger.close()
                        [os.remove(f) for f in pdfs]
                        pdfs = []
                        pdfs.append(file)

                        barcode_page = []
                        barcode_page.append(barcode_text.strip())

                        i +=1
                
                        os.remove(fdname +'page_1.png')
                else:
                    update_log(fdname, sfname, file )

            except:
                pass  
                
        if len(pdfs) > 0:
            merger = PdfFileMerger()

            for pdf in pdfs:
                #print(pdf)
                merger.append(pdf)
            
            sfname_1 = sfname.split('.pdf')
            sfname = sfname_1[0]
            
            merger.write(sfname+"_"+str(i)+".pdf")
            merger.close()
            [os.remove(f) for f in pdfs]
            pdfs = []
            pdfs.append(file)

            barcode_page = []
            barcode_page.append(barcode_text.strip())

            i +=1
    
            os.remove(fdname +'page_1.png')
            
    except Exception as e:
        error_log(e, 'Failed', 'Something Went Wrong. Process Stopped','')
        pass

   
    
    

def rename():
    #print('rename')
    global m
    try :
    #if True:
        today = date.today()
        year = today.year
        month = today.month
        day = today.day
        today = str(month)+'/'+str(day)+'/'+str(year)
        #print(today)
        if not os.path.exists(fdname + 'nrtlog.xlsx'):
            
            book = Workbook()
            ws = book.active
            ws.cell(1,1).value = 'Source File Name'
            ws.cell(1,2).value = 'Renamed File Name'
            ws.cell(1,3).value = 'File Rename Status'
            ws.cell(1,4).value = 'Date File Renamed'
            ws.cell(1,5).value = 'Zoom Update Status'
            ws.cell(1,6).value = 'Date zoom updated'
            ws.cell(1,7).value = 'File Upload Status'
            ws.cell(1,8).value = 'Date File Uploaded to ND'
            ws.cell(1,9).value = 'Document ID'
            ws.cell(1,10).value = 'Workspace Name'
            ws.cell(1,11).value = 'Receipt Number'
            
            ws.column_dimensions["A"].width = 50.0
            ws.column_dimensions["B"].width = 60.0
            ws.column_dimensions["C"].width = 40.0
            ws.column_dimensions["D"].width = 10.0
            ws.column_dimensions["E"].width = 40.0
            ws.column_dimensions["F"].width = 40.0
            ws.column_dimensions["G"].width = 40.0
            ws.column_dimensions["H"].width = 40.0
            ws.column_dimensions["I"].width = 40.0
            ws.column_dimensions["J"].width = 40.0
            ws.column_dimensions["K"].width = 40.0
            
            ws['A1'].font = Font(bold=True)
            ws['B1'].font = Font(bold=True)
            ws['C1'].font = Font(bold=True)
            ws['D1'].font = Font(bold=True)
            ws['E1'].font = Font(bold=True)
            ws['F1'].font = Font(bold=True)
            ws['G1'].font = Font(bold=True)
            ws['H1'].font = Font(bold=True)
            ws['I1'].font = Font(bold=True)
            ws['J1'].font = Font(bold=True)
            ws['K1'].font = Font(bold=True)
            
            book.save(fdname +'nrtlog.xlsx')
            
        
        
        mypath = fdname +'processing'
        mypathd = fdname +'processed'
        mypathe = fdname +'errors'
        month_list = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November','December']
        onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
        #print(onlyfiles)
        for file in onlyfiles:
            
            file_name = basename(file)
           
            if ('Denial' or 'denial') in file_name:
                file_name = 'denial'

            elif ('Revocation' or 'revocation') in file_name:
                file_name = 'revocation'

            elif ('RFE' or 'rfe') in file_name:
                file_name = 'RFE'

            elif ('EAD' or 'ead') in file_name:
                file_name = 'EAD'    

            
          
            book2 = load_workbook(fdname +'nrtlog.xlsx')
            sheet = book2.active
            new_file_name ='';
            doc = fitz.open(join(mypath, file))
            page = doc[0]
                    
            image_matrix = fitz.Matrix(2,2)
            image_matrix.preScale(2, 2)
            #image_matrix.preRotate(90)
            pix = page.getPixmap(alpha = True, matrix=image_matrix)
            pix.writePNG(fdname +'page_1.png')
            txt_len = 0
            text=""
            bbox=""
            ndtxt=""
            
            doc.close()
            
            file_name_arr= fdname +'page_1.png'

            response = request_ocr(api_key, file_name_arr )
            print('response  :',response)
            
            if response.status_code != 200 or response.json().get('error'):
                print('response text   :',response.text)
               
            else:
                for idx, resp in enumerate(response.json()['responses']):
                    t = resp['textAnnotations'][0]
                    #print(t['description'].casefold())
                    text = t['description'].replace('\n',' ')
                    barcode_text = BarcodeReader(file_name_arr)
                
                    
            print('text   :',text)
            
            if '1129' in text or 'i129' in text or '1-129' in text or 'I-129' in text or 'I129' in text or \
                '1539' in text or 'i539' in text or 'I-539' in text  or '1-539' in text or 'I539' in text or \
                '1765' in text or 'i765' in text or '1-765' in text or 'I-765'  in text or 'I765'  in text or \
                '1131' in text or 'i131' in text or '1-131' in text or 'I-131'  in text or 'I131'  in text or \
                '1130' in text or 'i130' in text or '1-130' in text or 'I-130'  in text or 'I130'  in text or \
                '1140' in text or 'i140' in text or '1-140' in text or 'I-140'  in text or 'I140'  in text or \
                '1485' in text or 'i485' in text or '1-485' in text or 'I-485' in text or 'I485' in text or \
                'N400' in text or 'N-400' in text :
                pass
            else:
                #print('Notice type not configured')
                sheet.append((file,'','Notice type not configured',today))
                dest = get_next_file(file, mypathe)
                shutil.move(mypath+'/'+file, dest) 
                
                sheet.column_dimensions["A"].width = 50.0
                sheet.column_dimensions["B"].width = 60.0
                sheet.column_dimensions["C"].width = 40.0
                sheet.column_dimensions["D"].width = 10.0
                sheet.column_dimensions["E"].width = 40.0
                sheet.column_dimensions["F"].width = 40.0
                sheet.column_dimensions["G"].width = 40.0
                sheet.column_dimensions["H"].width = 40.0
                sheet.column_dimensions["I"].width = 40.0
                sheet.column_dimensions["J"].width = 40.0
                sheet.column_dimensions["K"].width = 40.0
                book2.save(fdname +'nrtlog.xlsx')
                
                continue

            
            
            #print(file)
            #if re.search('I-485', text, re.IGNORECASE):
            k=1
            nottype = re.search('Notice Type: (.*) Notice', text)
            if(nottype==None):
                nottype = re.search('ALIEN NUMBER (.*) TYPE', text)
                if(nottype==None):
                    nottype = re.search('NOTICE TYPE (.*) TYPE', text)
                    if(nottype==None):
                        nottype = file_name
            
            

            nottype1 = re.search('BENEFIT. (.*) CASE TYPE', text)
            if(nottype1==None):
                nottype1 = re.search('CASE TYPE (.*) NOTICE DATE', text)
                
            nottype2 = re.search('FOR INITIAL (.*) Case Type', text)
            nottype3 = re.search('Notice of (.*) Department of', text)
            nottype4 = re.search('appear for an (.*) on your Application for Naturalization', text)
            nottype5 = re.search('BENEFIT. (.*) CASE TYPE', text)
            nottype6 = re.search('1-140 (.*)', text)                    

             
            
            try:
                if nottype == file_name:
                    notetype = file_name

                elif(nottype.group(1).split(' Notice')[0]!=None):
                    notetype=nottype.group(1).split(' Notice')[0]
                else:
                    notetype=nottype.group(1)
                
                # print('dd',notetype)
                # quit()

                if(notetype=="denial"):
                    catype= re.search('Citizenship and Immigration Services(.*)Application for', text)

                    catype = catype.group().split(' ')[-3][:-1].strip()
                    
                
                #print("first loop")            
                if(notetype=="Receipt"):
                    print('hi')
                    catype= re.search('Case Type(.*) - ', text)
                    
                    if(catype==None):
                        catype=re.search('Case (.*) - ', text)
                        catype2=catype.group(1).split(' - ')[0].strip().split(" ")[1]
                                        
                        
                    else:
                        catype2=catype.group(1).split(' - ')[0].strip() 

                
                        
                    rdate = re.search('Received Date(.*)Priority Date', text)   
                   
                    try:
                        contains_digit = any(map(str.isdigit, rdate.group(1)))
                        #print(contains_digit)
                       
                    except:
                        contains_digit = False
                        pass
                    bname=re.search('Beneficiary(.*)Notice Date', text)
                    ndate=re.search('Notice Date(.*)1 of', text).group()
                    
                    if(ndate==None):
                        ndate=re.search('Notice Dale(.*)Case Type', text)

                    ndate_y = ndate.split("Page")[1].strip()[6:10].strip()

                    if(rdate==None or contains_digit!=True):
                        rdate = re.search('Priority Date(.*)Notice Date', text)        
                        
                        if(rdate==None):
                            rdate = re.search('Priority Dale(.*)Notice Date', text)
                            if(rdate==None):
                                rdate = re.search('Received Date(.*)Applicant', text)
                                if(rdate==None):
                                    rdate = re.search('Applicant(.*)Notice Date', text)
                                    if(rdate==None):
                                        rdate = re.search('(.*)Case Type', text)

              
                    
                    if("1129" not in catype2):
                        if('1140' in catype2):
                            result = re.search('Beneficiary(.*)Notice Date', text)
                            if(result==None):
                                result = re.search('Beneficiary(.*)Notice Type:', text)
                            
                        

                        elif('1539' not in catype2):
                            result = re.search('Applicant(.*)Priority Date', text)
                            if(result==None):
                                result = re.search('Applicant(.*)Beneficiary', text) 
                        else:
                            result = re.search('Beneficiary(.*)Notice Date', text)
                            if(result==None):
                                result = re.search('Beneficiary(.*)Notice Type:', text) 
                            
                    else:
                        result = re.search('Petitioner(.*)Beneficiary', text)
                        if(result==None):
                            result = re.search('Petitioner(.*)Bencficiary', text)
                            if(result==None):
                                result = re.search('Pedtioner(.*)Notice Date', text)
                                if(result==None):
                                    result = re.search('Applicant(.*)Beneficiary', text)
                        cas = re.search('Name DOB COB Class Consulate/POE OCC (.*)', text)
                        if(cas==None):
                            cas=re.search('Name. DOB COB Class Consulate/POE OCC (.*)', text)
                            if(cas==None):
                                cas=re.search('Name(.*)Class Consulate', text)
                                if(cas==None):
                                    cas=re.search('Name(.*)DOB', text)
                                    if(cas==None):
                                        cas=re.search('Number of workers:(.*)DOB', text)
                        cas = cas.group(1).strip().split("/")[0]
                        cas = re.sub(r'\d', '', cas).replace("DOB COB","")
                        bname1=bname.group(1).strip()
                 

                    match = re.sub(r'\w*\d\w*', '', result.group(1))
                    match=match.replace("//","")
                    
                    name1=rem_duplicate(match.strip())
                    print(name1)
                    rdate1=rdate.group(1).replace(" |-","").replace("APPLICATION TO EXTEND/CHANGE NONIMMIGRANT STATUS","").split("/")
                    print(rdate1)
                    rdatem=rdate1[0].replace("Received Date","").replace("Priority Date","").strip()
                    rdatem=rdatem.replace("Petitioner","")
                    rdate3=rdate1[2][:4]
                    rdate2=rdate1[1]
                    rdate0=rdate1[0][-2:]
                    
                    
                    name1=name1.replace("|","")
                    name1=name1.replace("//","")
                    name1=name1.replace("Notice","")
                    name1=name1.replace("Date","")
                    name1=name1.replace("Received","")
                    name1=name1.replace("priority","")
                    name1=name1.replace("Page","")
                    name1=name1.replace("of","")
                    name1=re.sub(r'\w*\d\w*', '', name1)
                    if("SIX CONTINENTS" in name1.strip()):
                        name1="Six Continents"
                    elif("FEDEX" in name1.strip()):
                        name1="FedEx"
                    

                    
                    if("1765" in catype.group(1).split(' - ')[0].strip()):

                        
                        
                        creq= re.search('Class requested: (.*) We have received', text)

                        creq3=creq.group(1).strip()
                                              
                        creq3=creq3.replace(".","")
                        creq3=creq3.replace("!","")
                        creq3=creq3.replace("i","")
                        creq3=creq3.replace("1","")
                        creq3=creq3.replace("|","")
                        creq3=creq3.replace("I","")
                        creq3=creq3.replace(":","")
                        creq3=creq3.strip()
                        rdate2=rdate.group(1).strip().split("/")
                        rdate3=rdate2[2][:4]
                        rdate1=rdate2[1]
                        rdate0=rdate2[0][-2:]

                        file_rename=name1.strip()+" -I765 "+creq3+" Receipt Notice "+ndate_y+".pdf"
                        

                    elif("1129" in catype2):
                        print("jks123")
                        creq= re.search('Class requested: (.*) We have mailed', text)
                        if(creq==None):
                            creq= re.search('Class requested: (.*) We have received', text)
                        creq2=creq.group(1).split(" ")[0].strip().replace("HI","H1")
                        creq3=creq2[:1] + '-' + creq2[1:] 
                        print(creq)
                        ndate1=ndate.group(1).split("/")
                        print(ndate1)
                        ndate3=ndate1[2][:4]
                        print(ndate3)
                        print(bname1)                        
                        file_rename=bname1.strip()+" - "+name1.strip()+" - "+creq2+" Receipt Notice - "+ndate_y+".pdf" #done

                        

                    elif("1485" in catype2):
                        rdate1=rdate.group(1).split("/")
                        #print(rdate1)
                        rdate3=rdate1[2][:4]
                        rdate2=rdate1[1]
                        rdate0=rdate1[0][-2:] 
                        #print(rdate3)
                        #print(rdate2)
                        #print(rdate0)                    
                        file_rename=name1.strip()+" - I485 Receipt Notice - "+ndate_y+".pdf"
                        #done n date needed

                       
                        #print(file_rename)
                    elif('1539' in catype2):
                        #print("jksreceipt")
                        name2=name1.replace("//","")
                        name2=name2.replace("Notice","")
                        name2=name2.replace("Date","")
                        name2=name2.replace("Received","")
                        name2=name2.replace("priority","")
                        name2=name2.replace("Page","")
                        name2=name2.replace("of","")
                        ndate1=ndate.group(1).replace("EXTEND/CHANGE","")
                                   
                        ndate2=ndate1.strip().split("/")
                        ndate3=ndate2[2][:4]
                      
                                  
                        file_rename= "I539 Receipt Notice - "+name2.strip()+" - "+ndate_y+".pdf" 
                        #done rena
                        #print(file_rename)                    
                    else:
                       
                        file_rename=manual_replace(name1.strip()+catype.group(1).split(' - ')[0].strip(),'I-',0)+" - "+notetype+" - Notice - "+ndate_y+".pdf"      

                        file_rename=file_rename.replace("Receiver","").replace("I-485)","I-485J")

                    file_rename = fnameclean(file_rename)
                    os.rename(mypath+'/'+file,mypath+'/'+file_rename)
                    sheet.append((file,file_rename,'File was renamed',today,'Data not updated','','Notice type not configured'))
                    dest = get_next_file(file_rename, mypathd)
                    shutil.move(mypath+'/'+file_rename, dest)
                    k=0

                elif(notetype.strip()=="Recei t CASE" or notetype.strip()=="R ei t CASE" or notetype.strip()=="Reit CASE" or notetype.strip()==" it CASE" or notetype.strip()=="Receipt CASE" or notetype.strip()=="NOTICE DATE Receipt CASE" or notetype.strip()=="NOTICE TYPE Receipt CASE" or "NOTICE DATE Receipt" in notetype or "CASE" in notetype):

                    # print('xxx')
                    # quit()

                    # ndate=re.search('Notice Date(.*)1 of', text).group()
                    
                    # if(ndate==None):
                    #     ndate=re.search('Notice Dale(.*)Case Type', text)

                    # ndate_y = ndate.split("Page")[1].strip()[6:10].strip()
                    
                    
                    
                    catype = re.search('CASE TYPE(.*)', text)
                    if(catype==None):
                        catype = re.search('Recei t CASE TYPE(.*)', text)
                        if(catype==None):
                            catype = re.search('R ei t CASE TYPE(.*)', text)
                            if(catype==None):
                                catype = re.search('Receipt CASE TYPE(.*)', text)
                            
                    
                    catype=catype.group(1).strip().split(",")
                    catype3=catype[0].strip().split(".")[0]
                    
                    
                    
                    if("1-131" in catype[0].strip() or "L-131" in catype[0].strip().split(".")[0]):
                        
                        rdate = re.search('RECEIVED DATE(.*) NOTICE', text)
                        if(rdate==None):
                            rdate = re.search('RECEIVED DATE(.*) USCIS ALIEN NUMBER', text)
                            if(rdate==None):
                                rdate = re.search('RECEIVER DATE(.*) USCIS ALIEN NUMBER', text)
                        #  rdate = rdate.replace("RECEIVED DATE", "")
                        # if(catype.group(1).split(' - ')[0].strip()=="1485"):
                        
                        result = re.search('PAYMENT INFORMATION:(.*) C/O', text)
                        if(result.group(1).strip()!=""):
                            name=result.group(1).strip()
                            match = re.sub(r'\w*\d\w*', '', result.group(1))
                            name=match.strip().replace("$","")
                        else:
                            result = re.search('RECEIPT NUMBER(.*) RECEIVED DATE', text)
                            match = re.sub(r'\w*\d\w*', '', result.group(1))
                            name=match.strip().replace("$","")
                            
                        
                        rdate2=rdate.group(1).strip().replace(",","")
                        rdate2=rdate2.replace("'","")
                        rdate1=rdate.group(1).strip().split(" ")
                        rdatemon=mon_conversion(rdate1[1].strip())
                        
                        file_rename=name.replace(".","")+" - I131 Receipt Notice"+" - "+rdate1[3].strip()+".pdf"
                        #done
                        
                        file_rename = fnameclean(file_rename)
                        os.rename(mypath+'/'+file,mypath+'/'+file_rename)  
                        sheet.append((file,file_rename,'File was renamed',today,'Data not updated','','Notice type not configured'))
                        #shutil.move(mypath+'/'+file_rename, mypathd+'/'+file_rename)
                        dest = get_next_file(file_rename, mypathd)
                        shutil.move(mypath+'/'+file_rename, dest)
                        k=0
                    elif("1-1485" in catype[0].strip() or "1-485" in catype[0].strip() or "I-485" in catype[0].strip() or "1-485" in catype3):
                        rdate = re.search('PRIORITY DATE(.*)PREFERENCE CLASSIFICATION', text)
                        
                        

                        if(rdate==None):
                            rdate = re.search('RECEIVED DATE(.*)PREFERENCE CLASSIFICATION', text)
                            
                        
                        result = re.search('PREFERENCE CLASSIFICATION(.*) CIO', text)
                        if(result==None):
                            result = re.search('PREFERENCE CLASSIFICATION(.*) C/O', text)
                        
                        rdate1=rdate.group(1).strip().split(" ")
                        #print(rdate.group(1).strip().split(" "))
                        rdatemon=mon_conversion(rdate1[3].strip())
                        name=result.group(1).strip().replace(".","")
                        name=name.replace("Alien worker Form I-140","")
                        
                        #doubt

                        file_rename = name.strip()+" - I485 Receipt Notice - "+rdate1[5].strip()+".pdf"

                        file_rename = fnameclean(file_rename)
                                                
                        os.rename(mypath+'/'+file,mypath+'/'+file_rename)    
                        sheet.append((file,file_rename,'File was renamed',today,'Data not updated','','Notice type not configured'))
                        #shutil.move(mypath+'/'+file_rename, mypathd+'/'+file_rename)
                        dest = get_next_file(file_rename, mypathd)
                        shutil.move(mypath+'/'+file_rename, dest)
                        k=0 
                    elif("1-130" in catype[0].strip().split(".")[0]):
                        
                        rdate = re.search('RECEIVED DATE (.*) PREFERENCE CLASSIFICATION', text)
                        result = re.search('Name (.*) Date of', text)
                        
                        rdate1=rdate.group(1).strip().split(" ")
                        try:
                            rdatemon=mon_conversion(rdate1[0].strip())
                            file_rename="I-130 - "+"Receipt"+" - "+rdate1[2].strip()+"-"+str(rdatemon)+"-"+rdate1[1].strip().replace(",","")+" - "+result.group(1).strip()+".pdf"
                        except:                        
                            file_rename="I-130 - "+"Receipt"+" - "+rdate.group(1).strip()+" - "+result.group(1).strip()+".pdf"
                            pass
                        file_rename = fnameclean(file_rename)
                        os.rename(mypath+'/'+file,mypath+'/'+file_rename)    
                        sheet.append((file,file_rename,'File was renamed',today,'Data not updated','','Notice type not configured'))
                        #shutil.move(mypath+'/'+file_rename, mypathd+'/'+file_rename)
                        dest = get_next_file(file_rename, mypathd)
                        shutil.move(mypath+'/'+file_rename, dest)
                        k=0
                    elif("1-539" in text or "I-539" in text):
                        
                        rdate = re.search('RECEIVED DATE(.*)PAGE 1', text)
                        if(rdate==None):
                            rdate = re.search('RECEIVED DATE PAGE(.*)1 of', text) 
                        
                        #print(rdate)
                        result = re.search('PAYMENT INFORMATION:(.*) C/O', text)
                        if(result==None):
                            result = re.search('PAYMENT INFORMATION:(.*) CIO', text)
                        
                        
                        rdate1=rdate.group(1).strip().split(" ")
                        
                        rdatemon=mon_conversion(rdate1[1].strip())
                        #print(rdatemon)
                        try:
                            rdatemon=mon_conversion(rdate1[1].strip())

                            file_rename="I539 "+"Receipt Notice"+" - "+rdate1[3].strip()+" - "+result.group(1).strip()+".pdf"

                            # done I539 Receipt Notice - 2022 - Swan, Elizabeth
                        except:                        
                            file_rename="I539 "+"Receipt Notice"+" - "+rdate.group(1).strip()+" - "+result.group(1).strip()+".pdf"
                            # done I539 Receipt Notice - 2022 - Swan, Elizabeth
                            pass 
                        file_rename = fnameclean(file_rename)                        
                        os.rename(mypath+'/'+file,mypath+'/'+file_rename)    
                        sheet.append((file,file_rename,'File was renamed',today,'Data not updated','','Notice type not configured'))
                        #shutil.move(mypath+'/'+file_rename, mypathd+'/'+file_rename)
                        dest = get_next_file(file_rename, mypathd)
                        shutil.move(mypath+'/'+file_rename, dest)
                        k=0                        
                    elif("1-140" in text or "I-140" in text or "1-140" in text):
                        
                        rdate = re.search('RECEIVED DATE(.*)PREFERENCE CLASSIFICATION', text)
                        
                        result1 = re.search('Name Date of Birth Country of Birth Class (.*)', text)
                        if(result1==None):
                            result1=re.search('Name Country of Birth Class(.*)If this', text)
                            if(result1==None):
                                result1=re.search('Name Country of Birth. Class(.*)If this', text)
                        
                        
                        result1 = result1.group(1).strip()
                        
                        result1 = re.sub(r'\d', '', result1)
                        result1=result1.replace("(If Applicable)","").replace("//","").replace("Date of Birth","").strip()
                        result1=result1.split(' ')
                        rname=result1[0]+" "+result1[1]
                        
                        
                                            
                        result = re.search('PAYMENT INFORMATION: (.*)CIO', text)
                        
                        pc = re.search('203 (.*)', text)
                        if(result==None):
                            result = re.search('PAYMENT INFORMATION: (.*)C/O', text)
                            
                            if(result==None):
                                result = re.search('PRIORITY DATE (.*)RECEIVED DATE', text)
                                if(result==None):
                                    result = re.search('ABILITY (.*) CIO', text)
                                    if(result==None):
                                        result = re.search('ABILITY (.*) C/O', text)
                                        
                        
                        
                        rdate1=rdate.group(1).replace(" PAGE ","").replace("PRIORITY DATE ","").strip().split(" ")
                        
                        pc1=pc.group(1).strip().split(" ")
                        
                        
                        try:
                            rdatemon=mon_conversion(rdate1[1].strip())
                            
                            file_rename=result.group(1).strip()+" - "+rname.strip()+" - "+"I-140 E "+pc1[0]+" Receipt Notice - "+rdate1[3].strip()+".pdf"

                            # done check DOE, John - FedEx - I140 PP Receipt Notice - 2022

                        except:                        
                            file_rename=result.group(1).strip()+" - "+rname.strip()+" - "+"I-140 E"+pc1[0]+" Receipt Notice - "+rdate.group(1).strip()+".pdf"

                            # done check DOE, John - FedEx - I140 PP Receipt Notice - 2022

                            pass
                        file_rename = fnameclean(file_rename)                       
                        os.rename(mypath+'/'+file,mypath+'/'+file_rename)    
                        sheet.append((file,file_rename,'File was renamed',today,'Data not updated','','Notice type not configured'))
                        #shutil.move(mypath+'/'+file_rename, mypathd+'/'+file_rename)
                        dest = get_next_file(file_rename, mypathd)
                        shutil.move(mypath+'/'+file_rename, dest)
                        k=0
                    else:
                        rdate = re.search('RECEIVED DATE(.*) PAGE 1', text)
                        if(rdate==None):
                            rdate = re.search('RECEIVED DATE PAGE(.*) 1 of', text)
                        
                        result = re.search('PAYMENT INFORMATION:(.*) C/O', text)
                        #print(rdate)
                        #print(result)
                        if(result==None):
                            result = re.search('PAYMENT INFORMATION:(.*) CIO', text)
                        if(result.group(1).strip()!=""):
                            name=result.group(1).strip()
                        else:
                            result = re.search('RECEIPT NUMBER(.*) RECEIVED DATE', text)
                            match = re.sub(r'\w*\d\w*', '', result.group(1))
                            name=match.strip().replace("PRIORITY DATE","")
                        
                        rdate2= re.sub(r'[a-zA-Z]*\d+[a-zA-Z]+|[a-zA-Z]+\d+[a-zA-Z]*','',rdate.group(1).strip())                    
                        rdate1=rdate2.split(" ")
                        #print(rdate1)
                        creq = re.search('Eligibility Category: (.*) We have', text)
                        name = re.sub(r'\w*\d\w*', '', name)
                        name=name.replace("$.","")
                        #print(name)
                        #print(creq)
                        try:
                            rdatemon=mon_conversion(rdate1[1].strip())
                            #print(rdatemon)
                            #print(catype[0].strip())
                            file_rename=manual_replace(name+"-"+catype[0].strip(),'I',0)+" "+creq.group(1)+" - "+"Receipt Notice"+" - "+rdate1[3].strip()+".pdf"
                            #print(file_rename)
                        except:                        
                            file_rename=manual_replace(name+"-"+catype[0].strip(),'I',0)+" ("+creq.group(1)+") - "+"Receipt Notice"+" - "+rdate.group(1).strip().replace("PREFERENCE CLASSIFICATION r","")+".pdf"
                            pass                    
                        
                        #print(file_rename)
                                        
                        file_rename = fnameclean(file_rename)
                        os.rename(mypath+'/'+file,mypath+'/'+file_rename)
                        sheet.append((file,file_rename,'File was renamed',today,'Data not updated','','Notice type not configured'))
                        #shutil.move(mypath+'/'+file_rename, mypathd+'/'+file_rename)
                        dest = get_next_file(file_rename, mypathd)
                        shutil.move(mypath+'/'+file_rename, dest)
                        k=0                    
                elif(notetype=="Approval" or notetype=="-Approval"):
                    catype= re.search('Case (.*) - ', text)
                    

                    if(catype==None):
                        catype = re.search('case (.*) - ', text)
                        if(catype==None):
                            catype = re.search('Cnsc (.*) - ', text)
                            if(catype==None):
                                catype = re.search('Case (.*) PETITION ', text)
                            
                    
                    #catype=any(re.match(regex_str, text) for regex in ['Case Type(.*) - ','Cuse Type(.*) - '])
                   
                    #print(catype.group(1).split(' ')[1].strip())
                    if("1765" in catype.group(1)):
                        result = re.search('Applicant(.*)Notice Date', text)
                        if(result==None):
                            result = re.search('Applicant(.*)Priority Date', text)
                        creq= re.search('Class: (.*) Valid from', text)
                        rdate= re.search('to (.*) We have mailed', text)
                        if(rdate==None):
                            rdate= re.search('to (.*) We have approved', text) 
                        match = re.sub(r'\w*\d\w*', '', result.group(1))
                        name1=rem_duplicate(match.strip())
                        name1=name1.replace("Priority","").replace("Date","")
                        rdate1=rdate.group(1).split("/")
                        #print(rdate1)
                        #print(name1)
                        #print(creq)
                        
                        file_rename=name1.strip()+" - I765 "+creq.group(1).strip().split()[0]+" Approval Notice - "+rdate1[2].replace(".","").split(" ")[0].strip()+".pdf"

                        #  done date check John Doe - I765 C03C Approval Notice  - 2022
                        #print(file_rename)
                    elif("1539" in catype.group(1)):
                        result = re.search('Beneficiary(.*)Notice Date', text)
                        if(result==None):
                            result = re.search('Beneficiary(.*)Notice Dale', text)
                            if(result==None):
                                result = re.search('Beneficiary(.*)Page 1', text)
                                if(result==None):
                                    result = re.search('Applicant(.*)Notice Date', text)
                        creq= re.search('Class: (.*) Valid from', text)
                        if(creq==None):
                            creq= re.search('Class; (.*) Valid from', text)
                        ndate= re.search('Norice Date (.*) Page 1 of', text)
                        match = re.sub(r'\w*\d\w*', '', result.group(1))
                        name1=rem_duplicate(match.strip())
                        name1=name1.replace("Page","")
                        print(result)
                        print(creq)
                        print(ndate)
                        print(name1)
                        ndate1=ndate.group(1).split("/")
                        creq2=creq.group(1).strip().replace("HIB","H1B")
                        creq2=creq2.replace("LIB","L1B")
                        creq3=creq2[:1] + '-' + creq2[1:]
                        
                        ndate1[2]=ndate1[2][:4]
                        print (ndate1[2])
                        
                        
                        file_rename=creq2.strip()+" Approval Notice - "+ndate1[2]+" - "+name1.strip()+".pdf"

                        # done H4 Approval Notice - 2022 - Swan, Elizabeth

                        print(file_rename)
                    elif(catype.group(1).split(' ')[1].strip()=="1131" or catype.group(1).split(' ')[1].strip()=="I131" or "1131" in catype.group(1)):
                        result = re.search('Applicant(.*) Priority', text)
                        
                        match = re.sub(r'\w*\d\w*', '', result.group(1))
                        match = match.replace("-", "").replace("Received Date //","").replace("Received Date","")
                        test = re.search('Valid from(.*)', text)
                        rdate1=test.group(1).split(" ")
                        
                        rdate1=rdate1[3].split("/")
                        name1=rem_duplicate(match.strip())
                        if(name1!=""):
                            name1=name1
                        else:
                            result = re.search('Priority Date(.*) Notice Date', text)
                            #print(result)
                            match = re.sub(r'(\d{2})/(\d{2})/(\d{4})', '', result.group(1))
                            name1=rem_duplicate(match.strip())
                            
                        file_rename=name1.strip()+" - I131 Approval Notice - "+rdate1[2]+".pdf" 

                        # done check date John Doe - I131 Approval Notice  - 2022
                        
                    elif("1140" in catype.group(1).strip()):
                        rdate = re.search('Priority Date (.*) Page ', text)
                        if(rdate==None):
                            rdate = re.search('Priority Date (.*) 1 of ', text)
                        result = re.search('Petitioner (.*) c/o', text)
                        ndate= re.search('Norice Date (.*) Page 1 of', text)
                        if(ndate==None):
                          ndate= re.search('Notice Date (.*) Page 1 of', text)  
                        match = re.sub(r'\w*\d\w*', '', result.group(1))
                        csec = re.search('Section: (.*) Consulate:', text)
                        if(csec==None):
                            csec = re.search('Section: (.*) Coinsulate:', text)
                        
                    
                        ndate1=ndate.group(1).split("/")
                        ndate1[2]=ndate1[2][:4]
                        name1=match.strip().replace(":","").split('Beneficiary')
                        if("Beneficlary" in text):
                            name1=match.strip().replace(":","").split('Beneficlary')
                        rdate1=rdate.group(1).split("/")
                        csec1=csec.group(1).split("(")
                        
                        cname=name1[1].replace(name1[0].strip(),"").replace("Notice Date //","").replace("Page  of","").strip()
                        if "  " in cname:
                            cname = cname.strip().split("  ")[0]
                        if("SIX CONTINENTS" in name1[0].strip()):
                            name1[0]="Six Continents"
                        elif("FEDEX" in name1[0].strip()):
                            name1[0]="FedEx"
                        
                        file_rename=cname.strip()+" - "+name1[0].strip()+" - I140 Approval Notice - "+ndate1[2]+".pdf" 



                        # done  DOE, John - FedEx - I140 Approval Notice  - 2022
                        
                      
                    elif("1130" in catype.group(1)):
                        result = re.search('Petitioner (.*)Notice Date', text)
                        match = re.sub(r'\w*\d\w*', '', result.group(1))            
                    
                        #print(result.group(1))
                        #print(match)
                        name1=match.strip().split('Beneficiary')
                        
                        file_rename="I-130 - Approval -"+name1[1].strip()+".pdf" 
                        #print(file_rename)
                    
                        
                    elif(catype.group(1).split(' - ')[0].strip()=="Type Il 29" or catype.group(1).split(' - ')[0].strip()=="Il29" or catype.group(1).split(' - ')[0].strip()=="1l29" or catype.group(1).split(' ')[1].strip()=="1129" or "1129" in catype.group(1)):
                        
                        ndate = re.search('Notice Date (.*) Page 1', text)
                        if(ndate==None):
                            ndate = re.search('Notice Date (.*) 1 of', text)
                        fdate = re.search('Valid from (.*) We have mailed', text)
                        rdate = re.search('to (.*) We have mailed', text)
                        if(rdate==None):
                            rdate = re.search('to (.*) The abo', text)
                        result = re.search('Petitioner (.*)Notice Date', text)
                        if(result==None):
                            result = re.search('Petitioner (.*)Page', text) 
                            if(result==None):
                                result= re.search('Pétitioner (.*)Notice Date', text)
                        match = re.sub(r'\w*\d\w*', '', result.group(1))
                        creq= re.search('Class: (.*) Valid from', text)
                        print(rdate)
                        print(match)
                        print(creq)
                        print(ndate)
                                        
                        
                        
                        name1=match.strip().replace("Beneficia ry","Beneficiary").split('Beneficiary')
                        ndate1=ndate.group(1).strip().split("/")
                        print(ndate1)
                        
                        creq2=creq.group(1).strip().replace("HI","H1")
                        creq2=creq2.replace("LI","L1")
                        creq3=creq2[:1] + '-' + creq2[1:]
                        ndate1[2]=ndate1[2][:4]
                        if("SIX CONTINENTS" in name1[0].strip()):
                            name1[0]="Six Continents"
                        elif("FEDEX" in name1[0].strip()):
                            name1[0]="FedEx"
                                               
                        
                        
                        file_rename=name1[1].strip().replace("Page","")+" - "+name1[0].strip()+" - "+creq2+" Approval Notice - "+ndate1[2]+".pdf" 
                        print(file_rename)

                        # done DOE, John - FedEx - H1B Approval Notice  - 2022
                        
                    elif("1485J" in catype.group(1).strip() or "1485)" in catype.group(1).strip() or "14851" in catype.group(1).strip()):
                        result = re.search('Page 1 of(.*)c/o', text)
                        if(result==None):
                            result = re.search('1 of(.*)c/o', text)
                            if(result==None):
                                result = re.search('Received Date(.*)c/o', text)
                        
                        # match = re.sub(r'\w*\d\w*', '', result.group(1))
                        match = re.search('Applicant (.*)Priority Date', text).group().split('Priority')[0].strip()
                        match = match[23:]
                        name1=rem_duplicate(match.strip())

                        # print('xxx')
                        # print((match))
                        # print(type(ndate))
                        # quit()
                                        
                        file_rename=name1.strip().replace("Received Date","")+" - I485J Approval Notice"+".pdf"
                        #print(file_rename)                    

                        # done  John Doe - I485J Approval Notice  - 2022    
                    else:
                        result = re.search('Applicant(.*)Beneficiary', text)
                        
                        match = re.sub(r'\w*\d\w*', '', result.group(1))
                        name1=rem_duplicate(match.strip())
                        cname=catype.group(1).strip().split()
                        ndate = re.search('Notice Date (.*) Page 1', text)
                        if(ndate==None):
                            ndate = re.search('Notice Date (.*) 1 of', text)
                            if(ndate==None):
                                ndate = re.search('Notice Dale (.*) Page 1 of', text)
                        ndate1=ndate.group(1).split("/")
                        ndate1[2]=ndate1[2][:4]
                        if(len(cname)>2):
                            cname2=cname[len(cname)-1]
                        else:
                            cname2=catype.group(1).split(' ')[1].strip()
                            
                        if("1485" in catype.group(1)):           
                            file_rename=name1.strip()+" - "+cname2+" Approval Notice - "+ndate1[2]+".pdf" 

                            #done fed check DOE, John - FedEx - I485 Approval Notice - 2022
                        else:
                            file_rename=name1.strip()+manual_replace(cname2,'I-',0)+" - Approval - "+".pdf" 

                    
                    file_rename = fnameclean(file_rename)                    
                    
                    os.rename(mypath+'/'+file,mypath+'/'+file_rename)
                    if(ndtxt=="approvaln"):
                        sheet.append((file,file_rename,'File was renamed',today,'','','','','','',str(barcode_text)))
                    else:
                        sheet.append((file,file_rename,'File was renamed',today,'Data not updated','','Notice type not configured'))
                    #shutil.move(mypath+'/'+file_rename, mypathd+'/'+file_rename)
                    dest = get_next_file(file_rename, mypathd)
                    shutil.move(mypath+'/'+file_rename, dest)
                    
                    k=0
                elif(notetype=="Transfer"):
                    catype= re.search('Case Type(.*) - ', text)
                    if(catype==None):
                        catype= re.search('casc Type(.*) - ', text) 
                    #catype=any(re.match(regex_str, text) for regex in ['Case Type(.*) - ','Cuse Type(.*) - '])
                    
                    result = re.search('Applicant(.*)Notice Type', text)
                    match = re.sub(r'\w*\d\w*', '', result.group(1))
            
                    #print(catype.group(1))
                    #print(match)
                    name1=rem_duplicate(match.strip())
                
                    
                    file_rename=manual_replace(name1.strip()+"-"+catype.group(1).split(' - ')[0].strip(),'I-',0)+" - "+notetype+" Notice"+".pdf" 
                    #print(file_rename) 
                                    
                    file_rename = fnameclean(file_rename)
                    os.rename(mypath+'/'+file,mypath+'/'+file_rename)
                    sheet.append((file,file_rename,'File was renamed',today,'Data not updated','','Notice type not configured'))
                    #shutil.move(mypath+'/'+file_rename, mypathd+'/'+file_rename)
                    dest = get_next_file(file_rename, mypathd)
                    shutil.move(mypath+'/'+file_rename, dest)
                    k=0 
                                    
                elif("Premium Processing Receipt" in notetype):
                    catype= re.search('Case Type(.*) - ', text)
                    if(catype==None):
                        catype= re.search('Case Typo(.*) - ', text)
                        if(catype==None):
                            catype= re.search('lYpc(.*) - ', text)
                            if(catype==None):
                                catype= re.search('IVpc(.*) - ', text)  


                    rdate=re.search('Received Date Priority Date Petitioner(.*)Notice Date',text)
                    if(rdate==None):                      
                        rdate = re.search('Received Date(.*)Petitioner', text)
                        if(rdate==None):
                            rdate = re.search('Received Date(.*)Case Type', text)
                            if(rdate==None):
                                rdate = re.search('Receivedl Date(.*)Case Type', text)
                                if(rdate==None):
                                    rdate = re.search('Priority Date(.*)Notice Date', text)
                    # print(rdate)

                    ndate = re.search('Notice Date(.*)1 of', text)
                    if(ndate==None):
                        ndate = re.search('Norice Date(.*)1 of', text)
                    result = re.search('Notice Date Page 1 of 2(.*)Beneficiary', text)
                    if(result==None):
                        result = re.search('Petitioner (.*)Notice Date', text)
                        if(result==None):
                            result = re.search('Petitioner (.*)Page 1 of 2', text)
                        if(result==None):
                            result=re.search('Pctitioner (.*)Beneficiary', text)
                            if(result==None):
                                result=re.search('Petitioner (.*)Beneficiary', text)
                   
                     

                    # print(catype)
                    # print(rdate)
                    
                    if("1129" not in catype.group(1).split(' - ')[0].strip()):
                        csec = re.search('Section: (.*) ETA Case', text)
                        if(csec==None):
                            csec = re.search('Section: (.*) SOC Code', text)
                        csec1=csec.group(1).split("(")

                        
                    else:
                        
                        csec = re.search('Name DOB COB Class Consulate/POE OCC (.*)', text)
                        if(csec==None):
                            csec = re.search('Nalne (.*) DOB', text)
                            if(csec==None):
                                csec = re.search('Name (.*) California Service Center', text)
                                if(csec==None):
                                    csec = re.search('Name (.*) Class', text)
                                    if(csec==None):
                                        csec = re.search('Name (.*) DOB', text)
                                        if(csec==None):
                                            csec = re.search('Name (.*) OCC', text)
                             

                            if(csec==None):
                                csec1 = ""
                            else:
                                csec1 = csec.group(1).strip()
                        else:
                            csec1 = csec.group(1).strip().split("/")[0]
                            csec1 = re.sub(r'\d', '', csec1)   


                     
                        clreq = re.search('Class requested: (.*) We have', text)                      

                     
                    
                    # print("7890")
                    # print(csec1)
                    match = re.sub(r'\w*\d\w*', '', result.group(1))
                    match=match.replace("//","")

                  
                    
                    name1=match.strip().replace("Beneficinry","Beneficiary").split('Beneficiary')
                    
                    
                    
                    ndate1=ndate.group(1).split("/")
                    ndate3=ndate1[2][:4]
                    
                    if("SIX CONTINENTS" in name1[0].strip()):
                        name1[0]="Six Continents"
                    elif("FEDEX" in name1[0].strip()):
                        name1[0]="FedEx"
                    if(len(name1)==1):
                        bname=re.search('Beneficiary (.*) Page Norice', text)
                        match1 = re.sub(r'\w*\d\w*', '', bname.group(1))
                        match1=match1.replace("//","")

                    
                    
                  

                    if("1140" in catype.group(1)):
                       
                        file_rename=match1.strip()+" - "+name1[0].strip()+" - I140 PP Receipt Notice - "+ndate3.strip()+".pdf"

                        #done DOE, John - FedEx - I140 PP Receipt Notice - 2022

                    elif("1129" not in catype.group(1).split(' - ')[0].strip()):
                        file_rename=manual_replace(catype.group(1).split(' - ')[0].strip(),'I-',0)+" - ("+csec1[1]+"("+csec1[2]+" "+notetype+" - "+rdate3.strip()+"-"+rdate0.strip()+"-"+rdate2.strip()+" - "+name1[1].strip()+" - "+name1[0].strip()+".pdf"
                    else:
                        creq2=clreq.group(1).strip().split("We have")[0].replace("HI","H1")
                        creq2=creq2.replace("LI","L1")
                        creq2=creq2.replace(":","")
                        creq2=creq2.split(" ")[0]
                        creq3=creq2[:1] + '-' + creq2[1:]
                        
                       
                        
                        csec1=csec1.replace("Class Consulate/POE","")
                                     
                        file_rename=csec1.split("DOB")[0]+" - "+name1[0].strip()+" - "+creq2+" PP Receipt Notice - "+ndate3+".pdf"
                        
                        
                        
                    file_rename = fnameclean(file_rename)
                    os.rename(mypath+'/'+file,mypath+'/'+file_rename)
                    sheet.append((file,file_rename,'File was renamed',today,'Data not updated','','Notice type not configured'))
                    #shutil.move(mypath+'/'+file_rename, mypathd+'/'+file_rename)
                    dest = get_next_file(file_rename, mypathd)
                    shutil.move(mypath+'/'+file_rename, dest)
                    k=0                
            except Exception as e:
                #print(e)
                exc_type, exc_obj, exc_tb = sys.exc_info()
                fnamerr = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                #print(exc_type, fnamerr, exc_tb.tb_lineno)
                if(k!=0):
                    k=1
                pass  
                
            try:
                notetype=nottype.group(1).split(' Notice')[0]
                #print(notetype)    
                if(notetype=="Approval"):
                    
                    
                    catype = re.search('case (.*) PETITION ', text)
                    if(catype==None):
                        catype = re.search('Case (.*) PETITION ', text)
                        if(catype==None):
                            catype = re.search('Case (.*) APPLICATION TO ', text)
                            if(catype==None):
                                catype = re.search('Case (.*) APPLICATION FOR ', text)
                    
                    rdate = re.search('to (.*) We have mailed', text)
                    if(rdate==None):
                        rdate= re.search('to (.*) Consulate:', text)
                        if(rdate==None):
                            rdate= re.search('to (.*) The above', text)
                            if(rdate==None):
                                rdate= re.search('from (.*) We have mailed', text)
                                
                    ndate = re.search('Notice Date (.*) Page 1', text)
                    if(ndate==None):
                        ndate = re.search('Notice Date (.*) 1 of', text)
                    fdate = re.search('Valid from (.*) We have mailed', text)
                    
                    #print(catype)
                    #print(rdate)
                    if('1539' in text or 'i539' in text or 'I-539' in text  or '1-539' in text):
                        result = re.search('Beneficiary(.*)Notice Date', text)
                        if(result==None):
                            result = re.search('Beneficiary(.*)Notice Dale', text)
                            if(result==None):
                                result = re.search('Beneficiary(.*)Page 1', text)
                                if(result==None):
                                    result = re.search('Applicant(.*)Notice Date', text)
                        creq= re.search('Class: (.*) Valid from', text)
                        rdate= re.search('to (.*) We have mailed', text)
                        match = re.sub(r'\w*\d\w*', '', result.group(1))
                        match=match.replace("/","")
                        match=match.replace("|","")
                        #print(result)
                        #print(creq)
                        #print(rdate)
                        name1=rem_duplicate(match.strip())
                        name1=name1.replace("Page","")
                        #print(name1)
                        rdate1=rdate.group(1).split("/")
                        
                        rdate3=rdate1[2][:4]
                        rdate2=rdate1[1]
                        rdate0=rdate1[0][-2:]
                        
                        creq2=creq.group(1).strip().replace("HIB","H1B")
                        creq2=creq2.replace("LIB","L1B")
                        creq3=creq2[:1] + '-' + creq2[1:]
                        
                        
                        
                        file_rename=creq3+" "+notetype+" Notice - "+rdate3.strip()+" - "+name1.strip()+".pdf"
                        
                        # done H4 Approval Notice - 2022 - Swan, Elizabeth

                    else:
                        #print("jks2")
                        result = re.search('Petitioner (.*)Notice Date', text)
                        if(result==None):
                            result = re.search('Pctitioner (.*)Notice Date', text)
                            if(result==None):
                                result = re.search('Petitioner (.*)Page', text)
                                if(result==None):
                                    result = re.search('Applicant (.*)Notice Date', text)
                        
                        match = re.sub(r'\w*\d\w*', '', result.group(1))
                        match=match.replace("//","")
                        #print(result)
                        #print(rdate)
                        
                        name1=match.strip().replace("Beneficia ry","Beneficiary").split('Beneficiary')
                        #print(name1)
                        Z=0
                        if(len(name1)<2):
                            name2=name1[0].strip()
                            name3= re.search('Beneficiary (.*)c/o', text)
                            if(name3==None):
                                name3= re.search('Beneficiary (.*)cio', text)
                                if(name3==None):
                                    name3= re.search('Beneficiary (.*)clo', text)
                                    if(name3==None):
                                        name3= re.search('Beneficiary (.*)Notice Type:', text)
                                        if(name3==None):
                                            Z=1
                            if(Z!=1):
                                name4=name3.group(1).strip()
                                name4=re.sub(r'\w*\d\w*', '', name4)
                                name4=name4.replace("//","")
                                name4=name4.replace("Date","")
                                name4=name4.replace("Received","")
                                name4=name4.replace("Priority","")
                                name4=name4.replace("Page","")
                                name4=name4.replace("Of","")
                                name4=name4.replace("Notice","")
                            
                                name2=name2.replace("/","")
                                name2=name2.replace("Date","")
                                name2=name2.replace("Received","")
                                name2=name2.replace("Priority","")
                                name2=name2.replace("Page","")
                                name2=name2.replace("Of","")
                                name2=name2.replace("Notice","")
                                name4=name4.replace(name2,"")
                            else:
                                name2=name2.replace("/","")
                                name2=name2.replace("Date","")
                                name2=name2.replace("Received","")
                                name2=name2.replace("Priority","")
                                name2=name2.replace("Page","")
                                name2=name2.replace("Of","")
                                name2=name2.replace("Notice","")
                        else:
                            name2=name1[0].strip()
                            name4=name1[1].strip()
                            name4=re.sub(r'\w*\d\w*', '', name4)
                            name4=name4.replace("//","")
                            name4=name4.replace("Date","")
                            name4=name4.replace("Received","")
                            name4=name4.replace("Priority","")
                            name4=name4.replace("Page","")
                            name4=name4.replace("Of","")
                            name4=name4.replace("Notice","")
                            name2=name2.replace("/","")
                            name2=name2.replace("Date","")
                            name2=name2.replace("Received","")
                            name2=name2.replace("Priority","")
                            name2=name2.replace("Page","")
                            name2=name2.replace("Of","")
                            name2=name2.replace("Notice","")
                            name4=name4.replace(name2,"")
                            
                        rdate1=rdate.group(1).split("/")
                        fdate1=fdate.group(1).split("to")
                        ndate1=ndate.group(1).split("/")
                        rdate3=rdate1[2][:4]
                        rdate2=rdate1[1]
                        rdate0=rdate1[0][-2:]
                        ndate3=ndate1[2][:4]
                        
                        creq= re.search('Class: (.*) Valid from', text)
                        if(creq==None):
                            creq= re.search('Class: (.*) Valid froin', text)
                            if(creq==None):
                                creq= re.search('Class; (.*) Valid from', text)
                        #print(creq)
                        creq2=creq.group(1).strip().replace("HI","H1")
                        creq2=creq2.replace("LI","L1")
                        creq3=creq2[:1] + '-' + creq2[1:]
                        
                        #print(creq)                    
                        #print(creq3)
                       
                                            
                                        
                        if('1765' in text or 'i765' in text or 'I-765' in text  or '1-765' in text):
                            file_rename=name2.strip()+" - I765 "+creq2+" "+notetype+" Notice - "+rdate3+".pdf" 

                            

                            #  DOE, John - I765 C26 Approval Notice  - 2022

                        elif('1129' in text or 'i129' in text or 'ii29' in text):
                            name4=name4.replace(name2.strip(),"")
                            rdate2            
                            file_rename=name4.strip()+" - "+name2.strip()+" - "+creq3+" - Approval Notice"+" - "+ndate3+".pdf" 

                            print ('name1:',name1)
                            print ('name2:',name2)
                            print ('name3:',name3)

                         
                            # DOE, John - FedEx - H1B Approval Notice  - 2022
                                                        
                        else:

                         
                            name4=name4.replace(name2.strip(),"")
                            file_rename=creq3+" - "+notetype+" Notice - "+name4.strip()+" - "+name2.strip()+" - "+rdate3+".pdf" 
                      

                    file_rename = fnameclean(file_rename)
                    os.rename(mypath+'/'+file,mypath+'/'+file_rename)
                    sheet.append((file,file_rename,'File was renamed',today,'Data not updated','','Notice type not configured'))
                    #shutil.move(mypath+'/'+file_rename, mypathd+'/'+file_rename)
                    dest = get_next_file(file_rename, mypathd)
                    shutil.move(mypath+'/'+file_rename, dest)
                    
                    
                    k=0                
            except Exception as e:
                if(k!=0):
                    k=1
                pass                      
            try:
                
                notetype1=nottype1.group(1)
                
                #print(notetype1)
                if("ASC Appointment Notice" in notetype1):
                    notetype1="ASC Appointment Notice" 
                if(notetype1.strip()=="ASC Appointment Notice" or notetype1.strip()=="Biometric Notification"):
                    catype= re.search('NOTICE DATE (.*) APPLICATION TO ', text)
                    if(catype==None):
                        catype= re.search('CASE TYPE (.*) NOTICE DATE ', text)
                        if(catype==None):
                            catype= re.search('CASE TYPE (.*) APPLICATION TO ', text)
                            if(catype==None):
                                catype= re.search('CASE TYPE (.*) APPLICATION ', text)
                    #print(catype)
                    rdate = re.search('DATE AND TIME OF APPOINTMENT (.*) YOU GO', text)
                    if(rdate==None):
                        rdate = re.search('DATE AND TIME OF APPOINTMENT (.*) WHEN YOU APPEAR', text)
                    #print(rdate)
                    result = re.search('PAGE 1 of 2 (.*) c/o ', text)
                    if(result==None):
                        result = re.search('PAGE. 1 of 2 (.*) c/o ', text)
                        if(result==None):
                            result = re.search('PAGE CSC 1 of 2 (.*) c/o ', text)
                            if(result==None):
                                result = re.search('PAGE NSC 1 of 2 (.*) c/o ', text)
                                if(result==None):
                                    result = re.search('PAGE TSC 1 of 2 (.*) c/o ', text)
                                    if(result==None):
                                        result = re.search('PAGE TSC 1 of 2 (.*) c/o ', text)
                                        if(result==None):
                                            result = re.search('SC 1 of 2 (.*) c/o ', text)
                                            if(result==None):
                                                result = re.search('SC 1 of 3 (.*) c/o ', text)
                    
                    #print(result)
                    catype1=catype.group(1)
                    catype1=catype1.replace('; ', "-")
                    catype1=catype1.replace(' APPLICATION FOR NATURALIZATION', "")
                    catype1=catype1.replace('NOTICE DATE ', "")
                    catype2=catype1.split("-")[0].replace("NOTICE DATE ","")
                    rdate2=rdate.group(1).split("/")
                    match = re.sub(r'\w*\d\w*', '', result.group(1))
                    match=match.replace("!","")
                    name1=rem_duplicate(match.strip())
                    catype2=catype2.replace("APPLICATION TO EXTEND/CHANGE NONIMMIGRANT STATUS","")
                    catype2=catype2.replace("|","")
                    catype2 = re.sub(r'(\d{2})/(\d{2})/(\d{4})', '', catype2).strip()
                    
                    #print(catype1)
                    #print(name1)
                    #print(rdate2)
                    
                    
                    rdate3=rdate2[2][:4]
                    rdate1=rdate2[1]
                    rdate0=rdate2[0][-2:]
                    
                    if(catype1.strip()!="N400"):
                        if('1539' in catype1):
                            file_rename="I-539 - "+notetype1.strip()+" ("+rdate3.strip()+"-"+rdate1.strip()+"-"+rdate0.strip()+" ) - "+name1.strip()+".pdf"
                        else:
                            file_rename=manual_replace(catype2.strip(),'I-',0)+" - "+notetype1.strip()+" ("+rdate3.strip()+"-"+rdate1.strip()+"-"+rdate0.strip()+" ) - "+name1.strip()+".pdf"
                    else:
                        file_rename=manual_replace(catype.strip(),'N-',0)+" - "+notetype1.strip()+" ("+rdate3.strip()+"-"+rdate1.strip()+"-"+rdate0.strip()+") - "+name1.strip()+".pdf"                
                    #print(file_rename)
                    file_rename = fnameclean(file_rename)
                    os.rename(mypath+'/'+file,mypath+'/'+file_rename)
                    sheet.append((file,file_rename,'File was renamed',today,'Data not updated','','Notice type not configured'))
                    #shutil.move(mypath+'/'+file_rename, mypathd+'/'+file_rename)
                    dest = get_next_file(file_rename, mypathd)
                    shutil.move(mypath+'/'+file_rename, dest)
                    k=0                
            except Exception as e:
                #print(e)
                if(k!=0):
                    k=1
                pass    

            try:
                notetype2=nottype2.group(1).split(' Case Type')[0]
                
                if(notetype2.strip()=="INTERVIEW"):
                    catype= re.search('Case Type FORM (.*), APPLICATION', text)
                    rdate = re.search('ON: (.*) AT:', text)
                    result = re.search('A Number (.*) c/o ', text)
                    if(result==None):
                        result= re.search('Receipt Number (.*) c/o ', text)
                        match = re.sub(r'\w*\d\w*', '', result.group(1))
                    else:
                        match = result.group(1)
                    
            
                    
                    
                    catype1=catype.group(1)
                    rdate2=rdate.group(1).split(",")
                    
                    rdate1=rdate2[1].strip().split(" ")
                    
                    rdatemon=mon_conversion(rdate1[0].strip())
                    if(catype1.strip()!="N400"):
                        file_rename=manual_replace(catype1.strip(),'I-',0)+" - "+notetype2.strip()+" ("+rdate2[2].strip()+"-"+str(rdatemon)+"-"+rdate1[1].strip()+" ) - "+match.strip()+".pdf"
                    else:
                        file_rename=manual_replace(catype1.strip(),'N-',0)+" - "+notetype2.strip()+" ("+rdate2[2].strip()+"-"+str(rdatemon)+"-"+rdate1[1].strip()+") - "+match.strip()+".pdf"
                    
                    
                                    
                    file_rename = fnameclean(file_rename)
                    os.rename(mypath+'/'+file,mypath+'/'+file_rename)
                    sheet.append((file,file_rename,'File was renamed',today,'Data not updated','','Notice type not configured'))
                    #shutil.move(mypath+'/'+file_rename, mypathd+'/'+file_rename)
                    dest = get_next_file(file_rename, mypathd)
                    shutil.move(mypath+'/'+file_rename, dest)
                    k=0                
            except Exception as e:
                if(k!=0):
                    k=1
                pass  

            try:
                notetype3=nottype.group(1).split(' TYPE')[1]
                #print(notetype3)
                if(notetype3.strip()=="Recei t CASE" or notetype3.strip()=="Reit CASE" or notetype3==" it CASE" or notetype3==" Receipt CASE"):
                
                    catype= re.search('CASE TYPE (.*) RECEIPT NUMBER', text)
                    if(catype==None):
                        catype= re.search('CASE TYPE (.*) RECEiPT NUMBER', text)
                    
                         
                    if(catype.group(1).strip().split(" ")[0]=="1-539"):
                        rdate = re.search('RECEIVED DATE(.*)NOTICE DATE', text)
                        result = re.search('PAYMENT INFORMATION:(.*) C/O', text)
                        rdate1=rdate.group(1).strip().split(" ")
                        rdatemon=mon_conversion(rdate1[0].strip())
                        try:
                            rdatemon=mon_conversion(rdate1[0].strip())
                            file_rename="I-539 - "+"Receipt"+" - "+rdate1[2].strip()+"-"+str(rdatemon)+"-"+rdate1[1].strip()+" - "+result.group(1).strip()+".pdf"
                        except:                        
                            file_rename="I-539 - "+"Receipt"+" - "+rdate.group(1).strip()+" - "+result.group(1).strip()+".pdf"
                            pass                    
                        
                    elif(catype.group(1).strip().split(" ")[0]=="1-130" or catype.group(1).strip().split(" ")[0]=="1-130."):
                        rdate = re.search('RECEIVED DATE (.*) PREFERENCE CLASSIFICATION', text)
                        result = re.search('ame (.*) Date of', text)
                        rdate1=rdate.group(1).strip().split(" ")
                        try:
                            rdatemon=mon_conversion(rdate1[0].strip())
                            file_rename="I-130 - "+"Receipt"+" - "+rdate1[2].strip()+"-"+str(rdatemon)+"-"+rdate1[1].strip()+" - "+result.group(1).strip()+".pdf"
                        except:                        
                            file_rename="I-130 - "+"Receipt"+" - "+rdate.group(1).strip()+" - "+result.group(1).strip()+".pdf"
                            pass
                    else:
                        rdate = re.search('RECEIVED DATE(.*) PAGE 1', text)
                        
                        result = re.search('PAYMENT INFORMATION:(.*) C/O', text)
                        #print(rdate)
                        #print(result)
                        if(result==None):
                            result = re.search('PAYMENT INFORMATION:(.*) CIO', text)
                        if(result.group(1).strip()!=""):
                            name=result.group(1).strip()
                        else:
                            result = re.search('RECEIPT NUMBER(.*) RECEIVED DATE', text)
                            match = re.sub(r'\w*\d\w*', '', result.group(1))
                            name=match.strip().replace("PRIORITY DATE","")
                            
                        rdate1=rdate.group(1).strip().split(" ")
                        
                        creq = re.search('Eligibility Category: (.*) We have', text)
                        name = re.sub(r'\w*\d\w*', '', name)
                        name=name.replace("$.","")
                        try:
                            rdatemon=mon_conversion(rdate1[1].strip())
                            file_rename=manual_replace(catype.group(1).strip().split(" ")[0],'I',0)+" ("+creq.group(1)+") - "+"Receipt"+" - "+rdate1[3].strip()+"-"+str(rdatemon)+"-"+rdate1[2].strip().replace(",","")+" - "+name+".pdf"
                        except:                        
                            file_rename=manual_replace(catype.group(1).strip().split(" ")[0],'I',0)+" ("+creq.group(1)+") - "+"Receipt"+" - "+rdate.group(1).strip().replace("PREFERENCE CLASSIFICATION r","")+" - "+name+".pdf"
                            pass                    
                        
                    #print(file_rename)
                                    
                    file_rename = fnameclean(file_rename)
                    os.rename(mypath+'/'+file,mypath+'/'+file_rename)
                    sheet.append((file,file_rename,'File was renamed',today,'Data not updated','','Notice type not configured'))
                    #shutil.move(mypath+'/'+file_rename, mypathd+'/'+file_rename)
                    dest = get_next_file(file_rename, mypathd)
                    shutil.move(mypath+'/'+file_rename, dest)
                    k=0                
            except Exception as e:
                #print(e)
                if(k!=0):
                    k=1
                pass       
            
            try:
                notetype4=nottype3.group(1).strip().split(" ")[0]
                
                if(notetype4.strip()=="Naturalization"):                
                    rdate = re.search('Date and Time: (.*) at ', text)
                    result = re.search('A-Number (.*) c/o', text)
                    
                    rdate2=rdate.group(1).split("at")[0]
                    
                    match = re.sub(r'\w*\d\w*', '', result.group(1))
                    name1=rem_duplicate(match.strip())
                    
                    name2=name1.replace("//","")
                    name2=name2.replace("Date","")
                    
                    rdate1=rdate2.split(" ")
                    rdatemon=mon_conversion(rdate1[1].strip())
                    
                    
                    file_rename="Notice of Naturalization Oath Ceremony"+" ("+rdate1[3].strip()+"-"+str(rdatemon)+"-"+rdate1[2].strip().replace(",","")+" ) - "+name2.strip()+".pdf"
                    
                                                
                    file_rename = fnameclean(file_rename)
                    os.rename(mypath+'/'+file,mypath+'/'+file_rename)
                    sheet.append((file,file_rename,'File was renamed',today,'Data not updated','','Notice type not configured'))
                    #shutil.move(mypath+'/'+file_rename, mypathd+'/'+file_rename)
                    dest = get_next_file(file_rename, mypathd)
                    shutil.move(mypath+'/'+file_rename, dest)
                    k=0                
            except Exception as e:
                if(k!=0):
                    k=1
                pass

            try:
                notetype5=nottype4.group(1)
                
                if(notetype5.strip()=="interview"): 
                    catype= re.search('Case Type (.*) - ', text)            
                    rdate = re.search('On (.*) At', text)
                    result = re.search('Please come to:(.*) c/o', text)
                    if(result==None):
                        result = re.search('Page lof2 (.*) c/o', text)
                        if(result==None):
                            result = re.search('Page 1 of 2 (.*) c/o', text)
                    
                    
                    match = re.sub(r'\w*\d\w*', '', result.group(1))
                    name1=rem_duplicate(match.strip())
                    
                    rdate1=rdate.group(1).strip().split(" ")
                    rdatemon=mon_conversion(rdate1[2].strip())
                    
                    if(catype.group(1).split(' - ')[0].strip()!="N400"):
                        file_rename=manual_replace(catype.group(1).split(' - ')[0].strip(),'I-',0)+" - Interview Notice"+" ("+rdate1[4].strip()+"-"+str(rdatemon)+"-"+rdate1[3].strip().replace(",","")+" ) - "+name1+".pdf"
                    else:
                        file_rename=manual_replace(catype.group(1).split(' - ')[0].strip(),'N-',0)+" - Interview Notice"+" ("+rdate1[4].strip()+"-"+str(rdatemon)+"-"+rdate1[3].strip().replace(",","")+" ) - "+name1+".pdf"
                    
                    #print(file_rename)
                                    
                    file_rename = fnameclean(file_rename)
                    os.rename(mypath+'/'+file,mypath+'/'+file_rename)
                    sheet.append((file,file_rename,'File was renamed',today,'Data not updated','','Notice type not configured'))
                    #shutil.move(mypath+'/'+file_rename, mypathd+'/'+file_rename)
                    dest = get_next_file(file_rename, mypathd)
                    shutil.move(mypath+'/'+file_rename, dest)
                    k=0                
            except Exception as e:
                if(k!=0):
                    k=1
                pass            
                
            try:
                notetype6=nottype5.group(1)
                #print(nottype5.group(1))
                if('Receipt' in notetype6.strip()): 
                    catype= re.search('CASE TYPE (.*) Naturalization', text)
                    if(catype==None):
                        catype= re.search('CASE TYPE (.*), Application', text)
                        if(catype==None):
                            catype= re.search('CASE TYPE (.*). Application', text)                    
                    rdate = re.search('RECEIVED DATE (.*) APPLICANT', text)
                    if(rdate==None):
                        rdate = re.search('RECEIVED DATE (.*) USCIS ALIEN', text)
                        if(rdate==None):
                            rdate = re.search('RECEIVED DATE (.*) PRIORITY DATE', text)
                            if(rdate==None):
                                rdate = re.search('RECEIVED DATE (.*) PAGE 1', text)
                                if(rdate==None):
                                    rdate = re.search('RECEIVED DATE (.*) PAYMENT INFORMATION:', text)
                    result = re.search('Single Application Fee: (.*) CIO', text)
                    if(result==None):
                        result = re.search('Single Application Fee: (.*) C/O', text)
                        if(result==None):
                            result = re.search('PAYMENT INFORMATION: (.*) CIO', text)
                            if(result==None):
                                result = re.search('PAYMENT INFORMATION: (.*) C/O', text)
                    #print(catype)
                    #print(rdate)
                    #print(result)
                    match = re.sub(r'\w*\d\w*', '', result.group(1))
                    name1=rem_duplicate(match.strip())
                    name1=name1.replace("$","")
                    #print(name1)
                    
                    rdate1=rdate.group(1).replace("NOTICE DATE","").strip().split(" ")
                    #print(rdate1)
                    if(len(rdate1)>3):
                        rdatemon=mon_conversion(rdate1[1].strip())
                    else:
                        rdatemon=mon_conversion(rdate1[0].strip())
                    #print(rdatemon)
                    #print(catype)
                    if(catype.group(1).split(' ')[0].strip()!="N-400"):
                        #print("loop1")
                        #print(catype)
                        if('I-131' in catype.group(1)):
                            catype1='I-131'
                        elif('I-485' in catype.group(1)):
                            catype1='I-485'
                        else:
                            catype1=catype.group(1).split(' ')[0].strip()
                        #print(catype)   
                        if(len(rdate1)>3):
                            file_rename=manual_replace(catype1,'I',0)+" - Receipt"+" - "+rdate1[3].strip()+"-"+str(rdatemon)+"-"+rdate1[2].strip().replace(",","")+" - "+name1+".pdf"
                        else:
                            file_rename=manual_replace(catype1,'I',0)+" - Receipt"+" - "+rdate1[2].strip()+"-"+str(rdatemon)+"-"+rdate1[1].strip().replace(",","")+" - "+name1+".pdf"
                    else:
                        #print("loop2")
                        file_rename=manual_replace(catype.group(1).split(' ')[0].strip().replace(",",""),'N',0)+" - Receipt"+" - "+rdate1[3].strip()+"-"+str(rdatemon)+"-"+rdate1[2].strip().replace(",","")+" - "+name1+".pdf"
                    
                    #print(file_rename)
                                    
                    file_rename = fnameclean(file_rename)
                    os.rename(mypath+'/'+file,mypath+'/'+file_rename)
                    sheet.append((file,file_rename,'File was renamed',today,'Data not updated','','Notice type not configured'))
                    #shutil.move(mypath+'/'+file_rename, mypathd+'/'+file_rename)
                    dest = get_next_file(file_rename, mypathd)
                    shutil.move(mypath+'/'+file_rename, dest)
                    k=0                
            except Exception as e:
                if(k!=0):
                    k=1
                pass    

            try:
                notetype7=nottype6.group(1)
                
                if(nottype6!=None): 
                    creq= re.search('PREFERENCE CLASSIFICATION (.*) NOTICE DATE', text)            
                    rdate = re.search('RECEIVED DATE (.*) PREFERENCE CLASSIFICATION', text)
                    result = re.search('PAYMENT INFORMATION: (.*) C/O', text)
                    if(result==None):
                        result = re.search('PREFERENCE CLASSIFICATION (.*) C/O', text)
                    name2 = re.search('ame (.*) Date of', text)
                    
                    #print(creq.group(1).split(' ')[1].strip())
                    #print(rdate.group(1))
                    #print(result.group(1))
                    #print(name2)
                    match = re.sub(r'\w*\d\w*', '', result.group(1))
                    name1=rem_duplicate(match.strip())
                    
                    rdate1=rdate.group(1).strip().split(" ")
                    rdatemon=mon_conversion(rdate1[0].strip())
                    
                    
                    file_rename="I-140 - E"+creq.group(1).split(' ')[1].strip()+" - Receipt"+" - "+rdate1[2].strip()+"-"+str(rdatemon)+"-"+rdate1[1].strip().replace(",","")+" - "+name2.group(1).strip()+" - "+name1+".pdf"
                    
                    #print(file_rename)
                                    
                    file_rename = fnameclean(file_rename)
                    os.rename(mypath+'/'+file,mypath+'/'+file_rename)
                    sheet.append((file,file_rename,'File was renamed',today,'Data not updated','','Notice type not configured'))
                    #shutil.move(mypath+'/'+file_rename, mypathd+'/'+file_rename)
                    dest = get_next_file(file_rename, mypathd)
                    shutil.move(mypath+'/'+file_rename, dest)
                    k=0                
            except Exception as e:
                if(k!=0):
                    k=1
                pass 
            #print('cattype '+str(catype))
            
            if(k==1):
                
                sheet.append((file,'','File was not renamed',today))
                #shutil.move(mypath+'/'+file, mypathe+'/'+file) 
                dest = get_next_file(file, mypathe)
                #shutil.move(mypath+'/'+file, dest)            

            sheet.column_dimensions["A"].width = 50.0
            sheet.column_dimensions["B"].width = 60.0
            sheet.column_dimensions["C"].width = 40.0
            sheet.column_dimensions["D"].width = 10.0
            sheet.column_dimensions["E"].width = 40.0
            sheet.column_dimensions["F"].width = 40.0
            sheet.column_dimensions["G"].width = 40.0
            sheet.column_dimensions["H"].width = 40.0
            sheet.column_dimensions["I"].width = 40.0
            sheet.column_dimensions["J"].width = 40.0
            sheet.column_dimensions["K"].width = 40.0
            book2.save(fdname +'nrtlog.xlsx')
                
        if os.path.exists(fdname + 'page_1.png'):
            os.remove(fdname + 'page_1.png')
            #os.remove(file_name_arr)
        
    except Exception as e:
        error_log(e, 'Failed', 'Something Went Wrong. Process Stopped','')
        pass 
        
        
    
   
            
        
def request_ocr1(subscription_key, image_filenames):
    try:
        # Read the image into a byte array
        image_data = open(image_filenames, "rb").read()
        # Set Content-Type to octet-stream


        headers = {'Ocp-Apim-Subscription-Key': subscription_key, 'Content-Type': 'application/octet-stream'}
        params = {'language': 'en', 'detectOrientation': 'true'}
        # put the byte array into your post request
        response = requests.post(ocr_url, headers=headers, params=params, data = image_data)
        response.raise_for_status()
        analysis = response.json()
        return analysis 
    except Exception as e:
        error_log(e, 'Failed', 'Something Went Wrong. Process Stopped','')
        pass

def fnameclean(file_name):
    file_name=file_name.replace(":","")
    file_name=file_name.replace("*","")
    file_name=file_name.replace("?","")
    file_name=file_name.replace("\\","")
    file_name=file_name.replace("/","")
    file_name=file_name.replace("<","")
    file_name=file_name.replace(">","")
    file_name=file_name.replace("|","")
    file_name=file_name.replace('"',"")
    file_name=file_name.replace("GLOBAL IMMIGRATION PARTNERS INC Notice Type Approval Notice","")
    file_name=file_name.replace("GLOBAL IMMIGRATION PARTNERS INC. Notice Type Approval Notice","")
    file_name=file_name.replace("AGOURA HILLS CA 91301","")
    file_name=file_name.replace(" 30300 AGOURA RD STE B100","")
    file_name=file_name.replace(" 30300 AGOURA ROAD STE B100","")
    file_name=file_name.replace("GLOBAL IMMIGRATION PARTNERS INC","")
    file_name=file_name.replace(" AGOURA HLLS CA 930","")
    file_name=file_name.replace("Type Approval","")
    file_name=file_name.replace("ISCIS ALIEN NUMBER","")
    file_name=file_name.replace("Class","")
    file_name=file_name.replace("Date","")
    file_name=file_name.replace("Received","")
    file_name=file_name.replace("Priority","")
    file_name=file_name.replace("Page","")
    file_name=file_name.replace("Of","")
    file_name=file_name.replace("of","")
    file_name=file_name.replace("Norice","")
    
    file_name=file_name.replace("Application","")
    file_name=file_name.replace("Petition","")
    file_name=file_name.replace("Fee","")
    file_name=re.sub('\s+',' ',file_name)
    
    
    return file_name
    
def get_next_file(file_name, dest_dir):
    try:
        file_name=file_name.replace(":","")
        file_name=file_name.replace("*","")
        file_name=file_name.replace("?","")
        file_name=file_name.replace("\\","")
        file_name=file_name.replace("/","")
        file_name=file_name.replace("<","")
        file_name=file_name.replace(">","")
        file_name=file_name.replace("|","")
        file_name=file_name.replace('"',"")
        file_name=file_name.replace("Date","")
        file_name=file_name.replace("Received","")
        file_name=file_name.replace("Priority","")
        file_name=file_name.replace("Page","")
        file_name=file_name.replace("Of","")
        
        dest = os.path.join(dest_dir, file_name)
        num = 0

        while os.path.exists(dest):
            num += 1

            period = file_name.rfind('.')
            if period == -1:
                period = len(file_name)

            new_file = f'{file_name[:period]}({num}){file_name[period:]}'

            dest = os.path.join(dest_dir, new_file)
        
        
        
        return dest
    except Exception as e:
        error_log(e, 'Failed', 'Something Went Wrong. Process Stopped','')
        pass


def rem_duplicate(rtext):
    try:
        s = rtext.replace('.', "")
        l = s.split() 
        k = [] 
        for i in l: 
    
            # If condition is used to store unique string  
            # in another list 'k'  
            if (s.count(i)>1 and (i not in k)or s.count(i)==1): 
                k.append(i) 
        result=' '.join(k)
        result=result.split('Beneficiary')[0]
        return result 
    except Exception as e:
        #error_log(e, 'Failed', 'Something Went Wrong. Process Stopped')
        pass

def mon_conversion(monthWord):
    try:
        newWord = monthWord [0].upper() + monthWord [1:3].lower() 
        result=strptime(newWord,'%b').tm_mon
        
        if(result<10):
            result="0"+str(result)
        
        return result  
    except Exception as e:
        #error_log(e, 'Failed', 'Something Went Wrong. Process Stopped')
        pass

def manual_replace(s, char, index):
    try:
        return s[:index] + char + s[index +1:]    
    except Exception as e:
        #error_log(e, 'Failed', 'Something Went Wrong. Process Stopped')
        pass
 
def error_log(e, status='Failed', msg='',mstat=''):
    #print('error_log', e, msg)
    try: 
        date = datetime.today().strftime('%m/%d/%Y')
        end_time  = datetime.now().strftime("%H:%M") 
        if not os.path.exists(fdname + 'error_log.xlsx'):
            book = Workbook()
            ws = book.active
            ws.cell(1,1).value = 'Date'
            ws.cell(1,2).value = 'Start Time'
            ws.cell(1,3).value = 'End Time'
            ws.cell(1,4).value = 'Run Status'
            ws.cell(1,5).value = 'Error Message'
            ws.cell(1,6).value = 'Error Message Comments'
            ws.cell(1,7).value = 'Files Processed'
            
            ws.column_dimensions["A"].width = 20.0
            ws.column_dimensions["B"].width = 20.0
            ws.column_dimensions["C"].width = 20.0
            ws.column_dimensions["D"].width = 20.0
            ws.column_dimensions["E"].width = 50.0
            ws.column_dimensions["F"].width = 50.0
            ws.column_dimensions["G"].width = 200.0
            
            ws['A1'].font = Font(bold=True)
            ws['B1'].font = Font(bold=True)
            ws['C1'].font = Font(bold=True)
            ws['D1'].font = Font(bold=True)
            ws['E1'].font = Font(bold=True)
            ws['F1'].font = Font(bold=True)
            ws['G1'].font = Font(bold=True)
            
            book.save(fdname +'error_log.xlsx')

        book2 = load_workbook(fdname +'error_log.xlsx')
        sheet = book2.active                
        sheet.append((date, start_time, end_time, status, str(e), msg, ";".join(processed_files)))
        sheet.column_dimensions["A"].width = 20.0
        sheet.column_dimensions["B"].width = 20.0
        sheet.column_dimensions["C"].width = 20.0
        sheet.column_dimensions["D"].width = 20.0
        sheet.column_dimensions["E"].width = 50.0
        sheet.column_dimensions["F"].width = 50.0
        sheet.column_dimensions["G"].width = 200.0
        
        book2.save(fdname +'error_log.xlsx')
        if(mstat!=""):
            send_email(receiver_email, status)
        exit()
        pass
        
    except:
        pass
        
    
        
        
        
        
def send_email(email, msg):
    #print('mail', msg)
    
    try:
        if msg =='Failed':
            SUBJECT = 'Process Failed. Please refer to the Log File/Command Prompt Window for details. '
            TEXT = 'Your Process Failed. Please refer to the Log File/Command Prompt Window for details.'
        else:
            SUBJECT = 'Process Completed Successfully'
            TEXT = 'Your Process Completed Successfully.'
            
        
        message = 'Subject: {}\n\n{}'.format(SUBJECT, TEXT)
        context = ssl.create_default_context()
        with smtplib.SMTP_SSL(smtp_server, port, context=context) as server:
            server.login(sender_email, password)
            server.sendmail(sender_email, receiver_email.split(','), message)
        #return False
        exit()
    except Exception as e:
        print('Mail Error', e)
        if msg =='fail':
            error_log(e, 'Failed', 'Mail Error','')
        else:
            error_log(e, 'Successful', 'Mail not Sent.','')
        exit()
    
rename()




   

            
   
